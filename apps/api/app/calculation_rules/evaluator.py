"""Safe evaluation of validated `calc-ir-v1` nodes."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import math
import re
from typing import Any, Mapping, Sequence

from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.utils.datetime import (
    CALENDAR_MAC_1904,
    CALENDAR_WINDOWS_1900,
    from_excel,
    to_excel,
)

from .compiler import CalculationExpressionValidator
from .function_registry import FUNCTION_REGISTRY, FunctionDefinition
from .graph import CalculationGraphPlan
from .types import (
    CalculationRuleExtractionConfiguration,
    FormulaCompilation,
    WorkbookCatalog,
    WorkbookCellFact,
    WorkbookCellRef,
    WorkbookFormulaCell,
)


_ERROR_CODES = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A"}


@dataclass(frozen=True)
class ScalarValue:
    kind: str
    value: Any = None
    error_code: str | None = None
    iso_evidence: str | None = None

    @classmethod
    def number(cls, value: int | float) -> "ScalarValue":
        number = float(value)
        if not math.isfinite(number):
            return cls.error("#NUM!")
        return cls("number", number)

    @classmethod
    def boolean(cls, value: bool) -> "ScalarValue":
        return cls("boolean", bool(value))

    @classmethod
    def text(cls, value: str) -> "ScalarValue":
        return cls("text", str(value))

    @classmethod
    def blank(cls) -> "ScalarValue":
        return cls("blank")

    @classmethod
    def date_serial(cls, value: float, iso_evidence: str | None = None) -> "ScalarValue":
        if not math.isfinite(value):
            return cls.error("#NUM!")
        return cls("date_serial", float(value), iso_evidence=iso_evidence)

    @classmethod
    def error(cls, code: str) -> "ScalarValue":
        if code not in _ERROR_CODES:
            raise ValueError(f"Unknown Excel error code: {code}")
        return cls("error", error_code=code)

    @property
    def number_value(self) -> float:
        if self.kind not in {"number", "date_serial"}:
            raise TypeError("Scalar value is not numeric")
        return float(self.value)

    def to_json(self) -> dict[str, Any]:
        if self.kind == "error":
            return {"value_type": "error", "error_code": self.error_code}
        if self.kind in {"number", "date_serial"}:
            payload: dict[str, Any] = {
                "value_type": self.kind,
                "value": format(float(self.value), ".17g"),
            }
            if self.iso_evidence is not None:
                payload["iso_evidence"] = self.iso_evidence
            return payload
        return {"value_type": self.kind, "value": self.value}

    @classmethod
    def from_json(cls, payload: dict[str, Any] | None) -> "ScalarValue | None":
        if payload is None:
            return None
        kind = payload.get("value_type")
        if kind == "error":
            return cls.error(payload["error_code"])
        if kind == "number":
            return cls.number(float(payload["value"]))
        if kind == "date_serial":
            return cls.date_serial(
                float(payload["value"]),
                payload.get("iso_evidence"),
            )
        if kind == "boolean":
            return cls.boolean(bool(payload["value"]))
        if kind == "text":
            return cls.text(str(payload["value"]))
        if kind == "blank":
            return cls.blank()
        raise ValueError(f"Unknown persisted scalar value type: {kind}")


@dataclass(frozen=True)
class FormulaExecution:
    status: str
    value: ScalarValue | None
    error_code: str | None
    direct_input_trace: tuple[dict[str, Any], ...]
    warnings: tuple[str, ...] = ()


@dataclass
class CalculationExecutionContext:
    catalog: WorkbookCatalog
    formula_cell: WorkbookFormulaCell
    compilation: FormulaCompilation
    calculated_values: Mapping[WorkbookCellRef, ScalarValue]
    configuration: CalculationRuleExtractionConfiguration
    input_values: Mapping[WorkbookCellRef, ScalarValue] = field(default_factory=dict)


@dataclass(frozen=True)
class _RangeValue:
    values: tuple[ScalarValue, ...]
    rows: int
    columns: int


class SafeCalculationEvaluator:
    def __init__(
        self,
        *,
        function_registry: Mapping[str, FunctionDefinition] | None = None,
    ):
        self._function_registry = function_registry or FUNCTION_REGISTRY
        self._validator = CalculationExpressionValidator(
            function_registry=self._function_registry
        )

    def evaluate(
        self,
        expression: dict[str, Any],
        context: CalculationExecutionContext,
    ) -> FormulaExecution:
        self._validator.validate(
            expression,
            context.formula_cell,
            context.compilation.references,
            context.configuration,
        )
        trace: list[dict[str, Any]] = []
        value = self._evaluate_node(expression["root"], context, trace)
        if isinstance(value, _RangeValue):
            value = ScalarValue.error("#VALUE!")
        status = "execution_error" if value.kind == "error" else "executed"
        return FormulaExecution(
            status=status,
            value=value,
            error_code=value.error_code if value.kind == "error" else None,
            direct_input_trace=tuple(trace),
        )

    def execute(
        self,
        plan: CalculationGraphPlan,
        catalog: WorkbookCatalog,
        compilations: Sequence[FormulaCompilation],
        configuration: CalculationRuleExtractionConfiguration | None = None,
        *,
        evaluation_cells: Sequence[WorkbookCellRef] | None = None,
        initial_calculated_values: Mapping[WorkbookCellRef, ScalarValue] | None = None,
        input_values: Mapping[WorkbookCellRef, ScalarValue] | None = None,
    ) -> dict[WorkbookCellRef, FormulaExecution]:
        configuration = configuration or CalculationRuleExtractionConfiguration()
        formula_by_ref = catalog.formula_by_ref()
        compilation_by_formula_id = {
            compilation.formula_cell_id: compilation for compilation in compilations
        }
        results: dict[WorkbookCellRef, FormulaExecution] = {}
        calculated: dict[WorkbookCellRef, ScalarValue] = dict(
            initial_calculated_values or {}
        )
        selected = set(evaluation_cells) if evaluation_cells is not None else None
        for reference, status in plan.status_by_cell.items():
            if status == "ready":
                continue
            results[reference] = FormulaExecution(
                status=status,
                value=None,
                error_code=None,
                direct_input_trace=(),
                warnings=(status,),
            )
        for reference in plan.evaluation_order:
            if selected is not None and reference not in selected:
                continue
            formula_cell = formula_by_ref[reference]
            compilation = compilation_by_formula_id[formula_cell.id]
            if compilation.ir_json is None:
                raise ValueError("Executable graph references a formula without validated IR")
            context = CalculationExecutionContext(
                catalog=catalog,
                formula_cell=formula_cell,
                compilation=compilation,
                calculated_values=calculated,
                configuration=configuration,
                input_values=input_values or {},
            )
            execution = self.evaluate(compilation.ir_json, context)
            results[reference] = execution
            if execution.value is not None:
                calculated[reference] = execution.value
        return results

    def _evaluate_node(
        self,
        node: dict[str, Any],
        context: CalculationExecutionContext,
        trace: list[dict[str, Any]],
    ) -> ScalarValue | _RangeValue:
        node_type = node["node_type"]
        if node_type == "literal":
            literal_type = node["literal_type"]
            if literal_type == "number":
                return ScalarValue.number(float(node["value"]))
            if literal_type == "boolean":
                return ScalarValue.boolean(node["value"])
            if literal_type == "text":
                return ScalarValue.text(node["value"])
            return ScalarValue.blank()
        if node_type == "error_value":
            return ScalarValue.error(node["error_code"])
        if node_type == "cell_reference":
            return self._cell_value(node["cell"], context, trace)
        if node_type == "range_reference":
            return self._range_value(node, context, trace)
        if node_type == "unary_operation":
            operand = self._evaluate_node(node["operand"], context, trace)
            number = _coerce_numeric(operand)
            if isinstance(number, ScalarValue):
                return number
            if node["operator"] == "negative":
                number = -number
            elif node["operator"] == "percent":
                number /= 100.0
            return _finite_number(number)
        if node_type == "binary_operation":
            left_value = self._evaluate_node(node["left"], context, trace)
            right_value = self._evaluate_node(node["right"], context, trace)
            if isinstance(left_value, _RangeValue) or isinstance(right_value, _RangeValue):
                return _range_binary_operation(
                    left_value,
                    right_value,
                    node["operator"],
                )
            left = _coerce_numeric(left_value)
            if isinstance(left, ScalarValue):
                return left
            right = _coerce_numeric(right_value)
            if isinstance(right, ScalarValue):
                return right
            operator = node["operator"]
            try:
                if operator == "add":
                    result = left + right
                elif operator == "subtract":
                    result = left - right
                elif operator == "multiply":
                    result = left * right
                elif operator == "divide":
                    if right == 0:
                        return ScalarValue.error("#DIV/0!")
                    result = left / right
                else:
                    result = left**right
            except (ArithmeticError, OverflowError, ValueError):
                return ScalarValue.error("#NUM!")
            if isinstance(result, complex):
                return ScalarValue.error("#NUM!")
            return _finite_number(result)
        if node_type == "comparison":
            left = self._evaluate_node(node["left"], context, trace)
            if isinstance(left, _RangeValue):
                return ScalarValue.error("#VALUE!")
            if left.kind == "error":
                return left
            right = self._evaluate_node(node["right"], context, trace)
            if isinstance(right, _RangeValue):
                return ScalarValue.error("#VALUE!")
            if right.kind == "error":
                return right
            return _compare(left, right, node["operator"])
        if node_type == "function_call":
            return self._function(node, context, trace)
        raise ValueError(f"Unknown calculation node type: {node_type}")

    def _function(
        self,
        node: dict[str, Any],
        context: CalculationExecutionContext,
        trace: list[dict[str, Any]],
    ) -> ScalarValue:
        name = node["function_name"]
        arguments = node["arguments"]
        definition = self._function_registry.get(name)
        if definition is None:
            raise ValueError(f"Unregistered calculation function: {name}")
        if definition.lazy:
            if name == "IF":
                condition = self._evaluate_node(arguments[0], context, trace)
                truth = _truthy(condition)
                if isinstance(truth, ScalarValue):
                    return truth
                selected = arguments[1] if truth else arguments[2] if len(arguments) == 3 else None
                if selected is None:
                    return ScalarValue.boolean(False)
                value = self._evaluate_node(selected, context, trace)
                return value if isinstance(value, ScalarValue) else ScalarValue.error("#VALUE!")
            if name == "IFERROR":
                value = self._evaluate_node(arguments[0], context, trace)
                if isinstance(value, _RangeValue):
                    return ScalarValue.error("#VALUE!")
                if value.kind != "error":
                    return value
                fallback = self._evaluate_node(arguments[1], context, trace)
                return (
                    fallback
                    if isinstance(fallback, ScalarValue)
                    else ScalarValue.error("#VALUE!")
                )
            raise ValueError(f"Unsupported lazy calculation function: {name}")

        if name in {"AND", "OR"}:
            result = name == "AND"
            found_logical = False
            for argument in arguments:
                value = self._evaluate_node(argument, context, trace)
                values = value.values if isinstance(value, _RangeValue) else (value,)
                for item in values:
                    if item.kind == "error":
                        return item
                    if item.kind in {"text", "blank"}:
                        if isinstance(value, _RangeValue):
                            continue
                        return ScalarValue.error("#VALUE!")
                    truth = _truthy(item)
                    if isinstance(truth, ScalarValue):
                        return truth
                    found_logical = True
                    result = result and truth if name == "AND" else result or truth
            if not found_logical:
                return ScalarValue.error("#VALUE!")
            return ScalarValue.boolean(result)

        if name in {"COUNT", "COUNTA"}:
            count = 0
            for argument in arguments:
                value = self._evaluate_node(argument, context, trace)
                values = value.values if isinstance(value, _RangeValue) else (value,)
                for item in values:
                    if name == "COUNT" and item.kind in {"number", "date_serial"}:
                        count += 1
                    elif name == "COUNTA" and item.kind != "blank":
                        count += 1
            return ScalarValue.number(count)

        if name == "MATCH":
            lookup = self._evaluate_node(arguments[0], context, trace)
            lookup_array = self._evaluate_node(arguments[1], context, trace)
            if isinstance(lookup, _RangeValue) or not isinstance(
                lookup_array,
                _RangeValue,
            ):
                return ScalarValue.error("#VALUE!")
            if lookup_array.rows > 1 and lookup_array.columns > 1:
                return ScalarValue.error("#VALUE!")
            match_type = 1.0
            if len(arguments) == 3:
                raw_match_type = self._evaluate_node(arguments[2], context, trace)
                match_type_value = _coerce_numeric(raw_match_type)
                if isinstance(match_type_value, ScalarValue):
                    return match_type_value
                match_type = match_type_value
            if match_type not in {-1.0, 0.0, 1.0}:
                return ScalarValue.error("#VALUE!")
            return _match(lookup, lookup_array.values, int(match_type))

        if name == "COUNTIF":
            criteria_range = self._evaluate_node(arguments[0], context, trace)
            criteria = self._evaluate_node(arguments[1], context, trace)
            if isinstance(criteria, _RangeValue):
                return ScalarValue.error("#VALUE!")
            values = (
                criteria_range.values
                if isinstance(criteria_range, _RangeValue)
                else (criteria_range,)
            )
            count = 0
            for item in values:
                matched = _countif_match(item, criteria)
                if isinstance(matched, ScalarValue):
                    return matched
                count += int(matched)
            return ScalarValue.number(count)

        if name == "MINIFS":
            if len(arguments) < 3 or len(arguments) % 2 == 0:
                return ScalarValue.error("#VALUE!")
            minimum_range = self._evaluate_node(arguments[0], context, trace)
            if not isinstance(minimum_range, _RangeValue):
                return ScalarValue.error("#VALUE!")
            criteria_pairs: list[tuple[_RangeValue, ScalarValue]] = []
            for index in range(1, len(arguments), 2):
                criteria_range = self._evaluate_node(arguments[index], context, trace)
                criteria = self._evaluate_node(arguments[index + 1], context, trace)
                if (
                    not isinstance(criteria_range, _RangeValue)
                    or isinstance(criteria, _RangeValue)
                    or criteria_range.rows != minimum_range.rows
                    or criteria_range.columns != minimum_range.columns
                ):
                    return ScalarValue.error("#VALUE!")
                criteria_pairs.append((criteria_range, criteria))
            matched_values: list[float] = []
            for value_index, candidate in enumerate(minimum_range.values):
                matched = True
                for criteria_range, criteria in criteria_pairs:
                    criterion_match = _countif_match(
                        criteria_range.values[value_index],
                        criteria,
                    )
                    if isinstance(criterion_match, ScalarValue):
                        return criterion_match
                    if not criterion_match:
                        matched = False
                        break
                if not matched:
                    continue
                if candidate.kind == "error":
                    return candidate
                if candidate.kind in {"number", "date_serial"}:
                    matched_values.append(candidate.number_value)
            if not matched_values:
                return ScalarValue.number(0)
            return _finite_number(min(matched_values))

        if name == "IRR":
            values = self._evaluate_node(arguments[0], context, trace)
            if not isinstance(values, _RangeValue):
                return ScalarValue.error("#VALUE!")
            cash_flows = _numeric_range_values(values)
            if isinstance(cash_flows, ScalarValue):
                return cash_flows
            guess = 0.1
            if len(arguments) == 2:
                guess_value = self._evaluate_node(arguments[1], context, trace)
                guess_number = _coerce_numeric(guess_value)
                if isinstance(guess_number, ScalarValue):
                    return guess_number
                guess = guess_number
            return _irr(cash_flows, guess)

        if name == "XIRR":
            values = self._evaluate_node(arguments[0], context, trace)
            dates = self._evaluate_node(arguments[1], context, trace)
            paired = _dated_cash_flows(values, dates)
            if isinstance(paired, ScalarValue):
                return paired
            cash_flows, day_offsets = paired
            guess = 0.1
            if len(arguments) == 3:
                guess_value = self._evaluate_node(arguments[2], context, trace)
                guess_number = _coerce_numeric(guess_value)
                if isinstance(guess_number, ScalarValue):
                    return guess_number
                guess = guess_number
            return _xirr(cash_flows, day_offsets, guess)

        if name == "XNPV":
            rate_value = self._evaluate_node(arguments[0], context, trace)
            values = self._evaluate_node(arguments[1], context, trace)
            dates = self._evaluate_node(arguments[2], context, trace)
            rate = _coerce_numeric(rate_value)
            if isinstance(rate, ScalarValue):
                return rate
            paired = _dated_cash_flows(values, dates)
            if isinstance(paired, ScalarValue):
                return paired
            cash_flows, day_offsets = paired
            return _xnpv(rate, cash_flows, day_offsets)

        if name == "NPV":
            rate_value = self._evaluate_node(arguments[0], context, trace)
            rate = _coerce_numeric(rate_value)
            if isinstance(rate, ScalarValue):
                return rate
            cash_flows: list[float] = []
            for argument in arguments[1:]:
                value = self._evaluate_node(argument, context, trace)
                if isinstance(value, _RangeValue):
                    range_values = _numeric_range_values(value)
                    if isinstance(range_values, ScalarValue):
                        return range_values
                    cash_flows.extend(range_values)
                    continue
                number = _coerce_numeric(value)
                if isinstance(number, ScalarValue):
                    return number
                cash_flows.append(number)
            return _npv(rate, cash_flows)

        if name in {"SUM", "AVERAGE", "MIN", "MAX"}:
            numeric_values: list[float] = []
            for argument in arguments:
                value = self._evaluate_node(argument, context, trace)
                if isinstance(value, _RangeValue):
                    for item in value.values:
                        if item.kind == "error":
                            return item
                        if item.kind in {"number", "date_serial"}:
                            numeric_values.append(item.number_value)
                    continue
                number = _coerce_numeric(value)
                if isinstance(number, ScalarValue):
                    return number
                numeric_values.append(number)
            if name == "SUM":
                return _safe_fsum(numeric_values)
            if name == "AVERAGE":
                if not numeric_values:
                    return ScalarValue.error("#DIV/0!")
                total = _safe_fsum(numeric_values)
                if total.kind == "error":
                    return total
                return _finite_number(total.number_value / len(numeric_values))
            if not numeric_values:
                return ScalarValue.number(0)
            return _finite_number(
                min(numeric_values) if name == "MIN" else max(numeric_values)
            )

        first = self._evaluate_node(arguments[0], context, trace)
        first_number = _coerce_numeric(first)
        if isinstance(first_number, ScalarValue):
            return first_number
        if name == "MOD":
            divisor_value = self._evaluate_node(arguments[1], context, trace)
            divisor = _coerce_numeric(divisor_value)
            if isinstance(divisor, ScalarValue):
                return divisor
            if divisor == 0:
                return ScalarValue.error("#DIV/0!")
            try:
                result = first_number - divisor * math.floor(first_number / divisor)
            except (ArithmeticError, OverflowError, ValueError):
                return ScalarValue.error("#NUM!")
            return _finite_number(result)
        if name == "YEAR":
            return _year(first_number, context.catalog.workbook_date_system)
        if name == "ABS":
            return _finite_number(abs(first_number))
        if name == "ROUND":
            second = self._evaluate_node(arguments[1], context, trace)
            digits = _coerce_numeric(second)
            if isinstance(digits, ScalarValue):
                return digits
            try:
                quantum = Decimal(1).scaleb(-int(digits))
                rounded = Decimal(str(first_number)).quantize(
                    quantum,
                    rounding=ROUND_HALF_UP,
                )
            except (InvalidOperation, OverflowError, ValueError):
                return ScalarValue.error("#NUM!")
            return _finite_number(float(rounded))
        raise ValueError(f"Unregistered calculation function: {name}")

    def _cell_value(
        self,
        cell: dict[str, Any],
        context: CalculationExecutionContext,
        trace: list[dict[str, Any]],
    ) -> ScalarValue:
        reference = WorkbookCellRef(
            cell["workbook_version_id"],
            cell["sheet_name"],
            cell["sheet_position"],
            cell["cell_address"],
        )
        if reference in context.input_values:
            value = context.input_values[reference]
        elif reference in context.calculated_values:
            value = context.calculated_values[reference]
        else:
            fact = context.catalog.cell(reference)
            if fact.formula_kind is not None:
                value = ScalarValue.error("#REF!")
            else:
                value = _scalar_from_fact(fact, context.catalog.workbook_date_system)
        if len(trace) < context.configuration.max_trace_inputs:
            trace.append(
                {
                    "sheet_name": reference.sheet_name,
                    "sheet_position": reference.sheet_position,
                    "cell_address": reference.cell_address,
                    "value": value.to_json(),
                }
            )
        return value

    def _range_value(
        self,
        node: dict[str, Any],
        context: CalculationExecutionContext,
        trace: list[dict[str, Any]],
    ) -> _RangeValue:
        start = node["start_cell"]
        end = node["end_cell"]
        start_row, start_column = coordinate_to_tuple(start["cell_address"])
        end_row, end_column = coordinate_to_tuple(end["cell_address"])
        values: list[ScalarValue] = []
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                cell = dict(start)
                cell["cell_address"] = f"{get_column_letter(column)}{row}"
                values.append(self._cell_value(cell, context, trace))
        return _RangeValue(
            tuple(values),
            rows=end_row - start_row + 1,
            columns=end_column - start_column + 1,
        )


def _scalar_from_fact(fact: WorkbookCellFact, date_system: str) -> ScalarValue:
    return scalar_from_python(fact.value, fact.value_type, date_system)


def scalar_from_python(
    value: Any,
    value_type: str,
    date_system: str,
) -> ScalarValue:
    """Convert trusted workbook/cache data into the typed execution profile."""
    if value_type == "blank":
        return ScalarValue.blank()
    if value_type == "boolean":
        return ScalarValue.boolean(bool(value))
    if value_type == "number":
        return ScalarValue.number(value)
    if value_type == "text":
        return ScalarValue.text(str(value))
    if value_type == "error":
        code = str(value)
        return ScalarValue.error(code) if code in _ERROR_CODES else ScalarValue.error("#VALUE!")
    if value_type == "date" and isinstance(value, (date, datetime)):
        epoch = CALENDAR_MAC_1904 if date_system == "1904" else CALENDAR_WINDOWS_1900
        return ScalarValue.date_serial(
            float(to_excel(value, epoch)),
            value.isoformat(),
        )
    return ScalarValue.error("#VALUE!")


def _coerce_numeric(value: ScalarValue | _RangeValue) -> float | ScalarValue:
    if isinstance(value, _RangeValue):
        return ScalarValue.error("#VALUE!")
    if value.kind == "error":
        return value
    if value.kind in {"number", "date_serial"}:
        return value.number_value
    if value.kind == "blank":
        return 0.0
    if value.kind == "boolean":
        return 1.0 if value.value else 0.0
    return ScalarValue.error("#VALUE!")


def _range_binary_operation(
    left: ScalarValue | _RangeValue,
    right: ScalarValue | _RangeValue,
    operator: str,
) -> ScalarValue | _RangeValue:
    if (
        not isinstance(left, _RangeValue)
        or not isinstance(right, _RangeValue)
        or operator not in {"add", "subtract"}
        or (left.rows > 1 and left.columns > 1)
        or (right.rows > 1 and right.columns > 1)
        or left.rows != right.rows
        or left.columns != right.columns
    ):
        return ScalarValue.error("#VALUE!")
    results: list[ScalarValue] = []
    for left_item, right_item in zip(left.values, right.values):
        left_number = _coerce_numeric(left_item)
        if isinstance(left_number, ScalarValue):
            results.append(left_number)
            continue
        right_number = _coerce_numeric(right_item)
        if isinstance(right_number, ScalarValue):
            results.append(right_number)
            continue
        result = (
            left_number + right_number
            if operator == "add"
            else left_number - right_number
        )
        results.append(_finite_number(result))
    return _RangeValue(tuple(results), rows=left.rows, columns=left.columns)


def _numeric_range_values(
    value: _RangeValue,
) -> list[float] | ScalarValue:
    numbers: list[float] = []
    for item in value.values:
        if item.kind == "error":
            return item
        if item.kind in {"number", "date_serial"}:
            numbers.append(item.number_value)
    return numbers


def _strict_numeric_range_values(value: _RangeValue) -> list[float] | ScalarValue:
    if value.rows > 1 and value.columns > 1:
        return ScalarValue.error("#VALUE!")
    numbers: list[float] = []
    for item in value.values:
        if item.kind == "error":
            return item
        if item.kind not in {"number", "date_serial"}:
            return ScalarValue.error("#VALUE!")
        numbers.append(item.number_value)
    return numbers


def _dated_cash_flows(
    values: ScalarValue | _RangeValue,
    dates: ScalarValue | _RangeValue,
) -> tuple[list[float], list[float]] | ScalarValue:
    if not isinstance(values, _RangeValue) or not isinstance(dates, _RangeValue):
        return ScalarValue.error("#VALUE!")
    cash_flows = _strict_numeric_range_values(values)
    date_serials = _strict_numeric_range_values(dates)
    if isinstance(cash_flows, ScalarValue):
        return cash_flows
    if isinstance(date_serials, ScalarValue):
        return date_serials
    if not cash_flows or len(cash_flows) != len(date_serials):
        return ScalarValue.error("#VALUE!")
    whole_dates = [math.trunc(value) for value in date_serials]
    if any(value < 0 for value in whole_dates):
        return ScalarValue.error("#VALUE!")
    first_date = whole_dates[0]
    if any(value < first_date for value in whole_dates):
        return ScalarValue.error("#NUM!")
    return cash_flows, [float(value - first_date) for value in whole_dates]


def _xnpv_value(
    rate: float,
    cash_flows: Sequence[float],
    day_offsets: Sequence[float],
) -> float:
    if rate <= -1.0 or not math.isfinite(rate):
        raise ValueError("XNPV rate is outside the real-valued domain")
    base = 1.0 + rate
    return math.fsum(
        cash_flow / (base ** (day_offset / 365.0))
        for cash_flow, day_offset in zip(cash_flows, day_offsets)
    )


def _xnpv(
    rate: float,
    cash_flows: Sequence[float],
    day_offsets: Sequence[float],
) -> ScalarValue:
    try:
        value = _xnpv_value(rate, cash_flows, day_offsets)
    except (ArithmeticError, OverflowError, ValueError):
        return ScalarValue.error("#NUM!")
    return _finite_number(value)


def _year(serial: float, date_system: str) -> ScalarValue:
    if not math.isfinite(serial) or serial < 0:
        return ScalarValue.error("#NUM!")
    whole_serial = math.floor(serial)
    if date_system == "1900" and whole_serial in {0, 60}:
        return ScalarValue.number(1900)
    if date_system == "1904" and whole_serial == 0:
        return ScalarValue.number(1904)
    epoch = CALENDAR_MAC_1904 if date_system == "1904" else CALENDAR_WINDOWS_1900
    try:
        converted = from_excel(whole_serial, epoch)
        year = converted.year
    except (AttributeError, OverflowError, TypeError, ValueError):
        return ScalarValue.error("#NUM!")
    return ScalarValue.number(year)


_XIRR_MIN_RATE = -0.999999999
_XIRR_MAX_RATE = 1_000_000.0
_XIRR_MAX_ITERATIONS = 100
_XIRR_RATE_TOLERANCE = 1e-10
_XIRR_VALUE_TOLERANCE = 1e-8


def _xirr_derivative(
    rate: float,
    cash_flows: Sequence[float],
    day_offsets: Sequence[float],
) -> float:
    base = 1.0 + rate
    return math.fsum(
        -(day_offset / 365.0)
        * cash_flow
        / (base ** ((day_offset / 365.0) + 1.0))
        for cash_flow, day_offset in zip(cash_flows, day_offsets)
        if day_offset
    )


def _xirr_bracket(
    cash_flows: Sequence[float],
    day_offsets: Sequence[float],
    guess: float,
) -> tuple[float, float] | None:
    grid = sorted(
        {
            _XIRR_MIN_RATE,
            -0.99,
            -0.9,
            -0.75,
            -0.5,
            -0.25,
            0.0,
            0.1,
            0.25,
            0.5,
            1.0,
            2.0,
            5.0,
            10.0,
            100.0,
            1_000.0,
            _XIRR_MAX_RATE,
            min(max(guess, _XIRR_MIN_RATE), _XIRR_MAX_RATE),
        }
    )
    brackets: list[tuple[float, float]] = []
    previous_rate = grid[0]
    previous_value = _xnpv_value(previous_rate, cash_flows, day_offsets)
    for rate in grid[1:]:
        value = _xnpv_value(rate, cash_flows, day_offsets)
        if value == 0:
            return rate, rate
        if previous_value == 0:
            return previous_rate, previous_rate
        if math.copysign(1.0, value) != math.copysign(1.0, previous_value):
            brackets.append((previous_rate, rate))
        previous_rate, previous_value = rate, value
    if not brackets:
        return None
    return min(
        brackets,
        key=lambda pair: abs(((pair[0] + pair[1]) / 2.0) - guess),
    )


def _xirr(
    cash_flows: Sequence[float],
    day_offsets: Sequence[float],
    guess: float,
) -> ScalarValue:
    if (
        len(cash_flows) < 2
        or not any(value > 0 for value in cash_flows)
        or not any(value < 0 for value in cash_flows)
        or not math.isfinite(guess)
        or guess <= -1.0
    ):
        return ScalarValue.error("#NUM!")
    rate = min(guess, _XIRR_MAX_RATE)
    try:
        for _iteration in range(_XIRR_MAX_ITERATIONS):
            value = _xnpv_value(rate, cash_flows, day_offsets)
            if abs(value) <= _XIRR_VALUE_TOLERANCE:
                return _finite_number(rate)
            derivative = _xirr_derivative(rate, cash_flows, day_offsets)
            if not math.isfinite(derivative) or abs(derivative) <= 1e-15:
                break
            next_rate = rate - value / derivative
            if not (_XIRR_MIN_RATE < next_rate <= _XIRR_MAX_RATE):
                break
            if (
                abs(next_rate - rate) <= _XIRR_RATE_TOLERANCE
                and abs(_xnpv_value(next_rate, cash_flows, day_offsets))
                <= _XIRR_VALUE_TOLERANCE
            ):
                return _finite_number(next_rate)
            rate = next_rate

        bracket = _xirr_bracket(cash_flows, day_offsets, guess)
        if bracket is None:
            return ScalarValue.error("#NUM!")
        low, high = bracket
        if low == high:
            return _finite_number(low)
        low_value = _xnpv_value(low, cash_flows, day_offsets)
        for _iteration in range(_XIRR_MAX_ITERATIONS):
            middle = (low + high) / 2.0
            middle_value = _xnpv_value(middle, cash_flows, day_offsets)
            if (
                abs(middle_value) <= _XIRR_VALUE_TOLERANCE
                or abs(high - low) <= _XIRR_RATE_TOLERANCE
            ):
                return _finite_number(middle)
            if math.copysign(1.0, middle_value) == math.copysign(
                1.0,
                low_value,
            ):
                low, low_value = middle, middle_value
            else:
                high = middle
    except (ArithmeticError, OverflowError, ValueError):
        return ScalarValue.error("#NUM!")
    return ScalarValue.error("#NUM!")


def _irr(cash_flows: Sequence[float], guess: float) -> ScalarValue:
    if (
        len(cash_flows) < 2
        or not any(value > 0 for value in cash_flows)
        or not any(value < 0 for value in cash_flows)
        or not math.isfinite(guess)
    ):
        return ScalarValue.error("#NUM!")
    rate = guess
    for _iteration in range(20):
        if rate <= -1.0:
            return ScalarValue.error("#NUM!")
        try:
            base = 1.0 + rate
            value = math.fsum(
                cash_flow / (base**period)
                for period, cash_flow in enumerate(cash_flows)
            )
            derivative = math.fsum(
                -period * cash_flow / (base ** (period + 1))
                for period, cash_flow in enumerate(cash_flows)
                if period
            )
        except (ArithmeticError, OverflowError, ValueError):
            return ScalarValue.error("#NUM!")
        if (
            not math.isfinite(value)
            or not math.isfinite(derivative)
            or abs(derivative) <= 1e-15
        ):
            return ScalarValue.error("#NUM!")
        next_rate = rate - value / derivative
        if not math.isfinite(next_rate):
            return ScalarValue.error("#NUM!")
        if abs(next_rate - rate) <= 1e-7:
            return _finite_number(next_rate)
        rate = next_rate
    return ScalarValue.error("#NUM!")


def _npv(rate: float, cash_flows: Sequence[float]) -> ScalarValue:
    if rate == -1.0:
        return ScalarValue.error("#DIV/0!")
    try:
        base = 1.0 + rate
        value = math.fsum(
            cash_flow / (base**period)
            for period, cash_flow in enumerate(cash_flows, start=1)
        )
    except (ArithmeticError, OverflowError, ValueError):
        return ScalarValue.error("#NUM!")
    return _finite_number(value)


def _finite_number(value: float) -> ScalarValue:
    return ScalarValue.number(value) if math.isfinite(value) else ScalarValue.error("#NUM!")


def _safe_fsum(values: Sequence[float]) -> ScalarValue:
    try:
        return _finite_number(math.fsum(values))
    except (ArithmeticError, OverflowError, ValueError):
        return ScalarValue.error("#NUM!")


def _truthy(value: ScalarValue | _RangeValue) -> bool | ScalarValue:
    if isinstance(value, _RangeValue):
        return ScalarValue.error("#VALUE!")
    if value.kind == "error":
        return value
    if value.kind == "blank":
        return False
    if value.kind == "boolean":
        return bool(value.value)
    if value.kind in {"number", "date_serial"}:
        return value.number_value != 0
    return ScalarValue.error("#VALUE!")


def _wildcard_regex(source: str) -> re.Pattern[str]:
    pieces = ["^"]
    escaped = False
    for character in source:
        if escaped:
            pieces.append(re.escape(character))
            escaped = False
        elif character == "~":
            escaped = True
        elif character == "*":
            pieces.append(".*")
        elif character == "?":
            pieces.append(".")
        else:
            pieces.append(re.escape(character))
    if escaped:
        pieces.append(re.escape("~"))
    pieces.append("$")
    return re.compile("".join(pieces), re.IGNORECASE)


def _match(
    lookup: ScalarValue,
    candidates: Sequence[ScalarValue],
    match_type: int,
) -> ScalarValue:
    if lookup.kind == "error":
        return lookup
    if match_type == 0:
        pattern = (
            _wildcard_regex(str(lookup.value))
            if lookup.kind == "text"
            and any(
                character in str(lookup.value)
                for character in ("*", "?", "~")
            )
            else None
        )
        for index, candidate in enumerate(candidates, start=1):
            if candidate.kind == "error":
                return candidate
            if pattern is not None:
                matched = candidate.kind == "text" and pattern.fullmatch(
                    str(candidate.value)
                )
            else:
                compared = _compare(candidate, lookup, "equal")
                matched = compared.kind != "error" and bool(compared.value)
            if matched:
                return ScalarValue.number(index)
        return ScalarValue.error("#N/A")

    operator = "less_equal" if match_type == 1 else "greater_equal"
    best: int | None = None
    for index, candidate in enumerate(candidates, start=1):
        if candidate.kind == "error":
            return candidate
        compared = _compare(candidate, lookup, operator)
        if compared.kind == "error":
            continue
        if bool(compared.value):
            best = index
        elif best is not None:
            break
    return (
        ScalarValue.number(best)
        if best is not None
        else ScalarValue.error("#N/A")
    )


def _countif_match(
    value: ScalarValue,
    criteria: ScalarValue,
) -> bool | ScalarValue:
    if value.kind == "error":
        return value
    if criteria.kind == "error":
        return criteria
    operator = "equal"
    expected = criteria
    if criteria.kind == "text":
        source = str(criteria.value)
        if any(character in source for character in ("*", "?", "~")):
            return ScalarValue.error("#VALUE!")
        operators = (
            (">=", "greater_equal"),
            ("<=", "less_equal"),
            ("<>", "not_equal"),
            (">", "greater"),
            ("<", "less"),
            ("=", "equal"),
        )
        operand = source
        for prefix, registered in operators:
            if source.startswith(prefix):
                operator = registered
                operand = source[len(prefix) :]
                break
        try:
            number = float(operand)
        except ValueError:
            expected = ScalarValue.text(operand)
        else:
            if not math.isfinite(number):
                return ScalarValue.error("#VALUE!")
            expected = ScalarValue.number(number)
    if value.kind == "blank" and expected.kind in {"number", "date_serial"}:
        return False
    compared = _compare(value, expected, operator)
    if compared.kind == "error":
        return False
    return bool(compared.value)


def _compare(left: ScalarValue, right: ScalarValue, operator: str) -> ScalarValue:
    comparable = True
    if left.kind == "blank" and right.kind in {"blank", "number", "date_serial"}:
        left_value: Any = 0.0
        right_value = 0.0 if right.kind == "blank" else right.number_value
    elif right.kind == "blank" and left.kind in {"number", "date_serial"}:
        left_value = left.number_value
        right_value = 0.0
    elif left.kind in {"number", "date_serial"} and right.kind in {"number", "date_serial"}:
        left_value = left.number_value
        right_value = right.number_value
    elif left.kind == right.kind == "text":
        left_value = str(left.value).casefold()
        right_value = str(right.value).casefold()
    elif left.kind == right.kind == "boolean":
        left_value = bool(left.value)
        right_value = bool(right.value)
    elif left.kind == right.kind == "blank":
        left_value = right_value = 0.0
    else:
        comparable = False
        left_value = right_value = None

    if not comparable:
        if operator == "equal":
            return ScalarValue.boolean(False)
        if operator == "not_equal":
            return ScalarValue.boolean(True)
        return ScalarValue.error("#VALUE!")
    operations = {
        "equal": left_value == right_value,
        "not_equal": left_value != right_value,
        "less": left_value < right_value,
        "less_equal": left_value <= right_value,
        "greater": left_value > right_value,
        "greater_equal": left_value >= right_value,
    }
    return ScalarValue.boolean(operations[operator])
