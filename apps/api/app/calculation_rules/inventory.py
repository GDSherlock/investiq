"""Workbook-scoped formula and cached-value inventory."""

from __future__ import annotations

from hashlib import sha256
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from .types import (
    CalculationRuleExtractionConfiguration,
    FormulaIdFactory,
    WorkbookCatalog,
    WorkbookCellFact,
    WorkbookCellRef,
    WorkbookFormulaCell,
    value_type,
)


class WorkbookFormulaInventory:
    def __init__(
        self,
        configuration: CalculationRuleExtractionConfiguration | None = None,
    ):
        self._configuration = configuration or CalculationRuleExtractionConfiguration()

    def scan(self, content_bytes: bytes, workbook_version_id: str) -> WorkbookCatalog:
        formula_workbook = load_workbook(
            BytesIO(content_bytes),
            data_only=False,
            read_only=False,
            keep_links=False,
        )
        cached_workbook = load_workbook(
            BytesIO(content_bytes),
            data_only=True,
            read_only=False,
            keep_links=False,
        )
        try:
            return self._scan_workbooks(
                formula_workbook,
                cached_workbook,
                workbook_version_id,
            )
        finally:
            formula_workbook.close()
            cached_workbook.close()

    def _scan_workbooks(
        self,
        formula_workbook: Any,
        cached_workbook: Any,
        workbook_version_id: str,
    ) -> WorkbookCatalog:
        id_factory = FormulaIdFactory(workbook_version_id)
        cells: dict[tuple[int, str], WorkbookCellFact] = {}
        formulas: list[WorkbookFormulaCell] = []
        sheet_names = tuple(worksheet.title for worksheet in formula_workbook.worksheets)
        sheet_states = tuple(worksheet.sheet_state for worksheet in formula_workbook.worksheets)
        recalculation_required = self._requires_recalculation(formula_workbook)

        for sheet_position, worksheet in enumerate(formula_workbook.worksheets):
            cached_worksheet = cached_workbook[worksheet.title]
            # iter_rows() materializes the full rectangular dimension. Walking
            # loaded cells keeps sparse workbooks sparse while covering every
            # explicit formula, static value, and style-bearing cell.
            explicit_cells = sorted(
                worksheet._cells.values(),
                key=lambda item: (item.row, item.column),
            )
            for cell in explicit_cells:
                reference = WorkbookCellRef(
                    workbook_version_id,
                    worksheet.title,
                    sheet_position,
                    cell.coordinate,
                )
                (
                    formula_kind,
                    exact_formula,
                    special_range,
                    special_metadata,
                ) = self._formula_details(cell.value)
                if formula_kind is None:
                    cells[reference.key] = WorkbookCellFact(
                        ref=reference,
                        value=cell.value,
                        value_type=value_type(cell.value, cell.data_type),
                        data_type=cell.data_type,
                        number_format=cell.number_format,
                        sheet_state=worksheet.sheet_state,
                    )
                    continue

                cached_cell = cached_worksheet[cell.coordinate]
                cached_value = cached_cell.value
                cached_type = value_type(cached_value, cached_cell.data_type)
                cache_status = "available" if cached_value is not None else "missing"
                cache_freshness = (
                    "recalculation_required"
                    if recalculation_required
                    else "unknown" if cache_status == "available" else "missing"
                )
                formula = WorkbookFormulaCell(
                    id=id_factory.formula_cell_id(reference),
                    ref=reference,
                    exact_formula=exact_formula,
                    formula_sha256=sha256(exact_formula.encode("utf-8")).hexdigest(),
                    formula_kind=formula_kind,
                    cached_value=cached_value,
                    cached_value_type=cached_type,
                    cache_status=cache_status,
                    cache_freshness=cache_freshness,
                    number_format=cell.number_format,
                    data_type=cell.data_type,
                    sheet_state=worksheet.sheet_state,
                    special_range=special_range,
                    special_metadata=special_metadata,
                )
                formulas.append(formula)
                cells[reference.key] = WorkbookCellFact(
                    ref=reference,
                    value=None,
                    value_type="blank",
                    data_type=cell.data_type,
                    number_format=cell.number_format,
                    sheet_state=worksheet.sheet_state,
                    formula_kind=formula_kind,
                )
                if len(formulas) > self._configuration.max_formula_count:
                    raise ValueError(
                        "Workbook formula count exceeds configured limit"
                    )

        formulas.sort(
            key=lambda item: (
                item.ref.sheet_position,
                self._coordinate_sort_key(item.ref.cell_address),
            )
        )
        date_system = "1904" if formula_workbook.epoch.year == 1904 else "1900"
        return WorkbookCatalog(
            workbook_version_id=workbook_version_id,
            formulas=tuple(formulas),
            sheet_names=sheet_names,
            sheet_states=sheet_states,
            workbook_date_system=date_system,
            recalculation_required=recalculation_required,
            _cells=cells,
        )

    @staticmethod
    def _formula_details(
        value: Any,
    ) -> tuple[str | None, str, str | None, dict[str, Any] | None]:
        if isinstance(value, ArrayFormula):
            return "array", value.text or "", value.ref, dict(value)
        if isinstance(value, DataTableFormula):
            return "data_table", "", value.ref, dict(value)
        if isinstance(value, str) and value.startswith("="):
            return "scalar", value, None, None
        return None, "", None, None

    @staticmethod
    def _requires_recalculation(workbook: Any) -> bool:
        calculation = workbook.calculation
        return bool(
            calculation.fullCalcOnLoad
            or calculation.forceFullCalc
            or calculation.calcMode == "manual"
            or calculation.calcCompleted is False
        )

    @staticmethod
    def _coordinate_sort_key(address: str) -> tuple[int, int]:
        from openpyxl.utils.cell import coordinate_to_tuple

        return coordinate_to_tuple(address)
