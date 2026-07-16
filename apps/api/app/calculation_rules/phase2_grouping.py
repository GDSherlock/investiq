"""Deterministic model-scoped grouping of copied calculation formulas."""

from __future__ import annotations

from dataclasses import dataclass
import hashlib
import json
from typing import Mapping, Sequence
import uuid

from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

from .phase2_types import Phase2CalculationConfiguration
from .types import FormulaCompilation, WorkbookCatalog, WorkbookCellRef, WorkbookFormulaCell


@dataclass(frozen=True)
class CalculationRuleMember:
    id: str
    grouped_rule_id: str
    ordinal: int
    formula_cell_id: str
    expression_id: str
    sheet_name: str
    cell_address: str
    period_offset: int
    exact_formula: str


@dataclass(frozen=True)
class GroupedCalculationRule:
    id: str
    model_version_id: str
    grouping_profile: str
    group_fingerprint: str
    label: str
    normalized_expression: str
    orientation: str
    members: tuple[CalculationRuleMember, ...]
    exceptions: tuple[dict[str, str], ...]
    confidence: float
    approval_status: str
    compiler_version: str
    semantics_profile: str


class BusinessRuleGrouper:
    def __init__(
        self,
        configuration: Phase2CalculationConfiguration | None = None,
    ) -> None:
        self._configuration = configuration or Phase2CalculationConfiguration()

    def group(
        self,
        model_version_id: str,
        catalog: WorkbookCatalog,
        compilations: Sequence[FormulaCompilation],
    ) -> tuple[GroupedCalculationRule, ...]:
        model_id = str(uuid.UUID(model_version_id))
        formula_by_id = {item.id: item for item in catalog.formulas}
        compilation_by_formula = {
            item.formula_cell_id: item
            for item in compilations
            if item.support_status == "supported"
            and item.normalized_signature is not None
        }
        candidates = [
            (formula_by_id[formula_id], compilation)
            for formula_id, compilation in compilation_by_formula.items()
        ]
        candidates.sort(key=lambda item: _formula_sort_key(item[0]))

        grouped_formula_ids: set[str] = set()
        groups: list[GroupedCalculationRule] = []
        horizontal = _buckets(candidates, orientation="horizontal")
        for key, members in horizontal:
            if len(members) < 2:
                continue
            groups.append(
                self._group_from_members(
                    model_id,
                    catalog,
                    compilation_by_formula,
                    key,
                    members,
                    "horizontal",
                )
            )
            grouped_formula_ids.update(item[0].id for item in members)

        remaining = [
            item for item in candidates if item[0].id not in grouped_formula_ids
        ]
        vertical = _buckets(remaining, orientation="vertical")
        for key, members in vertical:
            if len(members) < 2:
                continue
            groups.append(
                self._group_from_members(
                    model_id,
                    catalog,
                    compilation_by_formula,
                    key,
                    members,
                    "vertical",
                )
            )
        return tuple(
            sorted(
                groups,
                key=lambda item: (
                    item.members[0].sheet_name,
                    coordinate_to_tuple(item.members[0].cell_address),
                ),
            )
        )

    def _group_from_members(
        self,
        model_version_id: str,
        catalog: WorkbookCatalog,
        compilation_by_formula: Mapping[str, FormulaCompilation],
        bucket_key: tuple[object, ...],
        member_pairs: Sequence[tuple[WorkbookFormulaCell, FormulaCompilation]],
        orientation: str,
    ) -> GroupedCalculationRule:
        signature = str(bucket_key[-1])
        fingerprint_payload = {
            "sheet_name": member_pairs[0][0].ref.sheet_name,
            "orientation": orientation,
            "fixed_axis": bucket_key[1],
            "normalized_signature": signature,
            "semantics_profile": self._configuration.semantics_profile,
        }
        fingerprint = hashlib.sha256(
            _canonical_json(fingerprint_payload).encode("utf-8")
        ).hexdigest()
        grouped_rule_id = str(
            uuid.uuid5(
                uuid.UUID(model_version_id),
                f"{self._configuration.grouping_profile}|{fingerprint}",
            )
        )
        ordered = sorted(
            member_pairs,
            key=lambda item: _formula_sort_key(item[0]),
        )
        members = tuple(
            CalculationRuleMember(
                id=str(
                    uuid.uuid5(
                        uuid.UUID(grouped_rule_id),
                        f"member|{formula.id}",
                    )
                ),
                grouped_rule_id=grouped_rule_id,
                ordinal=ordinal,
                formula_cell_id=formula.id,
                expression_id=compilation.expression_id,
                sheet_name=formula.ref.sheet_name,
                cell_address=formula.ref.cell_address,
                period_offset=ordinal,
                exact_formula=formula.exact_formula,
            )
            for ordinal, (formula, compilation) in enumerate(ordered)
        )
        exceptions = _group_exceptions(
            catalog,
            ordered,
            compilation_by_formula,
            signature,
            orientation,
        )
        first = members[0]
        last = members[-1]
        label = (
            f"Copied formula: {first.sheet_name}!"
            f"{first.cell_address}:{last.cell_address}"
        )
        return GroupedCalculationRule(
            id=grouped_rule_id,
            model_version_id=model_version_id,
            grouping_profile=self._configuration.grouping_profile,
            group_fingerprint=fingerprint,
            label=label,
            normalized_expression=signature,
            orientation=orientation,
            members=members,
            exceptions=exceptions,
            confidence=1.0 if not exceptions else 0.8,
            approval_status="unreviewed",
            compiler_version=self._configuration.compiler_version,
            semantics_profile=self._configuration.semantics_profile,
        )


def _buckets(
    candidates: Sequence[tuple[WorkbookFormulaCell, FormulaCompilation]],
    *,
    orientation: str,
) -> list[
    tuple[
        tuple[object, ...],
        list[tuple[WorkbookFormulaCell, FormulaCompilation]],
    ]
]:
    buckets: dict[
        tuple[object, ...],
        list[tuple[WorkbookFormulaCell, FormulaCompilation]],
    ] = {}
    for formula, compilation in candidates:
        row, column = coordinate_to_tuple(formula.ref.cell_address)
        fixed_axis = row if orientation == "horizontal" else column
        key = (
            formula.ref.sheet_name,
            fixed_axis,
            compilation.normalized_signature,
        )
        buckets.setdefault(key, []).append((formula, compilation))
    return sorted(buckets.items(), key=lambda item: tuple(str(value) for value in item[0]))


def _group_exceptions(
    catalog: WorkbookCatalog,
    members: Sequence[tuple[WorkbookFormulaCell, FormulaCompilation]],
    compilation_by_formula: Mapping[str, FormulaCompilation],
    signature: str,
    orientation: str,
) -> tuple[dict[str, str], ...]:
    formula_by_ref = catalog.formula_by_ref()
    ordered_formulas = [item[0] for item in members]
    occupied = {item.ref for item in ordered_formulas}
    exceptions: list[dict[str, str]] = []
    positions = [coordinate_to_tuple(item.ref.cell_address) for item in ordered_formulas]
    start = min(position[1] if orientation == "horizontal" else position[0] for position in positions)
    end = max(position[1] if orientation == "horizontal" else position[0] for position in positions)
    fixed = positions[0][0] if orientation == "horizontal" else positions[0][1]
    for coordinate in range(start, end + 2):
        row = fixed if orientation == "horizontal" else coordinate
        column = coordinate if orientation == "horizontal" else fixed
        address = f"{get_column_letter(column)}{row}"
        reference = WorkbookCellRef(
            catalog.workbook_version_id,
            ordered_formulas[0].ref.sheet_name,
            ordered_formulas[0].ref.sheet_position,
            address,
        )
        if reference in occupied:
            continue
        formula = formula_by_ref.get(reference)
        if formula is not None:
            compilation = compilation_by_formula.get(formula.id)
            if compilation is None or compilation.normalized_signature != signature:
                exceptions.append(
                    {
                        "sheet_name": reference.sheet_name,
                        "cell_address": reference.cell_address,
                        "reason": "formula_break",
                    }
                )
            continue
        fact = catalog.cell(reference)
        if fact.value_type != "blank":
            exceptions.append(
                {
                    "sheet_name": reference.sheet_name,
                    "cell_address": reference.cell_address,
                    "reason": "hardcode_break",
                }
            )
    return tuple(exceptions)


def _formula_sort_key(formula: WorkbookFormulaCell) -> tuple[int, int, int]:
    row, column = coordinate_to_tuple(formula.ref.cell_address)
    return formula.ref.sheet_position, row, column


def _canonical_json(payload: object) -> str:
    return json.dumps(payload, sort_keys=True, separators=(",", ":"))
