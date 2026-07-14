"""
EXPERIMENTAL — isolated. Backend-enforced observation coverage + hard caps.

Execution and observation are separate facts. A read_range call contributes geometric
coverage only after every complete JSON chunk has been delivered to the agent driver.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries


@dataclass
class HardCaps:
    max_tool_calls: int = 60
    max_iterations: int = 40
    max_range_cells: int = 500
    deadline_seconds: int = 500
    max_repeated_identical: int = 6
    max_internal_chunks_per_request: int = 64
    max_internal_chunks_per_run: int = 256
    max_observed_bytes_per_run: int = 4_000_000
    reserved_submit_call: int = 1


def _range_cells(cell_range: str) -> set[tuple[int, int]]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return {
        (row, col)
        for row in range(min_row, max_row + 1)
        for col in range(min_col, max_col + 1)
    }


def _compress_cells(cells: set[tuple[int, int]]) -> list[str]:
    """Return deterministic exact A1 rectangles, including vertical ranges."""
    by_row: dict[int, list[int]] = {}
    for row, col in sorted(cells):
        by_row.setdefault(row, []).append(col)

    horizontal_runs: list[tuple[int, int, int]] = []
    for row, cols in by_row.items():
        start = previous = cols[0]
        for col in cols[1:] + [None]:
            if col is not None and col == previous + 1:
                previous = col
                continue
            horizontal_runs.append((row, start, previous))
            if col is not None:
                start = previous = col

    rectangles: list[tuple[int, int, int, int]] = []
    runs_by_columns: dict[tuple[int, int], list[int]] = {}
    for row, start_col, end_col in horizontal_runs:
        runs_by_columns.setdefault((start_col, end_col), []).append(row)
    for (start_col, end_col), rows in runs_by_columns.items():
        start_row = previous_row = rows[0]
        for row in rows[1:] + [None]:
            if row is not None and row == previous_row + 1:
                previous_row = row
                continue
            rectangles.append((start_row, start_col, previous_row, end_col))
            if row is not None:
                start_row = previous_row = row

    ranges: list[str] = []
    for start_row, start_col, end_row, end_col in sorted(rectangles):
        first = f"{get_column_letter(start_col)}{start_row}"
        last = f"{get_column_letter(end_col)}{end_row}"
        ranges.append(first if first == last else f"{first}:{last}")
    return ranges


class CoverageTracker:
    """Tracks executed chunks separately from chunks successfully observed by the agent."""

    def __init__(self, tools):
        self.tools = tools
        self.meta = tools.get_workbook_metadata()
        self.all_sheets = {s["name"] for s in self.meta["sheets"]}
        self.hidden_sheets = {
            s["name"] for s in self.meta["sheets"] if s["state"] != "visible"
        }
        self.content_sheets = tools.content_sheets()
        self.has_named_ranges = bool(self.meta["named_ranges"])
        self.external_links = list(self.meta["external_links"])
        self.workbook_version = getattr(tools, "workbook_version", None)

        self.sheet_targets: dict[str, set[tuple[int, int]]] = {}
        self.sheet_required_ranges: dict[str, str | None] = {}
        for sheet in self.meta["sheets"]:
            required_range = sheet.get("required_range")
            if not required_range:
                max_row = sheet.get("max_row")
                max_col = sheet.get("max_col")
                if max_row and max_col:
                    required_range = f"A1:{get_column_letter(max_col)}{max_row}"
            self.sheet_required_ranges[sheet["name"]] = required_range
            if sheet["name"] not in self.content_sheets:
                continue
            if required_range:
                self.sheet_targets[sheet["name"]] = _range_cells(required_range)
            else:
                self.sheet_targets[sheet["name"]] = set()

        self.workbook_non_empty: dict[str, set[str]] = {}
        for sheet_name in self.content_sheets:
            if hasattr(tools, "non_empty_cell_references"):
                self.workbook_non_empty[sheet_name] = tools.non_empty_cell_references(
                    sheet_name
                )
            else:
                self.workbook_non_empty[sheet_name] = set()

        self.inspected: set[str] = set()
        self.regions_read: set[str] = set()
        self.formulas_inspected: set[str] = set()
        self.data_validations_inspected: set[str] = set()
        self.metadata_inspected = False
        self.list_sheets_completed = False
        self.named_ranges_inspected = False

        self.logical_model_tool_calls = 0
        self.internal_chunk_fetches = 0
        self.driver_observations = 0
        self.payload_retry_count = 0
        self.observed_bytes = 0
        self.duplicate_range_requests = 0
        self._sig_counts: dict[str, int] = {}
        self._logical_operations: list[dict[str, Any]] = []
        self.submit_attempts = 0
        self.coverage_rejections = 0
        self._requests: dict[str, dict[str, Any]] = {}
        self._sheet_observed_ranges: dict[str, set[tuple[int, int]]] = {
            sheet: set() for sheet in self.all_sheets
        }
        self._sheet_observed_non_empty: dict[str, set[str]] = {
            sheet: set() for sheet in self.all_sheets
        }

    def _request_state(self, result: dict[str, Any]) -> dict[str, Any]:
        request_id = result["request_id"]
        state = self._requests.get(request_id)
        if state is None:
            state = {
                "request_id": request_id,
                "sheet_name": result.get("sheet_name"),
                "requested_range": result.get("requested_range"),
                "workbook_version": result.get("workbook_version"),
                "chunk_count": result.get("chunk_count", 0),
                "executed_indices": set(),
                "observed_indices": set(),
                "executed_chunk_ids": set(),
                "observed_chunk_ids": set(),
                "observed_ranges": {},
                "continuation_tokens": {},
                "executed_index_counts": {},
                "observed_index_counts": {},
                "duplicate_chunk_ids": set(),
                "out_of_order_chunk_ids": [],
                "invalid_chunk_ids": set(),
                "binding_errors": 0,
                "workbook_version_errors": 0,
                "total_serialized_bytes": 0,
            }
            self._requests[request_id] = state
        return state

    @staticmethod
    def _payload_is_complete(result: dict[str, Any]) -> bool:
        try:
            payload = json.dumps(
                result,
                default=str,
                ensure_ascii=False,
                separators=(",", ":"),
            )
            json.loads(payload)
        except (TypeError, ValueError):
            return False
        return (
            result.get("is_complete") is True
            and result.get("serialized_bytes") == len(payload.encode("utf-8"))
        )

    def _validate_binding(
        self,
        state: dict[str, Any],
        args: dict[str, Any],
        result: dict[str, Any],
    ) -> bool:
        valid = True
        if (
            result.get("sheet_name") != state["sheet_name"]
            or result.get("requested_range") != state["requested_range"]
            or result.get("chunk_count") != state["chunk_count"]
            or args.get("sheet_name") != state["sheet_name"]
            or str(args.get("cell_range", "")).upper() != state["requested_range"]
        ):
            state["binding_errors"] += 1
            valid = False
        if (
            result.get("workbook_version") != state["workbook_version"]
            or (
                self.workbook_version is not None
                and result.get("workbook_version") != self.workbook_version
            )
        ):
            state["workbook_version_errors"] += 1
            valid = False

        index = result.get("chunk_index")
        if not isinstance(index, int) or not 0 <= index < state["chunk_count"]:
            state["binding_errors"] += 1
            valid = False
        elif index > 0 and index - 1 in state["continuation_tokens"]:
            if args.get("continuation_token") != state["continuation_tokens"][index - 1]:
                state["binding_errors"] += 1
                valid = False
        return valid

    @property
    def tool_call_count(self) -> int:
        """Backward-compatible alias; the budget now counts model calls only."""
        return self.logical_model_tool_calls

    def record_logical_call(self, name: str, args: dict) -> None:
        self.logical_model_tool_calls += 1
        sig = name + ":" + json.dumps(
            args, sort_keys=True, ensure_ascii=False, default=str
        )
        self._sig_counts[sig] = self._sig_counts.get(sig, 0) + 1
        self._logical_operations.append({
            "tool_name": name,
            "sheet_name": args.get("sheet_name"),
            "requested_range": (
                str(args.get("cell_range", "")).upper()
                if args.get("cell_range") is not None else None
            ),
        })
        if name == "submit_extraction_result":
            self.submit_attempts += 1

    def record_internal_chunk_fetch(self, count: int = 1) -> None:
        self.internal_chunk_fetches += count

    def record_driver_observation(self, serialized_bytes: int = 0) -> None:
        self.driver_observations += 1
        self.observed_bytes += serialized_bytes

    def record_payload_retry(self) -> None:
        self.payload_retry_count += 1

    def record_duplicate_range_request(self) -> None:
        self.duplicate_range_requests += 1

    def record_execution(self, name: str, args: dict, result: Any):

        if name == "inspect_sheet" and not (isinstance(result, dict) and result.get("error")):
            self.inspected.add(args.get("sheet_name"))
        elif name == "list_sheets" and not (isinstance(result, dict) and result.get("error")):
            self.list_sheets_completed = True
        elif name == "get_formulas":
            self.formulas_inspected.add(args.get("sheet_name"))
        elif name == "get_workbook_metadata":
            self.metadata_inspected = True
        elif name == "get_named_ranges":
            self.named_ranges_inspected = True
        elif name == "get_data_validations":
            self.data_validations_inspected.add(args.get("sheet_name"))
        if name != "read_range" or not isinstance(result, dict) or result.get("error"):
            return
        required = {
            "request_id", "chunk_id", "chunk_index", "chunk_count", "sheet_name",
            "requested_range", "returned_range", "serialized_bytes",
            "workbook_version", "is_complete", "has_more", "cells",
        }
        if not required.issubset(result):
            return
        state = self._request_state(result)
        chunk_id = result["chunk_id"]
        index = result["chunk_index"]
        if chunk_id in state["executed_chunk_ids"]:
            state["duplicate_chunk_ids"].add(chunk_id)
        state["executed_chunk_ids"].add(chunk_id)
        state["executed_indices"].add(index)
        state["executed_index_counts"][index] = (
            state["executed_index_counts"].get(index, 0) + 1
        )
        state["continuation_tokens"][index] = result.get("continuation_token")
        if not self._payload_is_complete(result):
            state["invalid_chunk_ids"].add(chunk_id)
        self._validate_binding(state, args, result)

    def record_observation(self, name: str, args: dict, result: Any):
        if name != "read_range" or not isinstance(result, dict) or result.get("error"):
            return
        if "request_id" not in result or result["request_id"] not in self._requests:
            return
        state = self._requests[result["request_id"]]
        chunk_id = result.get("chunk_id")
        index = result.get("chunk_index")
        state["observed_index_counts"][index] = (
            state["observed_index_counts"].get(index, 0) + 1
        )
        if chunk_id in state["observed_chunk_ids"]:
            state["duplicate_chunk_ids"].add(chunk_id)
            return
        if not self._payload_is_complete(result):
            state["invalid_chunk_ids"].add(chunk_id)
            return
        if not self._validate_binding(state, args, result):
            return
        if index not in state["executed_indices"]:
            state["binding_errors"] += 1
            return
        next_expected = len(state["observed_indices"])
        if index != next_expected:
            state["out_of_order_chunk_ids"].append(chunk_id)

        try:
            returned_cells = _range_cells(result["returned_range"])
            requested_cells = _range_cells(state["requested_range"])
        except (TypeError, ValueError):
            state["binding_errors"] += 1
            return
        if not returned_cells <= requested_cells:
            state["binding_errors"] += 1
            return
        if result.get("range_cell_count") != len(returned_cells):
            state["binding_errors"] += 1
            return

        state["observed_indices"].add(index)
        state["observed_chunk_ids"].add(chunk_id)
        state["observed_ranges"][index] = result["returned_range"]
        state["total_serialized_bytes"] += result.get("serialized_bytes", 0)
        sheet_name = state["sheet_name"]
        self._sheet_observed_ranges[sheet_name].update(returned_cells)
        self.regions_read.add(sheet_name)
        for cell in result.get("cells", []):
            cell_ref = cell.get("cell")
            if cell_ref:
                self._sheet_observed_non_empty[sheet_name].add(cell_ref.upper())

    def record(self, name: str, args: dict, result: Any):
        """Compatibility helper for callers that already delivered the result."""
        self.record_logical_call(name, args)
        self.record_execution(name, args, result)
        self.record_observation(name, args, result)
        payload = json.dumps(result, default=str, ensure_ascii=False, separators=(",", ":"))
        self.record_driver_observation(len(payload.encode("utf-8")))

    def max_repeat_count(self) -> int:
        return max(self._sig_counts.values(), default=0)

    def _request_complete(self, state: dict[str, Any]) -> bool:
        expected = set(range(state["chunk_count"]))
        observed_range_cells: set[tuple[int, int]] = set()
        for returned_range in state["observed_ranges"].values():
            observed_range_cells.update(_range_cells(returned_range))
        return (
            state["chunk_count"] > 0
            and state["executed_indices"] == expected
            and state["observed_indices"] == expected
            and observed_range_cells == _range_cells(state["requested_range"])
            and not state["invalid_chunk_ids"]
            and state["binding_errors"] == 0
            and state["workbook_version_errors"] == 0
        )

    def _request_telemetry(self, state: dict[str, Any]) -> dict[str, Any]:
        expected = set(range(state["chunk_count"]))
        observed_cells: set[tuple[int, int]] = set()
        for returned_range in state["observed_ranges"].values():
            observed_cells.update(_range_cells(returned_range))
        requested_cells = _range_cells(state["requested_range"])
        duplicate_indexes = sorted({
            index
            for index in expected | set(state["executed_index_counts"]) | set(state["observed_index_counts"])
            if state["executed_index_counts"].get(index, 0) > 1
            or state["observed_index_counts"].get(index, 0) > 1
        })
        return {
            "request_id": state["request_id"],
            "sheet_name": state["sheet_name"],
            "requested_range": state["requested_range"],
            "workbook_version": state["workbook_version"],
            "chunk_count": state["chunk_count"],
            "executed_chunk_count": len(state["executed_indices"]),
            "observed_chunk_count": len(state["observed_indices"]),
            "missing_chunk_indexes": sorted(expected - state["observed_indices"]),
            "duplicate_chunk_indexes": duplicate_indexes,
            "total_serialized_bytes": state["total_serialized_bytes"],
            "returned_ranges": [
                state["observed_ranges"][index]
                for index in sorted(state["observed_ranges"])
            ],
            "missing_ranges": _compress_cells(requested_cells - observed_cells),
            "coverage_complete": self._request_complete(state),
        }

    def request_telemetry(self, request_id: str) -> dict[str, Any] | None:
        state = self._requests.get(request_id)
        return self._request_telemetry(state) if state is not None else None

    def _sheet_telemetry(self, sheet_name: str) -> dict[str, Any]:
        requests = [
            state for state in self._requests.values()
            if state["sheet_name"] == sheet_name
        ]
        requests.sort(key=lambda state: state["request_id"])
        target = self.sheet_targets.get(sheet_name, set())
        observed = self._sheet_observed_ranges.get(sheet_name, set())
        missing = target - observed
        incomplete = [
            state["request_id"] for state in requests if not self._request_complete(state)
        ]
        complete_request_cells: set[tuple[int, int]] = set()
        for state in requests:
            if self._request_complete(state):
                complete_request_cells.update(_range_cells(state["requested_range"]))
        redundant_incomplete = [
            state["request_id"]
            for state in requests
            if not self._request_complete(state)
            and not state["invalid_chunk_ids"]
            and state["binding_errors"] == 0
            and state["workbook_version_errors"] == 0
            and _range_cells(state["requested_range"]) <= complete_request_cells
        ]
        blocking_incomplete = sorted(set(incomplete) - set(redundant_incomplete))
        chunk_count = sum(state["chunk_count"] for state in requests)
        observed_chunk_count = sum(len(state["observed_indices"]) for state in requests)
        invalid_count = sum(len(state["invalid_chunk_ids"]) for state in requests)
        binding_errors = sum(state["binding_errors"] for state in requests)
        version_errors = sum(state["workbook_version_errors"] for state in requests)
        observation_complete = bool(requests) and not (
            missing or blocking_incomplete or invalid_count or binding_errors or version_errors
        )
        return {
            "sheet_name": sheet_name,
            "required_sheet_range": self.sheet_required_ranges.get(sheet_name),
            "requested_ranges": sorted({state["requested_range"] for state in requests}),
            "observed_ranges": _compress_cells(observed & target),
            "returned_ranges": [
                state["observed_ranges"][index]
                for state in requests
                for index in sorted(state["observed_ranges"])
            ],
            "chunk_ids": [
                chunk_id
                for state in requests
                for chunk_id in sorted(state["observed_chunk_ids"])
            ],
            "chunk_count": chunk_count,
            "observed_chunk_count": observed_chunk_count,
            "observed_cell_count": len(self._sheet_observed_non_empty.get(sheet_name, set())),
            "workbook_non_empty_cell_count": len(self.workbook_non_empty.get(sheet_name, set())),
            "non_gating_metrics": {
                "observed_cell_count": len(
                    self._sheet_observed_non_empty.get(sheet_name, set())
                ),
                "workbook_non_empty_cell_count": len(
                    self.workbook_non_empty.get(sheet_name, set())
                ),
                "gating": False,
            },
            "missing_ranges": _compress_cells(missing),
            "invalid_json_count": invalid_count,
            "duplicate_chunk_ids": sorted({
                chunk_id for state in requests for chunk_id in state["duplicate_chunk_ids"]
            }),
            "out_of_order_chunk_ids": [
                chunk_id for state in requests for chunk_id in state["out_of_order_chunk_ids"]
            ],
            "binding_error_count": binding_errors,
            "workbook_version_error_count": version_errors,
            "incomplete_request_ids": incomplete,
            "blocking_incomplete_request_ids": blocking_incomplete,
            "redundant_incomplete_request_ids": redundant_incomplete,
            "inspect_preview_only": sheet_name in self.inspected,
            "observation_complete": observation_complete,
        }

    def coverage_summary(self) -> dict[str, Any]:
        formula_sheets = {s for s, _, _ in self.tools.iter_formulas()}
        telemetry = {
            sheet: self._sheet_telemetry(sheet) for sheet in sorted(self.all_sheets)
        }
        call_counts: dict[tuple[Any, Any, Any], int] = {}
        for operation in self._logical_operations:
            key = (
                operation["tool_name"],
                operation["sheet_name"],
                operation["requested_range"],
            )
            call_counts[key] = call_counts.get(key, 0) + 1
        return {
            "total_sheets": len(self.all_sheets),
            "inspected_sheets": len(self.inspected & self.all_sheets),
            "hidden_sheets_total": len(self.hidden_sheets),
            "hidden_sheets_inspected": len(self.inspected & self.hidden_sheets),
            "content_sheets": sorted(self.content_sheets),
            "regions_read_sheets": sorted(self.regions_read),
            "fully_observed_sheets": sorted(
                sheet for sheet, data in telemetry.items() if data["observation_complete"]
            ),
            "formula_sheets": sorted(formula_sheets),
            "metadata_inspected": self.metadata_inspected,
            "named_ranges_present": self.has_named_ranges,
            "external_links": [e["source_reference"] for e in self.external_links],
            "observation_telemetry": telemetry,
            "tool_call_count": self.logical_model_tool_calls,
            "logical_model_tool_calls": self.logical_model_tool_calls,
            "internal_chunk_fetches": self.internal_chunk_fetches,
            "driver_observations": self.driver_observations,
            "payload_retry_count": self.payload_retry_count,
            "observed_bytes": self.observed_bytes,
            "duplicate_range_requests": self.duplicate_range_requests,
            "logical_call_tally": [
                {
                    "tool_name": key[0],
                    "sheet_name": key[1],
                    "requested_range": key[2],
                    "call_count": count,
                }
                for key, count in sorted(
                    call_counts.items(), key=lambda item: tuple(
                        "" if value is None else str(value) for value in item[0]
                    )
                )
            ],
            "last_logical_operations": self._logical_operations[-15:],
            "request_telemetry": [
                self._request_telemetry(state)
                for state in sorted(
                    self._requests.values(), key=lambda state: state["request_id"]
                )
            ],
            "submit_attempts": self.submit_attempts,
            "coverage_rejections": self.coverage_rejections,
            "max_repeated_identical": self.max_repeat_count(),
        }

    def coverage_status(self) -> dict[str, Any]:
        sheets = []
        required_next_reads = []
        for sheet_name in sorted(self.content_sheets):
            telemetry = self._sheet_telemetry(sheet_name)
            sheet_status = {
                "sheet_name": sheet_name,
                "required_sheet_range": telemetry["required_sheet_range"],
                "observed_ranges": telemetry["observed_ranges"],
                "missing_ranges": telemetry["missing_ranges"],
                "inspected": sheet_name in self.inspected,
                "observation_complete": telemetry["observation_complete"],
            }
            sheets.append(sheet_status)
            required_next_reads.extend(
                {"sheet_name": sheet_name, "cell_range": cell_range}
                for cell_range in telemetry["missing_ranges"]
            )
        return {
            "type": "coverage_status",
            "metadata_loaded": self.metadata_inspected,
            "all_sheets_inspected": self.all_sheets <= self.inspected,
            "sheets": sheets,
            "required_next_reads": required_next_reads,
            "submission_allowed": self.submission_allowed(),
        }

    def is_range_observed(self, sheet_name: str, cell_range: str) -> bool:
        try:
            requested = _range_cells(cell_range.upper())
        except (TypeError, ValueError):
            return False
        return requested <= self._sheet_observed_ranges.get(sheet_name, set())

    def submission_allowed(self) -> bool:
        missing_inspect = self.all_sheets - self.inspected
        if not self.metadata_inspected or missing_inspect:
            return False
        return all(
            self._sheet_telemetry(sheet)["observation_complete"]
            for sheet in self.content_sheets
        )

    def submission_gate(self) -> tuple[bool, dict[str, Any]]:
        missing_inspect = sorted(self.all_sheets - self.inspected)
        telemetry = {
            sheet: self._sheet_telemetry(sheet) for sheet in sorted(self.content_sheets)
        }
        incomplete_observation = sorted(
            sheet for sheet, data in telemetry.items() if not data["observation_complete"]
        )
        missing_requests = [
            request
            for request in (
                self._request_telemetry(state)
                for state in sorted(
                    self._requests.values(), key=lambda state: state["request_id"]
                )
            )
            if not request["coverage_complete"]
            and request["request_id"] in {
                request_id
                for sheet in self.content_sheets
                for request_id in self._sheet_telemetry(sheet)[
                    "blocking_incomplete_request_ids"
                ]
            }
        ]
        actions: list[str] = []
        if not self.metadata_inspected:
            actions.append("Call get_workbook_metadata to enumerate named ranges and external links.")
        if missing_inspect:
            actions.append(f"Inspect all remaining sheets (incl. hidden): {missing_inspect}")
        if incomplete_observation:
            actions.append(
                "Fully observe every complete read_range chunk and close all range gaps: "
                f"{incomplete_observation}"
            )

        ok = not actions
        report = {
            "received": ok,
            "coverage": {
                "total_sheets": len(self.all_sheets),
                "inspected_sheets": len(self.inspected & self.all_sheets),
                "missing_sheets": sorted(set(missing_inspect) | set(incomplete_observation)),
                "observation_telemetry": telemetry,
                "missing_requests": missing_requests,
            },
            "required_next_actions": actions,
        }
        if not ok:
            report["error"] = {
                "code": "INSUFFICIENT_COVERAGE",
                "message": "Workbook observation is incomplete.",
            }
            self.coverage_rejections += 1
        return ok, report
