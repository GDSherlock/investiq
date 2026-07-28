"""Deterministic token- and byte-bounded workbook rectangle planner."""

from __future__ import annotations

from dataclasses import dataclass, replace
import hashlib
import json
import math
from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries

from partition_contract import (
    build_partition_envelope,
    request_measurement_payload,
)
from workbook_index import WorkbookIndex


PLANNER_VERSION = "partition-v1"


@dataclass(frozen=True)
class PartitionLimits:
    max_total_tokens: int = 200_000
    max_raw_evidence_tokens: int = 120_000
    max_request_bytes: int = 512 * 1024
    max_partitions: int = 512
    max_azure_calls: int = 768
    max_reconciliation_calls: int = 16
    max_retries_per_call: int = 2
    max_context_splits_per_partition: int = 1
    max_raw_evidence_bytes_per_run: int = 24 * 1024 * 1024
    deadline_seconds: int = 30 * 60


@dataclass(frozen=True)
class WorkbookPartition:
    workbook_version: str
    partition_id: str
    parent_partition_id: str | None
    split_depth: int
    sheet_name: str
    primary_range: str
    primary_facts: tuple[dict[str, Any], ...]
    dependency_references: tuple[str, ...]
    dependency_facts: tuple[dict[str, Any], ...]
    raw_evidence_bytes: int
    estimated_raw_tokens: int
    estimated_total_tokens: int
    request_bytes: int


class PartitionPlanningError(ValueError):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        sheet_name: str | None = None,
        cell: str | None = None,
    ):
        self.code = code
        self.sheet_name = sheet_name
        self.cell = cell
        super().__init__(message)


def _json_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        default=str,
        separators=(",", ":"),
    ).encode("utf-8")


def estimate_tokens(value: Any) -> int:
    return max(1, math.ceil(len(_json_bytes(value)) / 2))


def _compact_fact(fact: dict[str, Any]) -> dict[str, Any]:
    compact = dict(fact)
    compact.pop("python_type", None)
    if compact.get("displayed_value") is None:
        compact.pop("displayed_value", None)
    if compact.get("is_external_ref") is False:
        compact.pop("is_external_ref", None)
    if compact.get("is_error") is False:
        compact.pop("is_error", None)

    warnings = [
        warning
        for warning in compact.get("parse_warnings", ())
        if warning != "displayed_value_unavailable_via_openpyxl"
    ]
    if warnings:
        compact["parse_warnings"] = warnings
    else:
        compact.pop("parse_warnings", None)
    return compact


def stable_partition_id(
    workbook_version: str,
    sheet_name: str,
    primary_range: str,
    planner_version: str = PLANNER_VERSION,
) -> str:
    identity = "\x1f".join(
        (workbook_version, sheet_name, primary_range.upper(), planner_version)
    )
    return hashlib.sha256(identity.encode("utf-8")).hexdigest()


class PartitionPlanner:
    def __init__(self, limits: PartitionLimits | None = None):
        self.limits = limits or PartitionLimits()

    def plan(self, index: WorkbookIndex) -> list[WorkbookPartition]:
        planned: list[WorkbookPartition] = []
        for sheet_name in index.content_sheets:
            self._plan_range(
                index,
                sheet_name,
                index.required_ranges[sheet_name],
                parent_partition_id=None,
                split_depth=0,
                output=planned,
            )
        return planned

    def split(
        self,
        index: WorkbookIndex,
        partition: WorkbookPartition,
    ) -> tuple[WorkbookPartition, WorkbookPartition]:
        first_range, second_range = self._split_range(
            partition.sheet_name,
            partition.primary_range,
        )
        child_depth = partition.split_depth + 1
        return (
            self._build_partition(
                index,
                partition.sheet_name,
                first_range,
                parent_partition_id=partition.partition_id,
                split_depth=child_depth,
            ),
            self._build_partition(
                index,
                partition.sheet_name,
                second_range,
                parent_partition_id=partition.partition_id,
                split_depth=child_depth,
            ),
        )

    def _plan_range(
        self,
        index: WorkbookIndex,
        sheet_name: str,
        cell_range: str,
        *,
        parent_partition_id: str | None,
        split_depth: int,
        output: list[WorkbookPartition],
    ) -> None:
        partition = self._build_partition(
            index,
            sheet_name,
            cell_range,
            parent_partition_id=parent_partition_id,
            split_depth=split_depth,
        )
        if self._fits(partition):
            output.append(partition)
            if len(output) > self.limits.max_partitions:
                raise PartitionPlanningError(
                    "partition_count_exceeded",
                    f"Partition count exceeds {self.limits.max_partitions}.",
                    sheet_name=sheet_name,
                )
            return

        try:
            first_range, second_range = self._split_range(sheet_name, cell_range)
        except PartitionPlanningError as exc:
            raise PartitionPlanningError(
                "partition_cell_too_large",
                f"Cell {exc.cell} cannot fit the configured partition budgets.",
                sheet_name=sheet_name,
                cell=exc.cell,
            ) from exc

        parent_id = partition.partition_id
        self._plan_range(
            index,
            sheet_name,
            first_range,
            parent_partition_id=parent_id,
            split_depth=split_depth + 1,
            output=output,
        )
        self._plan_range(
            index,
            sheet_name,
            second_range,
            parent_partition_id=parent_id,
            split_depth=split_depth + 1,
            output=output,
        )

    def _build_partition(
        self,
        index: WorkbookIndex,
        sheet_name: str,
        cell_range: str,
        *,
        parent_partition_id: str | None,
        split_depth: int,
    ) -> WorkbookPartition:
        primary_facts = tuple(
            _compact_fact(fact)
            for fact in index.facts_for_range(sheet_name, cell_range)
        )
        dependency_references = index.related_references(sheet_name, cell_range)
        dependency_facts = self._dependency_facts(index, dependency_references)
        raw_evidence = {
            "primary_evidence": primary_facts,
            "dependency_evidence": dependency_facts,
        }
        raw_evidence_bytes = len(_json_bytes(raw_evidence))
        provisional = WorkbookPartition(
            workbook_version=index.workbook_version,
            partition_id=stable_partition_id(
                index.workbook_version,
                sheet_name,
                cell_range,
            ),
            parent_partition_id=parent_partition_id,
            split_depth=split_depth,
            sheet_name=sheet_name,
            primary_range=cell_range.upper(),
            primary_facts=primary_facts,
            dependency_references=dependency_references,
            dependency_facts=dependency_facts,
            raw_evidence_bytes=raw_evidence_bytes,
            estimated_raw_tokens=estimate_tokens(raw_evidence),
            estimated_total_tokens=0,
            request_bytes=0,
        )
        envelope = build_partition_envelope(index, provisional)
        request_payload = request_measurement_payload(envelope)
        return replace(
            provisional,
            estimated_total_tokens=estimate_tokens(request_payload),
            request_bytes=len(_json_bytes(request_payload)),
        )

    @staticmethod
    def _dependency_facts(
        index: WorkbookIndex,
        references: tuple[str, ...],
    ) -> tuple[dict[str, Any], ...]:
        facts: list[dict[str, Any]] = []
        for reference in references:
            try:
                sheet_name, cell_range = reference.rsplit("!", 1)
                facts.extend(
                    _compact_fact(fact)
                    for fact in index.facts_for_range(sheet_name, cell_range)
                )
            except (TypeError, ValueError):
                continue
        seen: set[str] = set()
        unique: list[dict[str, Any]] = []
        for fact in facts:
            source = fact["source_reference"]
            if source not in seen:
                seen.add(source)
                unique.append(fact)
        return tuple(unique)

    def _fits(self, partition: WorkbookPartition) -> bool:
        return (
            partition.estimated_total_tokens <= self.limits.max_total_tokens
            and partition.estimated_raw_tokens <= self.limits.max_raw_evidence_tokens
            and partition.request_bytes <= self.limits.max_request_bytes
        )

    @staticmethod
    def _split_range(
        sheet_name: str,
        cell_range: str,
    ) -> tuple[str, str]:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        row_count = max_row - min_row + 1
        col_count = max_col - min_col + 1
        if row_count == 1 and col_count == 1:
            cell = f"{get_column_letter(min_col)}{min_row}"
            raise PartitionPlanningError(
                "partition_cell_too_large",
                f"Cell {cell} cannot be split.",
                sheet_name=sheet_name,
                cell=cell,
            )
        if row_count >= col_count and row_count > 1:
            midpoint = (min_row + max_row) // 2
            return (
                _range_text(min_col, min_row, max_col, midpoint),
                _range_text(min_col, midpoint + 1, max_col, max_row),
            )
        midpoint = (min_col + max_col) // 2
        return (
            _range_text(min_col, min_row, midpoint, max_row),
            _range_text(midpoint + 1, min_row, max_col, max_row),
        )


def _range_text(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    first = f"{get_column_letter(min_col)}{min_row}"
    last = f"{get_column_letter(max_col)}{max_row}"
    return first if first == last else f"{first}:{last}"


__all__ = [
    "PLANNER_VERSION",
    "PartitionLimits",
    "PartitionPlanner",
    "PartitionPlanningError",
    "WorkbookPartition",
    "estimate_tokens",
    "stable_partition_id",
]
