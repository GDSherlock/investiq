"""Deterministic reconciliation of workbook-partition candidate results."""

from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
import hashlib
import re
from typing import Any, Callable, Iterable

from openpyxl.utils import get_column_letter, range_boundaries

from workbook_index import WorkbookIndex


FINAL_LIST_BUCKETS = (
    "metadata",
    "all_assumption_candidates",
    "parameter_candidates",
    "derived_value_candidates",
    "output_candidates",
    "financial_series_candidates",
    "scenario_structures",
    "sensitivity_structures",
    "unclassified_inputs",
    "review_candidates",
)
CANDIDATE_BUCKETS = (
    "metadata",
    "all_assumption_candidates",
    "parameter_candidates",
    "derived_value_candidates",
    "output_candidates",
    "financial_series_candidates",
    "unclassified_inputs",
    "review_candidates",
)
STRUCTURE_BUCKETS = ("scenario_structures", "sensitivity_structures")
_SAFE_SHEET = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
_CANDIDATE_SOURCE_ERROR_CODES = {
    "candidate_source_missing",
    "candidate_source_invalid",
    "candidate_source_not_found",
}


class ReconciliationError(ValueError):
    def __init__(self, code: str, message: str):
        self.code = code
        super().__init__(message)


@dataclass(frozen=True)
class ReconciliationOutcome:
    final_extraction: dict[str, Any]
    accepted_candidates: int
    deduplicated_candidates: int
    conflicts: int
    reconciliation_calls: int
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _CandidateRecord:
    bucket: str
    candidate: dict[str, Any]
    sources: tuple[str, ...]
    facts: tuple[dict[str, Any], ...]


@dataclass(frozen=True)
class _RangeRef:
    sheet_name: str
    cell_range: str
    bounds: tuple[int, int, int, int]
    orientation: str | None
    length: int

    @property
    def qualified(self) -> str:
        return f"{_sheet_for_output(self.sheet_name)}!{self.cell_range}"


def _hash_identity(*parts: str) -> str:
    return hashlib.sha256("\x1f".join(parts).encode("utf-8")).hexdigest()


def deterministic_candidate_id(
    workbook_version: str,
    semantic_bucket: str,
    normalized_sources: Iterable[str],
) -> str:
    return _hash_identity(
        workbook_version,
        semantic_bucket,
        *sorted(normalized_sources),
    )


def deterministic_series_id(
    workbook_version: str,
    period_range: str,
    value_range: str,
) -> str:
    return _hash_identity(workbook_version, period_range, value_range)


def _sheet_for_output(sheet_name: str) -> str:
    if _SAFE_SHEET.fullmatch(sheet_name):
        return sheet_name
    return "'" + sheet_name.replace("'", "''") + "'"


def _normalize_source(
    index: WorkbookIndex,
    source: dict[str, Any],
) -> tuple[dict[str, str], str, tuple[dict[str, Any], ...]]:
    sheet_name = source.get("sheet_name")
    cell_range = source.get("cell")
    if not isinstance(sheet_name, str) or not isinstance(cell_range, str):
        raise ReconciliationError(
            "candidate_source_invalid",
            "Candidate source reference must contain sheet_name and cell.",
        )
    normalized_range = cell_range.replace("$", "").upper()
    try:
        range_boundaries(normalized_range)
    except ValueError as exc:
        raise ReconciliationError(
            "candidate_source_invalid",
            "Candidate source reference is not a valid A1 cell or range.",
        ) from exc
    facts = index.facts_for_range(sheet_name, normalized_range)
    if not facts:
        raise ReconciliationError(
            "candidate_source_not_found",
            "Candidate source reference is absent from workbook evidence.",
        )
    normalized = {"sheet_name": sheet_name, "cell": normalized_range}
    return normalized, f"{sheet_name}!{normalized_range}", facts


def _normalize_candidate(
    index: WorkbookIndex,
    bucket: str,
    candidate: dict[str, Any],
) -> _CandidateRecord:
    submitted_sources = candidate.get("source_references")
    if not isinstance(submitted_sources, list) or not submitted_sources:
        raise ReconciliationError(
            "candidate_source_missing",
            "Every partial candidate requires at least one source reference.",
        )
    normalized_sources: list[dict[str, str]] = []
    source_names: list[str] = []
    facts: list[dict[str, Any]] = []
    for source in submitted_sources:
        if not isinstance(source, dict):
            raise ReconciliationError(
                "candidate_source_invalid",
                "Candidate source reference must be an object.",
            )
        normalized, source_name, source_facts = _normalize_source(index, source)
        normalized_sources.append(normalized)
        source_names.append(source_name)
        facts.extend(source_facts)

    normalized_candidate = deepcopy(candidate)
    authoritative = facts[0]
    normalized_candidate.update({
        "candidate_id": deterministic_candidate_id(
            index.workbook_version,
            bucket,
            source_names,
        ),
        "source_references": normalized_sources,
        "raw_value": authoritative.get("raw_value"),
        "displayed_value": authoritative.get("displayed_value"),
        "formula_status": authoritative.get("formula_status"),
        "number_format": authoritative.get("number_format"),
        "data_type": authoritative.get("data_type"),
        "reconciliation_status": "validated_source",
    })
    return _CandidateRecord(
        bucket=bucket,
        candidate=normalized_candidate,
        sources=tuple(sorted(source_names)),
        facts=tuple(deepcopy(fact) for fact in facts),
    )


def _source_rejected_candidate(
    index: WorkbookIndex,
    partition_id: str,
    bucket: str,
    item_index: int,
    candidate: dict[str, Any],
) -> dict[str, Any]:
    rejected = deepcopy(candidate)
    candidate_id = rejected.get("candidate_id")
    if not isinstance(candidate_id, str) or not candidate_id.strip():
        rejected["candidate_id"] = _hash_identity(
            index.workbook_version,
            partition_id,
            bucket,
            str(item_index),
            "source-rejected",
        )
    rejected["source_contract_bucket"] = bucket
    return rejected


def _series_range_rejected_candidate(
    index: WorkbookIndex,
    partition_id: str,
    item_index: int,
    descriptor: dict[str, Any],
) -> dict[str, Any]:
    rejected = deepcopy(descriptor)
    rejected["candidate_id"] = _hash_identity(
        index.workbook_version,
        partition_id,
        "financial_series",
        str(item_index),
        "range-rejected",
    )
    rejected["original_label"] = str(
        descriptor.get("label")
        or descriptor.get("series_id")
        or "Unlabelled financial series"
    )
    rejected["submitted_role"] = "financial_series"
    rejected["source_references"] = []
    rejected["source_contract_bucket"] = "financial_series"
    rejected["reconciliation_rejection_reason"] = "series_range_invalid"
    rejected.setdefault("raw_value", None)
    rejected.setdefault("displayed_value", None)
    rejected.setdefault("period", None)
    rejected.setdefault("formula_status", None)
    rejected.setdefault("canonical_name", descriptor.get("series_id"))
    rejected.setdefault("evidence", [])
    return rejected


def _semantic_signature(record: _CandidateRecord) -> tuple[Any, ...]:
    candidate = record.candidate
    return (
        record.bucket,
        candidate.get("submitted_role"),
        candidate.get("business_role"),
        candidate.get("scenario"),
        candidate.get("period"),
    )


def _formula_compatible(records: list[_CandidateRecord]) -> list[_CandidateRecord]:
    contains_formula = any(fact.get("formula") for record in records for fact in record.facts)
    if contains_formula:
        compatible = [
            record
            for record in records
            if str(record.candidate.get("submitted_role", "")).startswith("formula_")
        ]
    else:
        compatible = [
            record
            for record in records
            if not str(record.candidate.get("submitted_role", "")).startswith("formula_")
        ]
    return compatible


def _review_candidate(
    index: WorkbookIndex,
    records: list[_CandidateRecord],
) -> dict[str, Any]:
    first = deepcopy(records[0].candidate)
    buckets = sorted({record.bucket for record in records})
    first.update({
        "candidate_id": deterministic_candidate_id(
            index.workbook_version,
            "review_candidates",
            records[0].sources,
        ),
        "reconciliation_status": "review_required",
        "conflicting_buckets": buckets,
    })
    return first


class PartitionReconciler:
    def __init__(self, *, max_reconciliation_calls: int = 16):
        self.max_reconciliation_calls = max_reconciliation_calls

    def reconcile(
        self,
        index: WorkbookIndex,
        partials: list[dict[str, Any]],
        conflict_resolver: Callable[[dict[str, Any]], dict[str, Any] | None] | None = None,
    ) -> ReconciliationOutcome:
        final: dict[str, Any] = {bucket: [] for bucket in FINAL_LIST_BUCKETS}
        records_by_source: dict[tuple[str, ...], list[_CandidateRecord]] = {}
        series: list[dict[str, Any]] = []
        deduplicated = 0
        conflict_count = 0
        reconciliation_calls = 0
        warnings: list[str] = []

        for partial in partials:
            if partial.get("workbook_version") != index.workbook_version:
                raise ReconciliationError(
                    "partition_workbook_mismatch",
                    "Partial result belongs to a different workbook.",
                )
            result = partial.get("result")
            if not isinstance(result, dict):
                raise ReconciliationError(
                    "partition_result_invalid",
                    "Partial result is missing its typed result object.",
                )
            for bucket in CANDIDATE_BUCKETS:
                items = result.get(bucket, [])
                if not isinstance(items, list):
                    raise ReconciliationError(
                        "partition_bucket_invalid",
                        f"Partial bucket {bucket} must be a list.",
                    )
                for item_index, item in enumerate(items):
                    if not isinstance(item, dict):
                        raise ReconciliationError(
                            "partition_candidate_invalid",
                            f"Partial bucket {bucket} contains a non-object.",
                        )
                    try:
                        record = _normalize_candidate(index, bucket, item)
                    except ReconciliationError as exc:
                        if exc.code not in _CANDIDATE_SOURCE_ERROR_CODES:
                            raise
                        final["review_candidates"].append(
                            _source_rejected_candidate(
                                index,
                                str(partial.get("partition_id")),
                                bucket,
                                item_index,
                                item,
                            )
                        )
                        warnings.append(
                            "candidate_source_rejected:"
                            f"{partial.get('partition_id')}:{bucket}:"
                            f"{item_index}:{exc.code}"
                        )
                        continue
                    records_by_source.setdefault(record.sources, []).append(record)

            for bucket in STRUCTURE_BUCKETS:
                for item_index, structure in enumerate(result.get(bucket, []) or []):
                    try:
                        record = _normalize_candidate(index, bucket, structure)
                    except ReconciliationError as exc:
                        if exc.code not in _CANDIDATE_SOURCE_ERROR_CODES:
                            raise
                        final["review_candidates"].append(
                            _source_rejected_candidate(
                                index,
                                str(partial.get("partition_id")),
                                bucket,
                                item_index,
                                structure,
                            )
                        )
                        warnings.append(
                            "candidate_source_rejected:"
                            f"{partial.get('partition_id')}:{bucket}:"
                            f"{item_index}:{exc.code}"
                        )
                        continue
                    final[bucket].append(record.candidate)

            submitted_series = result.get("financial_series", [])
            if not isinstance(submitted_series, list):
                raise ReconciliationError(
                    "partition_series_invalid",
                    "financial_series must be a list.",
                )
            for item_index, descriptor in enumerate(submitted_series):
                if not isinstance(descriptor, dict):
                    raise ReconciliationError(
                        "partition_series_invalid",
                        "Financial-series descriptor must be an object.",
                    )
                try:
                    _parse_range(
                        descriptor.get("period_range"),
                        default_sheet=descriptor.get("sheet_name"),
                    )
                    _parse_range(
                        descriptor.get("value_range"),
                        default_sheet=descriptor.get("sheet_name"),
                    )
                except ReconciliationError as exc:
                    if exc.code != "series_range_invalid":
                        raise
                    final["review_candidates"].append(
                        _series_range_rejected_candidate(
                            index,
                            str(partial.get("partition_id")),
                            item_index,
                            descriptor,
                        )
                    )
                    warnings.append(
                        "financial_series_rejected:"
                        f"{partial.get('partition_id')}:{item_index}:{exc.code}"
                    )
                    continue
                series.append(deepcopy(descriptor))

        for sources, records in records_by_source.items():
            unique_by_signature: dict[tuple[Any, ...], _CandidateRecord] = {}
            for record in records:
                signature = _semantic_signature(record)
                if signature in unique_by_signature:
                    deduplicated += 1
                else:
                    unique_by_signature[signature] = record
            unique = list(unique_by_signature.values())
            if len(unique) == 1:
                final[unique[0].bucket].append(unique[0].candidate)
                continue

            compatible = _formula_compatible(unique)
            if len(compatible) == 1:
                final[compatible[0].bucket].append(compatible[0].candidate)
                conflict_count += 1
                continue

            conflict_count += 1
            selected: _CandidateRecord | None = None
            if (
                conflict_resolver is not None
                and reconciliation_calls < self.max_reconciliation_calls
            ):
                conflict_id = _hash_identity(index.workbook_version, *sources)
                resolution = conflict_resolver({
                    "conflict_id": conflict_id,
                    "allowed_buckets": sorted({record.bucket for record in unique}),
                    "candidates": [
                        {
                            "bucket": record.bucket,
                            "submitted_role": record.candidate.get("submitted_role"),
                            "business_role": record.candidate.get("business_role"),
                            "source_references": record.candidate["source_references"],
                        }
                        for record in unique
                    ],
                    "validated_facts": [
                        {
                            "source_reference": fact.get("source_reference"),
                            "raw_value": fact.get("raw_value"),
                            "formula": fact.get("formula"),
                            "formula_status": fact.get("formula_status"),
                        }
                        for fact in unique[0].facts
                    ],
                })
                reconciliation_calls += 1
                if isinstance(resolution, dict):
                    selected_bucket = resolution.get("selected_bucket")
                    selected = next(
                        (
                            record
                            for record in unique
                            if record.bucket == selected_bucket
                        ),
                        None,
                    )
            if selected is not None:
                selected.candidate["reconciliation_status"] = "conflict_resolved"
                final[selected.bucket].append(selected.candidate)
            else:
                final["review_candidates"].append(_review_candidate(index, unique))

        final["financial_series"] = self._reconcile_series(index, series)
        final["coverage_declaration"] = {
            "partitioned": True,
            "completed_partitions": len(partials),
            "workbook_version": index.workbook_version,
        }
        accepted_candidates = sum(len(final[bucket]) for bucket in FINAL_LIST_BUCKETS)
        return ReconciliationOutcome(
            final_extraction=final,
            accepted_candidates=accepted_candidates,
            deduplicated_candidates=deduplicated,
            conflicts=conflict_count,
            reconciliation_calls=reconciliation_calls,
            warnings=tuple(warnings),
        )

    def _reconcile_series(
        self,
        index: WorkbookIndex,
        submitted: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
        for descriptor in submitted:
            if not isinstance(descriptor, dict):
                raise ReconciliationError(
                    "partition_series_invalid",
                    "Financial-series descriptor must be an object.",
                )
            normalized = deepcopy(descriptor)
            period = _parse_range(
                normalized.get("period_range"),
                default_sheet=normalized.get("sheet_name"),
            )
            value = _parse_range(
                normalized.get("value_range"),
                default_sheet=normalized.get("sheet_name"),
            )
            if not index.facts_for_range(period.sheet_name, period.cell_range):
                raise ReconciliationError(
                    "series_source_not_found",
                    "Financial-series period range has no workbook evidence.",
                )
            if not index.facts_for_range(value.sheet_name, value.cell_range):
                raise ReconciliationError(
                    "series_source_not_found",
                    "Financial-series value range has no workbook evidence.",
                )
            normalized["period_range"] = period.qualified
            normalized["value_range"] = value.qualified
            key = tuple(
                normalized.get(field)
                for field in (
                    "label",
                    "semantic_role",
                    "business_role",
                    "category",
                    "unit",
                    "frequency",
                    "scenario",
                    "entity",
                    "currency",
                )
            )
            grouped.setdefault(key, []).append(normalized)

        reconciled: list[dict[str, Any]] = []
        for descriptors in grouped.values():
            descriptors.sort(key=_series_sort_key)
            merged: list[dict[str, Any]] = []
            for descriptor in descriptors:
                if merged:
                    joined = _merge_series_pair(merged[-1], descriptor)
                    if joined is not None:
                        merged[-1] = joined
                        continue
                merged.append(descriptor)
            for descriptor in merged:
                descriptor["series_id"] = deterministic_series_id(
                    index.workbook_version,
                    descriptor["period_range"],
                    descriptor["value_range"],
                )
                descriptor["reconciliation_status"] = "validated_source"
                reconciled.append(descriptor)
        return reconciled


def _parse_range(value: Any, *, default_sheet: Any = None) -> _RangeRef:
    if not isinstance(value, str) or not value.strip():
        raise ReconciliationError(
            "series_range_invalid",
            "Financial-series range must be a non-empty A1 range.",
        )
    raw = value.strip()
    if "!" in raw:
        sheet_token, cell_range = raw.rsplit("!", 1)
        sheet_name = sheet_token.strip("'").replace("''", "'")
        if not sheet_name.strip():
            raise ReconciliationError(
                "series_range_invalid",
                "Qualified financial-series range requires a sheet name.",
            )
    elif isinstance(default_sheet, str) and default_sheet:
        sheet_name = default_sheet
        cell_range = raw
    else:
        raise ReconciliationError(
            "series_range_invalid",
            "Unqualified financial-series range requires sheet_name.",
        )
    normalized_range = cell_range.replace("$", "").upper()
    try:
        bounds = range_boundaries(normalized_range)
    except ValueError as exc:
        raise ReconciliationError(
            "series_range_invalid",
            "Financial-series range is not valid A1 notation.",
        ) from exc
    min_col, min_row, max_col, max_row = bounds
    rows = max_row - min_row + 1
    cols = max_col - min_col + 1
    orientation = "horizontal" if rows == 1 else "vertical" if cols == 1 else None
    return _RangeRef(
        sheet_name=sheet_name,
        cell_range=normalized_range,
        bounds=bounds,
        orientation=orientation,
        length=rows * cols,
    )


def _series_sort_key(descriptor: dict[str, Any]) -> tuple[Any, ...]:
    period = _parse_range(
        descriptor["period_range"],
        default_sheet=descriptor.get("sheet_name"),
    )
    min_col, min_row, _, _ = period.bounds
    return (period.sheet_name, period.orientation or "", min_row, min_col)


def _merge_series_pair(
    left: dict[str, Any],
    right: dict[str, Any],
) -> dict[str, Any] | None:
    left_period = _parse_range(left["period_range"], default_sheet=left.get("sheet_name"))
    right_period = _parse_range(right["period_range"], default_sheet=right.get("sheet_name"))
    left_value = _parse_range(left["value_range"], default_sheet=left.get("sheet_name"))
    right_value = _parse_range(right["value_range"], default_sheet=right.get("sheet_name"))
    if (
        left_period.sheet_name != right_period.sheet_name
        or left_value.sheet_name != right_value.sheet_name
        or left_period.orientation is None
        or left_period.orientation != right_period.orientation
        or left_value.orientation != right_value.orientation
        or left_period.length != left_value.length
        or right_period.length != right_value.length
    ):
        return None

    if left_period.orientation == "horizontal":
        adjacent = (
            left_period.bounds[1] == right_period.bounds[1]
            and left_value.bounds[1] == right_value.bounds[1]
            and right_period.bounds[0] == left_period.bounds[2] + 1
            and right_value.bounds[0] == left_value.bounds[2] + 1
        )
    else:
        adjacent = (
            left_period.bounds[0] == right_period.bounds[0]
            and left_value.bounds[0] == right_value.bounds[0]
            and right_period.bounds[1] == left_period.bounds[3] + 1
            and right_value.bounds[1] == left_value.bounds[3] + 1
        )
    if not adjacent:
        return None

    merged = deepcopy(left)
    merged["period_range"] = _joined_range(left_period, right_period)
    merged["value_range"] = _joined_range(left_value, right_value)
    return merged


def _joined_range(left: _RangeRef, right: _RangeRef) -> str:
    min_col = min(left.bounds[0], right.bounds[0])
    min_row = min(left.bounds[1], right.bounds[1])
    max_col = max(left.bounds[2], right.bounds[2])
    max_row = max(left.bounds[3], right.bounds[3])
    first = f"{get_column_letter(min_col)}{min_row}"
    last = f"{get_column_letter(max_col)}{max_row}"
    cell_range = first if first == last else f"{first}:{last}"
    return f"{_sheet_for_output(left.sheet_name)}!{cell_range}"


__all__ = [
    "FINAL_LIST_BUCKETS",
    "PartitionReconciler",
    "ReconciliationError",
    "ReconciliationOutcome",
    "deterministic_candidate_id",
    "deterministic_series_id",
]
