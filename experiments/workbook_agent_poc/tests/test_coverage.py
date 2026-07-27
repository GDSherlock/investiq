"""TDD for backend-enforced observation coverage and hard caps."""

import os
import sys

import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from workbook_tools import WorkbookToolset
from coverage_gate import CoverageTracker, HardCaps


def _tracker(tmp_path):
    path = tmp_path / "coverage.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model"
    for row in range(1, 7):
        for col in range(1, 7):
            ws.cell(row, col, f"r{row}c{col}-" + "x" * 180)
    wb.save(path)
    tools = WorkbookToolset(file_path=str(path))
    cov = CoverageTracker(tools)
    cov.record("get_workbook_metadata", {}, tools.get_workbook_metadata())
    cov.record("inspect_sheet", {"sheet_name": "Model"}, tools.inspect_sheet("Model"))
    return tools, cov


def _chunks(tools, budget=3_500):
    chunks = []
    args = []
    token = None
    while True:
        call_args = {"sheet_name": "Model", "cell_range": "A1:F6"}
        if token:
            call_args["continuation_token"] = token
        chunk = tools.read_range(
            "Model",
            "A1:F6",
            continuation_token=token,
            max_serialized_bytes=budget,
        )
        chunks.append(chunk)
        args.append(call_args)
        if not chunk["has_more"]:
            return args, chunks
        token = chunk["continuation_token"]


def _execute_all(cov, args, chunks):
    for call_args, chunk in zip(args, chunks):
        cov.record_execution("read_range", call_args, chunk)


def _observe_all(cov, args, chunks):
    for call_args, chunk in zip(args, chunks):
        cov.record_observation("read_range", call_args, chunk)


def _record_range(tools, cov, sheet_name, cell_range, budget=12_000):
    cov.record_logical_call(
        "read_range", {"sheet_name": sheet_name, "cell_range": cell_range}
    )
    token = None
    while True:
        args = {"sheet_name": sheet_name, "cell_range": cell_range}
        if token:
            args["continuation_token"] = token
        chunk = tools.read_range(
            sheet_name,
            cell_range,
            continuation_token=token,
            max_serialized_bytes=budget,
        )
        cov.record_execution("read_range", args, chunk)
        cov.record_observation("read_range", args, chunk)
        if not chunk["has_more"]:
            return
        token = chunk["continuation_token"]


def test_last_chunk_not_observed_blocks_submission(tmp_path):
    tools, cov = _tracker(tmp_path)
    args, chunks = _chunks(tools)
    _execute_all(cov, args, chunks)
    _observe_all(cov, args[:-1], chunks[:-1])

    ok, report = cov.submission_gate()

    assert ok is False
    assert report["error"]["code"] == "INSUFFICIENT_COVERAGE"
    telemetry = cov.coverage_summary()["observation_telemetry"]["Model"]
    assert telemetry["observed_chunk_count"] == len(chunks) - 1
    assert telemetry["observation_complete"] is False
    assert telemetry["incomplete_request_ids"] == [chunks[0]["request_id"]]


def test_duplicate_and_out_of_order_chunks_are_recorded_but_deduplicated(tmp_path):
    tools, cov = _tracker(tmp_path)
    args, chunks = _chunks(tools)
    _execute_all(cov, args, chunks)

    for index in reversed(range(len(chunks))):
        cov.record_observation("read_range", args[index], chunks[index])
    cov.record_observation("read_range", args[0], chunks[0])

    ok, _ = cov.submission_gate()
    telemetry = cov.coverage_summary()["observation_telemetry"]["Model"]

    assert ok is True
    assert telemetry["observed_chunk_count"] == len(chunks)
    assert chunks[0]["chunk_id"] in telemetry["duplicate_chunk_ids"]
    assert telemetry["out_of_order_chunk_ids"]
    assert telemetry["missing_ranges"] == []


def test_range_gap_blocks_even_when_all_chunk_indexes_were_observed(tmp_path):
    tools, cov = _tracker(tmp_path)
    args, chunks = _chunks(tools)
    broken = [dict(chunk) for chunk in chunks]
    broken[-1]["returned_range"] = "A6:E6"
    broken[-1]["range_cell_count"] = 5
    broken[-1] = tools._with_serialized_bytes(broken[-1])
    _execute_all(cov, args, broken)
    _observe_all(cov, args, broken)

    ok, _ = cov.submission_gate()
    telemetry = cov.coverage_summary()["observation_telemetry"]["Model"]

    assert ok is False
    assert "F6" in telemetry["missing_ranges"]
    assert telemetry["observation_complete"] is False


def test_cross_request_binding_and_workbook_version_change_block_coverage(tmp_path):
    tools, cov = _tracker(tmp_path)
    args, chunks = _chunks(tools)
    _execute_all(cov, args, chunks)
    cov.record_observation("read_range", args[0], chunks[0])

    bad_binding = dict(chunks[1])
    bad_binding["requested_range"] = "A1:E6"
    bad_binding = tools._with_serialized_bytes(bad_binding)
    cov.record_observation("read_range", args[1], bad_binding)

    bad_version = dict(chunks[-1])
    bad_version["workbook_version"] = "stale-workbook"
    bad_version = tools._with_serialized_bytes(bad_version)
    cov.record_observation("read_range", args[-1], bad_version)

    ok, _ = cov.submission_gate()
    telemetry = cov.coverage_summary()["observation_telemetry"]["Model"]

    assert ok is False
    assert telemetry["binding_error_count"] >= 1
    assert telemetry["workbook_version_error_count"] >= 1


def test_inspect_sheet_is_preview_only_and_cannot_complete_coverage(tmp_path):
    tools, cov = _tracker(tmp_path)

    ok, _ = cov.submission_gate()
    telemetry = cov.coverage_summary()["observation_telemetry"]["Model"]

    assert ok is False
    assert telemetry["chunk_count"] == 0
    assert telemetry["observed_cell_count"] == 0
    assert telemetry["observation_complete"] is False


def test_observation_telemetry_reports_cells_ranges_and_workbook_count(tmp_path):
    tools, cov = _tracker(tmp_path)
    args, chunks = _chunks(tools)
    _execute_all(cov, args, chunks)
    _observe_all(cov, args, chunks)

    telemetry = cov.coverage_summary()["observation_telemetry"]["Model"]

    assert telemetry["requested_ranges"] == ["A1:F6"]
    assert telemetry["returned_ranges"] == [chunk["returned_range"] for chunk in chunks]
    assert telemetry["chunk_count"] == len(chunks)
    assert telemetry["observed_chunk_count"] == len(chunks)
    assert telemetry["observed_cell_count"] == 36
    assert telemetry["workbook_non_empty_cell_count"] == 36
    assert telemetry["invalid_json_count"] == 0
    assert telemetry["observation_complete"] is True

    request = cov.coverage_summary()["request_telemetry"][0]
    assert request["request_id"] == chunks[0]["request_id"]
    assert request["requested_range"] == "A1:F6"
    assert request["chunk_count"] == len(chunks)
    assert request["executed_chunk_count"] == len(chunks)
    assert request["observed_chunk_count"] == len(chunks)
    assert request["missing_chunk_indexes"] == []
    assert request["duplicate_chunk_indexes"] == []
    assert request["total_serialized_bytes"] == sum(
        chunk["serialized_bytes"] for chunk in chunks
    )
    assert request["coverage_complete"] is True


def test_completed_geometry_is_not_poisoned_by_redundant_incomplete_request(tmp_path):
    tools, cov = _tracker(tmp_path)
    args, chunks = _chunks(tools)
    _execute_all(cov, args, chunks)
    _observe_all(cov, args, chunks)

    duplicate_args, duplicate_chunks = _chunks(tools)
    cov.record_execution("read_range", duplicate_args[0], duplicate_chunks[0])
    cov.record_observation("read_range", duplicate_args[0], duplicate_chunks[0])

    ok, _ = cov.submission_gate()
    telemetry = cov.coverage_summary()["observation_telemetry"]["Model"]

    assert ok is True
    assert telemetry["observation_complete"] is True
    assert telemetry["redundant_incomplete_request_ids"] == [
        duplicate_chunks[0]["request_id"]
    ]


def test_gate_reports_exact_missing_request_chunks_and_ranges(tmp_path):
    tools, cov = _tracker(tmp_path)
    args, chunks = _chunks(tools)
    _execute_all(cov, args, chunks)
    _observe_all(cov, args[:-1], chunks[:-1])

    ok, report = cov.submission_gate()

    assert ok is False
    missing = report["coverage"]["missing_requests"][0]
    assert missing["request_id"] == chunks[0]["request_id"]
    assert missing["missing_chunk_indexes"] == [len(chunks) - 1]
    assert missing["missing_ranges"]


def test_blank_boundary_geometry_returns_compact_vertical_missing_ranges(tmp_path):
    path = tmp_path / "blank-boundary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue"
    ws["A1"].number_format = "0"
    for row in range(3, 17):
        for col in range(2, 24):
            ws.cell(row, col, f"r{row}c{col}")
    wb.save(path)
    tools = WorkbookToolset(file_path=str(path))
    cov = CoverageTracker(tools)
    cov.record("get_workbook_metadata", {}, tools.get_workbook_metadata())
    cov.record("inspect_sheet", {"sheet_name": "Revenue"}, tools.inspect_sheet("Revenue"))

    _record_range(tools, cov, "Revenue", "B3:W16")

    status = cov.coverage_status()
    sheet = status["sheets"][0]
    assert status["submission_allowed"] is False
    assert sheet["required_sheet_range"] == "A1:W16"
    assert sheet["observed_ranges"] == ["B3:W16"]
    assert sheet["missing_ranges"] == ["A1:W2", "A3:A16"]
    assert cov.coverage_summary()["observation_telemetry"]["Revenue"][
        "observation_complete"
    ] is False

    _record_range(tools, cov, "Revenue", "A1:W2")
    _record_range(tools, cov, "Revenue", "A3:A16")

    completed = cov.coverage_status()
    assert completed["submission_allowed"] is True
    assert completed["sheets"][0]["missing_ranges"] == []
    assert cov.coverage_summary()["observation_telemetry"]["Revenue"][
        "observation_complete"
    ] is True


def test_repeated_identical_calls_tracked(tmp_path):
    _, cov = _tracker(tmp_path)
    for _ in range(4):
        cov.record_logical_call(
            "get_cell", {"sheet_name": "Model", "cell_reference": "C3"}
        )
    assert cov.max_repeat_count() >= 4


def test_hard_caps_defaults():
    caps = HardCaps()
    assert caps.max_tool_calls > 0
    assert caps.max_iterations > 0
    assert caps.max_repeated_identical > 0
    assert caps.deadline_seconds == 30 * 60
    assert caps.max_internal_chunks_per_request > 0
    assert caps.max_internal_chunks_per_run >= caps.max_internal_chunks_per_request
    assert caps.max_observed_bytes_per_run > 0
    assert caps.reserved_submit_call == 1
