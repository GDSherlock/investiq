"""Backend-owned materialization for canonical financial time series."""

from datetime import datetime
import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from time_series import (
    SeriesMaterializationError,
    canonical_series_to_points,
    materialize_financial_series,
)
from workbook_tools import WorkbookToolset


def _save_series_workbook(
    tmp_path,
    *,
    periods=None,
    values=None,
    sheet_name="Series",
    vertical=False,
):
    periods = periods if periods is not None else ["2025", "2026", "2027", "2028"]
    values = values if values is not None else [1.25, 2.25, 3.25, 4.25]
    path = tmp_path / "series.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if vertical:
        ws["A3"] = "TOTAL REVENUE"
        for offset, (period, value) in enumerate(zip(periods, values), start=3):
            ws.cell(offset, 2, period)
            ws.cell(offset, 3, value)
    else:
        ws["B4"] = "TOTAL REVENUE"
        for offset, (period, value) in enumerate(zip(periods, values), start=3):
            ws.cell(3, offset, period)
            ws.cell(4, offset, value)
    wb.save(path)
    return path


def _descriptor(
    *,
    series_id="revenue_total",
    label="TOTAL REVENUE",
    period_range="Series!C3:F3",
    value_range="Series!C4:F4",
    **extra,
):
    descriptor = {
        "series_id": series_id,
        "label": label,
        "semantic_role": "financial_series",
        "category": "revenue",
        "unit": "USD M",
        "frequency": "annual",
        "period_range": period_range,
        "value_range": value_range,
        "label_reference": "Series!B4",
        "reasoning_summary": "Complete annual revenue row.",
        "llm_confidence": 0.98,
    }
    descriptor.update(extra)
    return descriptor


def _materialize(path, extraction):
    return materialize_financial_series(
        WorkbookToolset(file_path=str(path)), extraction
    )


def test_descriptor_only_materializes_complete_horizontal_series(tmp_path):
    path = _save_series_workbook(tmp_path)
    extraction = {"financial_series": [_descriptor()]}

    outcome = _materialize(path, extraction)

    assert outcome["summary"]["submitted_descriptors"] == 1
    assert outcome["summary"]["materialized_series"] == 1
    assert outcome["summary"]["validated_series"] == 1
    assert len(extraction["financial_series"]) == 1
    series = extraction["financial_series"][0]
    assert series["orientation"] == "horizontal"
    assert [point["display_label"] for point in series["period_axis"]["periods"]] == [
        "2025", "2026", "2027", "2028"
    ]
    assert [point["source_cell"] for point in series["period_axis"]["periods"]] == [
        "Series!C3", "Series!D3", "Series!E3", "Series!F3"
    ]
    assert [point["value"] for point in series["value_axis"]["values"]] == [
        1.25, 2.25, 3.25, 4.25
    ]
    assert [point["source_cell"] for point in series["value_axis"]["values"]] == [
        "Series!C4", "Series!D4", "Series!E4", "Series!F4"
    ]
    assert series["materialization_status"] == "materialized"
    assert series["validation_status"] == "validated"
    assert extraction["financial_series_descriptors"] == [_descriptor()]


def test_materialization_preserves_business_role_in_canonical_and_validation_results(tmp_path):
    path = _save_series_workbook(tmp_path)
    extraction = {"financial_series": [_descriptor(business_role="cfads")]}

    outcome = _materialize(path, extraction)

    assert extraction["financial_series"][0]["business_role"] == "cfads"
    assert outcome["validation_results"][0]["business_role"] == "cfads"


def test_backend_range_resolution_adds_audit_warning_and_rereads_workbook_points(tmp_path):
    path = _save_series_workbook(tmp_path)
    extraction = {
        "financial_series": [
            _descriptor(
                period_axis={"periods": ["wrong"] * 4},
                value_axis={"values": [999] * 4},
            )
        ],
        "range_resolutions": [
            {
                "field": "period_range",
                "submitted": "2025-2028",
                "resolved": "Series!C3:F3",
                "strategy": "unique_integer_span_match",
                "partition_id": "partition-series",
            }
        ],
    }

    outcome = _materialize(path, extraction)

    series = extraction["financial_series"][0]
    assert series["period_axis"]["source_range"] == "Series!C3:F3"
    assert [point["raw_label"] for point in series["period_axis"]["periods"]] == [
        "2025", "2026", "2027", "2028"
    ]
    assert [point["value"] for point in series["value_axis"]["values"]] == [
        1.25, 2.25, 3.25, 4.25
    ]
    assert "PERIOD_RANGE_RESOLVED_FROM_WORKBOOK_EVIDENCE" in series["warnings"]
    assert (
        "PERIOD_RANGE_RESOLVED_FROM_WORKBOOK_EVIDENCE"
        in outcome["validation_results"][0]["validation_warnings"]
    )


def test_unqualified_ranges_use_explicit_sheet_name_and_quoted_ranges_work(tmp_path):
    path = _save_series_workbook(tmp_path, sheet_name="Debt Schedule")
    extraction = {
        "financial_series": [
            _descriptor(
                label_reference="'Debt Schedule'!B4",
                period_range="C3:F3",
                value_range="C4:F4",
                sheet_name="Debt Schedule",
            ),
            _descriptor(
                series_id="quoted",
                period_range="'Debt Schedule'!C3:F3",
                value_range="'Debt Schedule'!C4:F4",
                label_reference="'Debt Schedule'!B4",
                scenario="Stress",
            ),
        ]
    }

    outcome = _materialize(path, extraction)

    assert outcome["summary"]["materialized_series"] == 2
    assert {series["period_axis"]["source_range"] for series in extraction["financial_series"]} == {
        "'Debt Schedule'!C3:F3"
    }


def test_legacy_arrays_are_ignored_as_source_of_truth(tmp_path):
    path = _save_series_workbook(tmp_path)
    legacy = {
        **_descriptor(),
        "period_axis": {
            "source_range": "Series!C3:F3",
            "periods": ["wrong", "wrong", "wrong", "wrong"],
        },
        "value_axis": {
            "source_range": "Series!C4:F4",
            "values": [999, 999, 999, 999],
        },
        "calculation_type": "formula",
        "formula_pattern": {"formula_cell_count": 99},
    }
    legacy.pop("period_range")
    legacy.pop("value_range")
    extraction = {"financial_series_candidates": [legacy]}

    outcome = _materialize(path, extraction)

    assert outcome["summary"]["submitted_descriptors"] == 0
    assert outcome["summary"]["legacy_series_detected"] == 1
    series = extraction["financial_series"][0]
    assert [point["value"] for point in series["value_axis"]["values"]] == [
        1.25, 2.25, 3.25, 4.25
    ]
    assert series["calculation_type"] == "hardcoded"
    assert series["formula_pattern"]["formula_cell_count"] == 0
    assert "LEGACY_PERIOD_ARRAY_DISAGREEMENT" in series["warnings"]
    assert "LEGACY_VALUE_ARRAY_DISAGREEMENT" in series["warnings"]
    assert extraction["financial_series_candidates"] == [legacy]


def test_formula_series_remains_financial_and_uses_backend_telemetry(tmp_path):
    values = ["=C5*2", "=D5*2", "=E5*2", "=F5*2"]
    path = _save_series_workbook(tmp_path, values=values)
    extraction = {"financial_series": [_descriptor()]}

    _materialize(path, extraction)

    series = extraction["financial_series"][0]
    assert series["semantic_role"] == "financial_series"
    assert series["calculation_type"] == "formula"
    assert series["formula_pattern"] == {
        "is_formula_series": True,
        "formula_cell_count": 4,
        "static_cell_count": 0,
        "blank_cell_count": 0,
        "pattern_consistent": True,
    }
    assert all(point["formula"] for point in series["value_axis"]["values"])
    assert all(point["cached_value_available"] is False for point in series["value_axis"]["values"])
    assert all(point["cached_value_freshness"] == "unknown" for point in series["value_axis"]["values"])


def test_hardcoded_zero_row_is_not_blank(tmp_path):
    path = _save_series_workbook(tmp_path, values=[0, 0, 0, 0])
    extraction = {"financial_series": [_descriptor()]}

    _materialize(path, extraction)

    series = extraction["financial_series"][0]
    assert series["calculation_type"] == "hardcoded"
    assert series["formula_pattern"]["static_cell_count"] == 4
    assert series["formula_pattern"]["blank_cell_count"] == 0
    assert [point["value"] for point in series["value_axis"]["values"]] == [0, 0, 0, 0]


def test_all_blank_row_is_blank_and_text_markers_are_not_zero(tmp_path):
    blank_path = _save_series_workbook(tmp_path, values=[None, None, None, None])
    blank_extraction = {"financial_series": [_descriptor()]}
    _materialize(blank_path, blank_extraction)

    blank_series = blank_extraction["financial_series"][0]
    assert blank_series["calculation_type"] == "blank"
    assert blank_series["formula_pattern"]["blank_cell_count"] == 4

    marker_path = _save_series_workbook(
        tmp_path,
        values=["N/A", "-", "NM", -2.5],
    )
    marker_extraction = {"financial_series": [_descriptor()]}
    _materialize(marker_path, marker_extraction)

    assert [point["value"] for point in marker_extraction["financial_series"][0]["value_axis"]["values"]] == [
        "N/A", "-", "NM", -2.5
    ]


def test_mixed_formula_static_and_blank_counts_are_backend_owned(tmp_path):
    path = _save_series_workbook(tmp_path, values=["=C5*2", 2.25, None, "=F5*2"])
    extraction = {"financial_series": [_descriptor()]}

    _materialize(path, extraction)

    series = extraction["financial_series"][0]
    assert series["calculation_type"] == "mixed"
    assert series["formula_pattern"]["formula_cell_count"] == 2
    assert series["formula_pattern"]["static_cell_count"] == 1
    assert series["formula_pattern"]["blank_cell_count"] == 1
    assert [point["value"] for point in series["value_axis"]["values"]] == [
        None, 2.25, None, None
    ]
    assert "PARTIALLY_BLANK_SERIES" in series["warnings"]
    assert "MIXED_FORMULA_STATIC_SERIES" in series["warnings"]


def test_period_normalization_preserves_raw_labels_and_safe_fields(tmp_path):
    period_values = [2025, "FY25", "2026E", "Q1 2027", datetime(2028, 1, 31)]
    path = _save_series_workbook(tmp_path, periods=period_values, values=[1, 2, 3, 4, 5])
    extraction = {
        "financial_series": [
            _descriptor(period_range="Series!C3:G3", value_range="Series!C4:G4")
        ]
    }

    _materialize(path, extraction)

    points = extraction["financial_series"][0]["period_axis"]["periods"]
    assert [point["raw_label"] for point in points] == period_values
    expected_annual = {
        "display_label": "2025", "period_type": "annual", "year": 2025,
        "quarter": None, "month": None, "is_forecast": None,
    }
    assert {key: points[0][key] for key in expected_annual} == expected_annual
    assert points[1]["year"] == 2025
    assert points[2]["year"] == 2026 and points[2]["is_forecast"] is True
    assert points[3]["quarter"] == 1 and points[3]["year"] == 2027
    assert points[4]["year"] == 2028 and points[4]["month"] == 1


def test_misaligned_ranges_return_structured_failure(tmp_path):
    path = _save_series_workbook(tmp_path, periods=[2025, 2026, 2027, 2028, 2029], values=[1, 2, 3, 4, 5])
    extraction = {
        "financial_series": [
            _descriptor(period_range="Series!C3:G3", value_range="Series!C4:F4")
        ]
    }

    outcome = _materialize(path, extraction)

    assert extraction["financial_series"] == []
    result = outcome["validation_results"][0]
    assert result["materialization_status"] == "failed"
    assert result["validation_status"] == "rejected"
    assert result["error_code"] == "PERIOD_VALUE_LENGTH_MISMATCH"
    assert result["number_of_periods"] == 0
    assert result["source_ranges"] == {
        "period_axis": "Series!C3:G3",
        "value_axis": "Series!C4:F4",
    }
    assert outcome["summary"]["period_value_mismatches"] == 1


def test_invalid_descriptor_returns_structured_failure_instead_of_key_error(tmp_path):
    path = _save_series_workbook(tmp_path)
    extraction = {
        "financial_series": [{
            "series_id": "missing_fields",
            "label": "TOTAL REVENUE",
            "semantic_role": "financial_series",
        }]
    }

    outcome = _materialize(path, extraction)

    assert extraction["financial_series"] == []
    result = outcome["validation_results"][0]
    assert result["error_code"] == "INVALID_SERIES_DESCRIPTOR"
    assert result["validation_status"] == "rejected"


def test_two_dimensional_and_orientation_mismatch_ranges_are_rejected(tmp_path):
    path = _save_series_workbook(tmp_path)
    extraction = {
        "financial_series": [
            _descriptor(series_id="two_d", period_range="Series!C3:F4"),
            _descriptor(series_id="orientation", value_range="Series!F1:F4"),
        ]
    }

    outcome = _materialize(path, extraction)

    by_id = {item["series_id"]: item for item in outcome["validation_results"]}
    assert by_id["two_d"]["error_code"] == "TWO_DIMENSIONAL_RANGE"
    assert by_id["orientation"]["error_code"] == "ORIENTATION_MISMATCH"


def test_missing_sheet_and_merged_axis_return_structured_failures(tmp_path):
    path = _save_series_workbook(tmp_path)
    wb = openpyxl.load_workbook(path)
    wb["Series"].merge_cells("C3:D3")
    wb.save(path)
    extraction = {
        "financial_series": [
            _descriptor(series_id="missing", period_range="Missing!C3:F3"),
            _descriptor(series_id="merged"),
        ]
    }

    outcome = _materialize(path, extraction)

    by_id = {item["series_id"]: item for item in outcome["validation_results"]}
    assert by_id["missing"]["error_code"] == "SHEET_NOT_FOUND"
    assert by_id["merged"]["error_code"] == "MERGED_CELL_AMBIGUITY"


def test_representative_cell_from_multi_period_row_is_rejected(tmp_path):
    periods = [str(year) for year in range(2025, 2035)]
    path = _save_series_workbook(tmp_path, periods=periods, values=list(range(10)))
    extraction = {
        "financial_series": [
            _descriptor(period_range="Series!H3", value_range="Series!H4")
        ]
    }

    outcome = _materialize(path, extraction)

    assert outcome["validation_results"][0]["error_code"] == "REPRESENTATIVE_CELL_ONLY"
    assert outcome["summary"]["representative_cell_only"] == 1


def test_vertical_series_materializes_in_source_order(tmp_path):
    path = _save_series_workbook(tmp_path, periods=["Q1", "Q2", "Q3"], values=[10, 20, 30], vertical=True)
    extraction = {
        "financial_series": [
            _descriptor(
                period_range="Series!B3:B5",
                value_range="Series!C3:C5",
                label_reference="Series!A3",
            )
        ]
    }

    _materialize(path, extraction)

    series = extraction["financial_series"][0]
    assert series["orientation"] == "vertical"
    assert [point["source_cell"] for point in series["value_axis"]["values"]] == [
        "Series!C3", "Series!C4", "Series!C5"
    ]


def test_exact_duplicate_ranges_produce_one_canonical_series_with_alias(tmp_path):
    path = _save_series_workbook(tmp_path)
    extraction = {
        "financial_series": [
            _descriptor(),
            _descriptor(series_id="revenue_alias", label="Revenue Total"),
        ]
    }

    outcome = _materialize(path, extraction)

    assert len(extraction["financial_series"]) == 1
    assert extraction["financial_series"][0]["aliases"] == ["Revenue Total"]
    assert outcome["summary"]["duplicate_series"] == 1
    assert outcome["summary"]["rejected_series"] == 1
    duplicate = [
        result for result in outcome["validation_results"]
        if result.get("error_code") == "DUPLICATE_SERIES"
    ][0]
    assert duplicate["duplicate_of"] == "revenue_total"
    assert duplicate["source_ranges"] == {
        "period_axis": "Series!C3:F3",
        "value_axis": "Series!C4:F4",
    }
    assert duplicate["calculation_type"] == "hardcoded"


def test_detailed_schedule_is_preferred_over_dashboard_copy(tmp_path):
    path = _save_series_workbook(tmp_path)
    wb = openpyxl.load_workbook(path)
    dashboard = wb.create_sheet("Dashboard")
    dashboard["B4"] = "TOTAL REVENUE"
    for column, (period, value) in enumerate(
        zip(["2025", "2026", "2027", "2028"], [1.25, 2.25, 3.25, 4.25]),
        start=3,
    ):
        dashboard.cell(3, column, period)
        dashboard.cell(4, column, value)
    wb.save(path)
    extraction = {
        "financial_series": [
            _descriptor(
                series_id="dashboard_revenue",
                period_range="Dashboard!C3:F3",
                value_range="Dashboard!C4:F4",
                label_reference="Dashboard!B4",
            ),
            _descriptor(),
        ]
    }

    outcome = _materialize(path, extraction)

    assert [series["series_id"] for series in extraction["financial_series"]] == [
        "revenue_total"
    ]
    dashboard_result = next(
        result for result in outcome["validation_results"]
        if result["series_id"] == "dashboard_revenue"
    )
    assert dashboard_result["error_code"] == "DUPLICATE_SERIES"
    assert dashboard_result["duplicate_of"] == "revenue_total"


def test_same_label_different_scenarios_are_not_merged(tmp_path):
    path = _save_series_workbook(tmp_path)
    extraction = {
        "financial_series": [
            _descriptor(scenario="Base"),
            _descriptor(series_id="stress", scenario="Stress"),
        ]
    }

    outcome = _materialize(path, extraction)

    assert len(extraction["financial_series"]) == 2
    assert outcome["summary"]["duplicate_series"] == 0


def test_same_label_different_units_and_evidence_remain_distinct_with_warning(tmp_path):
    path = _save_series_workbook(tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Series"]
    for column, value in enumerate([10, 20, 30, 40], start=3):
        ws.cell(5, column, value)
    wb.save(path)
    extraction = {
        "financial_series": [
            _descriptor(),
            _descriptor(
                series_id="revenue_units",
                value_range="Series!C5:F5",
                unit="USD",
            ),
        ]
    }

    outcome = _materialize(path, extraction)

    assert len(extraction["financial_series"]) == 2
    assert outcome["summary"]["duplicate_series"] == 0
    assert all(
        "DUPLICATE_LABEL_DIFFERENT_RANGE" not in series["warnings"]
        for series in extraction["financial_series"]
    )


def test_same_label_context_with_different_ranges_is_warned_not_merged(tmp_path):
    path = _save_series_workbook(tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Series"]
    for column, value in enumerate([10, 20, 30, 40], start=3):
        ws.cell(5, column, value)
    wb.save(path)
    extraction = {
        "financial_series": [
            _descriptor(),
            _descriptor(series_id="different_revenue", value_range="Series!C5:F5"),
        ]
    }

    _materialize(path, extraction)

    assert len(extraction["financial_series"]) == 2
    assert all(
        "DUPLICATE_LABEL_DIFFERENT_RANGE" in series["warnings"]
        for series in extraction["financial_series"]
    )


def test_scenario_and_sensitivity_structures_are_never_materialized(tmp_path):
    path = _save_series_workbook(tmp_path)
    extraction = {
        "scenario_structures": [{"scenarios": ["Base", "Stress", "Upside"]}],
        "sensitivity_structures": [{"row_driver": "Tariff", "column_driver": "Volume"}],
    }

    outcome = _materialize(path, extraction)

    assert extraction["financial_series"] == []
    assert outcome["summary"]["materialized_series"] == 0
    assert extraction["scenario_structures"][0]["scenarios"] == ["Base", "Stress", "Upside"]
    assert extraction["sensitivity_structures"][0]["row_driver"] == "Tariff"


def test_canonical_series_to_points_returns_chart_ready_alignment(tmp_path):
    path = _save_series_workbook(tmp_path)
    extraction = {"financial_series": [_descriptor()]}
    _materialize(path, extraction)

    points = canonical_series_to_points(extraction["financial_series"][0])

    assert points[0] == {
        "period_index": 0,
        "period_label": "2025",
        "year": 2025,
        "quarter": None,
        "month": None,
        "is_forecast": None,
        "value": 1.25,
        "source_cell": "Series!C4",
    }
    assert len(points) == 4


def test_canonical_series_to_points_rejects_misaligned_canonical_axes():
    with pytest.raises(SeriesMaterializationError) as exc_info:
        canonical_series_to_points({
            "period_axis": {"periods": [{"display_label": "2025"}]},
            "value_axis": {"values": []},
        })

    assert exc_info.value.code == "PERIOD_VALUE_LENGTH_MISMATCH"
