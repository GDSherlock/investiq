"""RED/GREEN contracts for complete workbook range observations."""

import json
import os
import sys

import openpyxl
import pytest
from openpyxl.utils import range_boundaries

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from workbook_tools import ToolError, WorkbookToolset


SHEET_RANGES = {
    "Project Setup": "A1:F25",
    "Build Schedule": "A1:AB8",
    "Solar Operations": "A1:AB11",
    "Debt Sculpting": "A1:AB10",
    "Cash Flow": "A1:AB11",
    "Investor Returns": "A1:F10",
}


def _range_cells(cell_range):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return {
        (row, col)
        for row in range(min_row, max_row + 1)
        for col in range(min_col, max_col + 1)
    }


def _write_solar_shape(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, cell_range in SHEET_RANGES.items():
        ws = wb.create_sheet(sheet_name)
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        if sheet_name == "Investor Returns":
            ws.cell(min_row, min_col, "Investor Returns Summary")
            for row in range(5, 11):
                ws.cell(row, 2, row / 100)
            ws.cell(max_row, max_col).number_format = "0"
        elif sheet_name == "Project Setup":
            ws.cell(1, 1, "Project Setup")
            for row in range(2, 26):
                ws.cell(row, 3, f"input-{row}-" + "x" * 180)
            ws.cell(max_row, max_col).number_format = "0"
        else:
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row, col, f"{sheet_name}-{row}-{col}-" + "x" * 90)
    wb.save(path)


@pytest.fixture
def solar_tools(tmp_path):
    path = tmp_path / "solar-shape.xlsx"
    _write_solar_shape(path)
    return WorkbookToolset(file_path=str(path))


def _all_chunks(tools, sheet_name, requested_range, *, budget=4_000):
    chunks = []
    chunk = tools.read_range(
        sheet_name,
        requested_range,
        max_serialized_bytes=budget,
    )
    while True:
        chunks.append(chunk)
        if not chunk["has_more"]:
            return chunks
        chunk = tools.read_range(
            sheet_name,
            requested_range,
            continuation_token=chunk["continuation_token"],
            max_serialized_bytes=budget,
        )


def test_large_range_is_multiple_complete_json_payloads(solar_tools):
    chunks = _all_chunks(solar_tools, "Solar Operations", "A1:AB11")

    assert len(chunks) > 1
    assert sum(len(json.dumps(c, ensure_ascii=False)) for c in chunks) > 12_000
    for chunk in chunks:
        payload = json.dumps(chunk, ensure_ascii=False, separators=(",", ":"))
        assert json.loads(payload) == chunk
        assert len(payload.encode("utf-8")) == chunk["serialized_bytes"]
        assert chunk["serialized_bytes"] <= 4_000
        assert payload.endswith("}")


def test_project_setup_chunks_observe_c25(solar_tools):
    chunks = _all_chunks(solar_tools, "Project Setup", "A1:F25")
    observed = {cell["cell"] for chunk in chunks for cell in chunk["cells"]}

    assert "C25" in observed
    assert chunks[-1]["has_more"] is False


@pytest.mark.parametrize(
    "sheet_name,requested_range",
    [
        ("Build Schedule", "A1:AB8"),
        ("Debt Sculpting", "A1:AB10"),
        ("Solar Operations", "A1:AB11"),
        ("Cash Flow", "A1:AB11"),
    ],
)
def test_chunk_ranges_cover_requested_range_without_gaps(
    solar_tools, sheet_name, requested_range
):
    chunks = _all_chunks(solar_tools, sheet_name, requested_range)
    covered = set()
    for chunk in chunks:
        covered |= _range_cells(chunk["returned_range"])

    assert covered == _range_cells(requested_range)
    assert sum(chunk["range_cell_count"] for chunk in chunks) == len(covered)


def test_single_row_over_budget_falls_back_to_column_windows(tmp_path):
    path = tmp_path / "wide-row.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wide"
    for col in range(1, 29):
        ws.cell(1, col, "x" * 250)
    wb.save(path)
    tools = WorkbookToolset(file_path=str(path))

    chunks = _all_chunks(tools, "Wide", "A1:AB1", budget=3_500)

    assert len(chunks) > 1
    assert all(range_boundaries(c["returned_range"])[1] == 1 for c in chunks)
    assert all(range_boundaries(c["returned_range"])[3] == 1 for c in chunks)
    assert any(c["returned_range"] != "A1:AB1" for c in chunks)
    assert set().union(*(_range_cells(c["returned_range"]) for c in chunks)) == _range_cells("A1:AB1")


def test_continuation_token_is_bound_to_request_and_workbook_version(solar_tools):
    first = solar_tools.read_range(
        "Project Setup", "A1:F25", max_serialized_bytes=4_000
    )
    token = first["continuation_token"]

    with pytest.raises(ToolError, match="continuation_binding_mismatch"):
        solar_tools.read_range(
            "Project Setup",
            "A1:F24",
            continuation_token=token,
            max_serialized_bytes=4_000,
        )

    solar_tools._workbook_version = "changed"
    with pytest.raises(ToolError, match="workbook_version_changed"):
        solar_tools.read_range(
            "Project Setup",
            "A1:F25",
            continuation_token=token,
            max_serialized_bytes=4_000,
        )


def test_chunk_telemetry_and_opaque_token_are_present(solar_tools):
    first = solar_tools.read_range(
        "Debt Sculpting", "A1:AB10", max_serialized_bytes=4_000
    )

    assert first["request_id"]
    assert first["chunk_id"]
    assert first["chunk_index"] == 0
    assert first["chunk_count"] > 1
    assert first["sheet_name"] == "Debt Sculpting"
    assert first["requested_range"] == "A1:AB10"
    assert first["returned_range"]
    assert first["workbook_version"] == solar_tools.workbook_version
    assert first["has_more"] is True
    assert first["continuation_token"]
    assert "Debt Sculpting" not in first["continuation_token"]
    assert "A1:AB10" not in first["continuation_token"]


def test_small_investor_returns_range_is_one_chunk(solar_tools):
    chunk = solar_tools.read_range(
        "Investor Returns", "A1:F10", max_serialized_bytes=12_000
    )

    assert chunk["chunk_count"] == 1
    assert chunk["returned_range"] == "A1:F10"
    assert chunk["has_more"] is False
    assert chunk["continuation_token"] is None


def test_oversized_logical_range_is_partitioned_at_physical_cell_limit(tmp_path):
    path = tmp_path / "oversized-logical-range.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Large"
    ws["A1"] = "content"
    ws["W22"].number_format = "0"
    wb.save(path)
    tools = WorkbookToolset(file_path=str(path))

    chunks = _all_chunks(tools, "Large", "A1:W22", budget=12_000)

    assert len(chunks) > 1
    assert {chunk["request_id"] for chunk in chunks} == {chunks[0]["request_id"]}
    assert {chunk["requested_range"] for chunk in chunks} == {"A1:W22"}
    assert all(chunk["range_cell_count"] <= 500 for chunk in chunks)
    assert set().union(
        *(_range_cells(chunk["returned_range"]) for chunk in chunks)
    ) == _range_cells("A1:W22")


def test_cover_516_cell_logical_range_completes_in_one_request(tmp_path):
    path = tmp_path / "cover-regression.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = "Cover"
    ws["L43"].number_format = "0"
    wb.save(path)
    tools = WorkbookToolset(file_path=str(path))

    chunks = _all_chunks(tools, "Cover", "A1:L43", budget=12_000)

    assert len(chunks) == 2
    assert [chunk["chunk_index"] for chunk in chunks] == [0, 1]
    assert all(chunk["range_cell_count"] <= 500 for chunk in chunks)
    assert sum(chunk["range_cell_count"] for chunk in chunks) == 516
