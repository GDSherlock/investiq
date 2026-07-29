"""Behavior tests for sequential atomic partition orchestration."""

import json
import logging
import os
import sys

import openpyxl
import pytest


sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from partition_driver import (
    PartitionAuthenticationError,
    PartitionContextLimitError,
)
from partition_pipeline import PartitionPipelineError, run_partitioned_extraction
from partition_planner import PartitionLimits
from time_series import materialize_financial_series
from validator import validate_extraction
from workbook_tools import WorkbookToolset


def _tools(tmp_path):
    path = tmp_path / "pipeline.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Model"
    for row in range(1, 5):
        sheet.cell(row, 1, f"Label {row}")
        sheet.cell(row, 2, row * 10)
        sheet.cell(row, 3, "secret-pipeline-cell-" + "x" * 900)
    workbook.save(path)
    return WorkbookToolset(file_path=str(path))


def _limits(**overrides):
    values = {
        "max_total_tokens": 13_000,
        "max_raw_evidence_tokens": 2_000,
        "max_request_bytes": 26_000,
        "max_azure_calls": 100,
        "deadline_seconds": 60,
    }
    values.update(overrides)
    return PartitionLimits(**values)


def _candidate_result(partition):
    candidates = []
    if any(fact["source_reference"] == "Model!B1" for fact in partition.primary_facts):
        candidates.append({
            "candidate_id": "model-id",
            "original_label": "Label 1",
            "submitted_role": "hardcoded_input",
            "raw_value": 999,
            "source_references": [{"sheet_name": "Model", "cell": "B1"}],
            "reasoning_summary": "Explicit workbook value.",
            "llm_confidence": 0.9,
        })
    return {
        "workbook_version": partition.workbook_version,
        "partition_id": partition.partition_id,
        "sheet_name": partition.sheet_name,
        "primary_range": partition.primary_range,
        "result": {
            "all_assumption_candidates": candidates,
            "output_candidates": [],
        },
    }


class RecordingPartitionDriver:
    max_calls_per_operation = 1
    usage_prompt = 0
    usage_completion = 0
    _deployment = "recording-partition-driver"

    def __init__(self):
        self.call_count = 0
        self.calls = []
        self.active_calls = 0
        self.max_concurrent_calls = 0

    def extract(self, partition, envelope):
        self.active_calls += 1
        self.max_concurrent_calls = max(
            self.max_concurrent_calls,
            self.active_calls,
        )
        self.call_count += 1
        self.calls.append((partition, json.dumps(envelope, sort_keys=True, default=str)))
        result = _candidate_result(partition)
        self.active_calls -= 1
        return result

    def resolve_conflict(self, _conflict):
        raise AssertionError("No conflict is expected in this fixture.")


class ContextOverflowThenSuccessDriver(RecordingPartitionDriver):
    def __init__(self):
        super().__init__()
        self.failed = False

    def extract(self, partition, envelope):
        self.call_count += 1
        rendered = json.dumps(envelope, sort_keys=True, default=str)
        self.calls.append((partition, rendered))
        if not self.failed:
            self.failed = True
            raise PartitionContextLimitError("too large")
        return _candidate_result(partition)


class FailAfterOnePartitionDriver(RecordingPartitionDriver):
    def extract(self, partition, envelope):
        self.call_count += 1
        self.calls.append((partition, json.dumps(envelope, sort_keys=True, default=str)))
        if self.call_count > 1:
            raise PartitionAuthenticationError("denied")
        return _candidate_result(partition)


class MixedSourcePartitionDriver(RecordingPartitionDriver):
    def __init__(self):
        super().__init__()
        self.emitted = False

    def extract(self, partition, envelope):
        result = super().extract(partition, envelope)
        if not self.emitted:
            self.emitted = True
            result["result"]["all_assumption_candidates"].append({
                "candidate_id": "source-less",
                "original_label": "Unbound candidate",
                "submitted_role": "hardcoded_input",
                "raw_value": 123,
                "source_references": [],
            })
        return result


def test_pipeline_processes_partitions_sequentially_and_submits_once(tmp_path):
    tools = _tools(tmp_path)
    driver = RecordingPartitionDriver()

    run = run_partitioned_extraction(driver, tools, limits=_limits())

    assert len(driver.calls) > 1
    assert driver.max_concurrent_calls == 1
    assert run["submitted"] is True
    assert run["stop_reason"] == "submitted"
    assert run["coverage"]["submission_allowed"] is True
    assert run["coverage"]["missing_partition_ids"] == []
    assert run["final_extraction"]["all_assumption_candidates"][0]["raw_value"] == 10


def test_pipeline_output_is_accepted_by_existing_materializer_and_validator(tmp_path):
    tools = _tools(tmp_path)
    run = run_partitioned_extraction(
        RecordingPartitionDriver(),
        tools,
        limits=_limits(),
    )

    series_outcome = materialize_financial_series(
        tools,
        run["final_extraction"],
    )
    validation = validate_extraction(
        tools,
        run["final_extraction"],
        financial_series_outcome=series_outcome,
    )

    assert validation
    assert all(item["validation_status"] != "rejected" for item in validation)


def test_source_less_candidate_is_rejected_without_failing_workbook(tmp_path):
    tools = _tools(tmp_path)
    run = run_partitioned_extraction(
        MixedSourcePartitionDriver(),
        tools,
        limits=_limits(),
    )
    series_outcome = materialize_financial_series(
        tools,
        run["final_extraction"],
    )
    validation = validate_extraction(
        tools,
        run["final_extraction"],
        financial_series_outcome=series_outcome,
    )

    assert run["submitted"] is True
    assert run["stop_reason"] == "submitted"
    assert run["coverage"]["submission_allowed"] is True
    assert any(
        item["candidate_id"] == "source-less"
        and item["validation_status"] == "rejected"
        and item["invalid_source"] is True
        for item in validation
    )
    assert any(
        item["validation_status"] != "rejected"
        for item in validation
    )


def test_context_overflow_splits_once_and_never_retries_same_envelope(tmp_path):
    tools = _tools(tmp_path)
    driver = ContextOverflowThenSuccessDriver()

    run = run_partitioned_extraction(
        driver,
        tools,
        limits=_limits(max_context_splits_per_partition=1),
    )

    first_partition, first_payload = driver.calls[0]
    assert all(
        partition.partition_id != first_partition.partition_id
        and payload != first_payload
        for partition, payload in driver.calls[1:]
    )
    assert run["coverage"]["split_count"] == 1
    assert run["submitted"] is True


def test_terminal_partition_failure_discards_all_partial_results(tmp_path):
    tools = _tools(tmp_path)
    driver = FailAfterOnePartitionDriver()

    with pytest.raises(PartitionPipelineError) as exc:
        run_partitioned_extraction(driver, tools, limits=_limits())

    assert exc.value.code == "azure_authentication_failed"
    assert exc.value.completed_partition_count == 1
    assert exc.value.final_extraction is None
    assert exc.value.azure_failure is True


def test_global_call_budget_is_enforced_before_next_partition(tmp_path):
    tools = _tools(tmp_path)
    driver = RecordingPartitionDriver()

    with pytest.raises(PartitionPipelineError) as exc:
        run_partitioned_extraction(
            driver,
            tools,
            limits=_limits(max_azure_calls=1),
        )

    assert exc.value.code == "partition_call_limit_exceeded"
    assert driver.call_count == 1


def test_raw_evidence_run_budget_is_enforced_before_azure(tmp_path):
    tools = _tools(tmp_path)
    driver = RecordingPartitionDriver()

    with pytest.raises(PartitionPipelineError) as exc:
        run_partitioned_extraction(
            driver,
            tools,
            limits=_limits(max_raw_evidence_bytes_per_run=1),
        )

    assert exc.value.code == "partition_raw_evidence_limit_exceeded"
    assert driver.call_count == 0


def test_deadline_is_enforced_before_azure(monkeypatch, tmp_path):
    tools = _tools(tmp_path)
    driver = RecordingPartitionDriver()
    ticks = iter([10.0, 11.0])
    monkeypatch.setattr(
        "partition_pipeline.time.monotonic",
        lambda: next(ticks),
    )

    with pytest.raises(PartitionPipelineError) as exc:
        run_partitioned_extraction(
            driver,
            tools,
            limits=_limits(deadline_seconds=0),
        )

    assert exc.value.code == "partition_deadline_exceeded"
    assert driver.call_count == 0


def test_pipeline_logs_exclude_raw_workbook_values(tmp_path, caplog):
    tools = _tools(tmp_path)
    caplog.set_level(logging.INFO)

    run_partitioned_extraction(
        RecordingPartitionDriver(),
        tools,
        limits=_limits(),
    )

    rendered = "\n".join(record.getMessage() for record in caplog.records)
    assert "secret-pipeline-cell" not in rendered
