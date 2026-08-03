"""Behavior tests for workbook-backed partition reconciliation."""

import os
import sys

from openpyxl.utils import get_column_letter
import pytest


sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from partition_planner import WorkbookPartition
from partition_reconciler import (
    PartitionReconciler,
    ReconciliationError,
    deterministic_candidate_id,
    deterministic_series_id,
)
from workbook_index import WorkbookIndex


def _fact(sheet, cell, value, *, formula=None, status=None):
    return {
        "sheet_name": sheet,
        "cell": cell,
        "source_reference": f"{sheet}!{cell}",
        "raw_value": value,
        "displayed_value": None,
        "formula": formula,
        "formula_status": status or (
            "formula_with_cached_value" if formula else "static_value"
        ),
        "data_type": "n" if isinstance(value, (int, float)) else "s",
        "number_format": "0.00",
        "parse_warnings": [],
    }


def _index():
    inputs = (
        _fact("Inputs", "A1", "Tax rate"),
        _fact("Inputs", "B1", 0.25),
        _fact(
            "Inputs",
            "C1",
            None,
            formula="=B1*2",
            status="formula_no_cache",
        ),
    )
    forecast = []
    for offset, col in enumerate(range(3, 11)):
        letter = get_column_letter(col)
        forecast.append(_fact("Forecast", f"{letter}3", 2025 + offset))
        forecast.append(_fact("Forecast", f"{letter}8", 100 + offset))
    vertical = []
    for row in range(2, 6):
        vertical.append(_fact("Vertical", f"A{row}", 2023 + row))
        vertical.append(_fact("Vertical", f"B{row}", row * 10))
    facts = {
        "Inputs": inputs,
        "Forecast": tuple(forecast),
        "Vertical": tuple(vertical),
    }
    return WorkbookIndex(
        workbook_version="d" * 64,
        manifest={
            "sheet_count": 3,
            "hidden_sheet_count": 0,
            "sheets": [
                {"name": "Inputs", "state": "visible", "required_range": "A1:C1"},
                {"name": "Forecast", "state": "visible", "required_range": "A1:J8"},
                {"name": "Vertical", "state": "visible", "required_range": "A1:B5"},
            ],
            "required_sheet_ranges": {
                "Inputs": "A1:C1",
                "Forecast": "A1:J8",
                "Vertical": "A1:B5",
            },
            "named_ranges": [],
            "external_links": [],
        },
        content_sheets=("Inputs", "Forecast", "Vertical"),
        required_ranges={
            "Inputs": "A1:C1",
            "Forecast": "A1:J8",
            "Vertical": "A1:B5",
        },
        facts=facts,
        formulas={"Inputs!C1": "=B1*2"},
        defined_names={},
        dependency_graph={
            "precedents": {"Inputs!C1": ["Inputs!B1"]},
            "dependents": {"Inputs!B1": ["Inputs!C1"]},
            "external_refs": {},
            "ranges": {"Inputs!C1": []},
        },
        non_empty_cell_count=sum(len(items) for items in facts.values()),
    )


def _partition(identifier, sheet="Inputs", cell_range="A1:C1"):
    return WorkbookPartition(
        workbook_version="d" * 64,
        partition_id=identifier,
        parent_partition_id=None,
        split_depth=0,
        sheet_name=sheet,
        primary_range=cell_range,
        primary_facts=(),
        dependency_references=(),
        dependency_facts=(),
        raw_evidence_bytes=0,
        estimated_raw_tokens=0,
        estimated_total_tokens=0,
        request_bytes=0,
    )


def _candidate(sheet, cell, *, raw_value=999_999, role="hardcoded_input"):
    return {
        "candidate_id": "model-authored-id",
        "original_label": "Tax rate",
        "submitted_role": role,
        "raw_value": raw_value,
        "source_references": [{"sheet_name": sheet, "cell": cell}],
        "reasoning_summary": "Model explanation only.",
        "llm_confidence": 0.9,
    }


def _bound(partition, *, bucket=None, item=None, series=None):
    result = {
        "all_assumption_candidates": [],
        "output_candidates": [],
    }
    if bucket is not None:
        result.setdefault(bucket, []).append(item)
    if series is not None:
        result["financial_series"] = [series]
    return {
        "workbook_version": partition.workbook_version,
        "partition_id": partition.partition_id,
        "sheet_name": partition.sheet_name,
        "primary_range": partition.primary_range,
        "result": result,
    }


def _series(label, period_range, value_range):
    return {
        "series_id": "model-series",
        "label": label,
        "semantic_role": "financial_series",
        "business_role": "revenue",
        "category": "revenue",
        "unit": "USD",
        "frequency": "annual",
        "scenario": None,
        "entity": None,
        "currency": "USD",
        "period_range": period_range,
        "value_range": value_range,
        "reasoning_summary": "Fragment.",
        "llm_confidence": 0.9,
    }


def _add_horizontal_axis(
    index,
    *,
    sheet,
    header_row,
    value_row,
    first_column,
    start,
    count,
    add_values=True,
):
    facts = list(index.facts.get(sheet, ()))
    for offset in range(count):
        column = get_column_letter(first_column + offset)
        facts.append(_fact(sheet, f"{column}{header_row}", start + offset))
        if add_values:
            facts.append(_fact(sheet, f"{column}{value_row}", 100 + offset))
    index.facts[sheet] = tuple(facts)


def _add_vertical_axis(
    index,
    *,
    sheet,
    header_column,
    value_column,
    first_row,
    start,
    count,
):
    facts = list(index.facts.get(sheet, ()))
    for offset in range(count):
        row = first_row + offset
        facts.append(_fact(sheet, f"{header_column}{row}", start + offset))
        facts.append(_fact(sheet, f"{value_column}{row}", 100 + offset))
    index.facts[sheet] = tuple(facts)


def _rejected_series(outcome):
    assert outcome.final_extraction["financial_series"] == []
    rejected = outcome.final_extraction["review_candidates"]
    assert len(rejected) == 1
    assert rejected[0]["reconciliation_rejection_reason"] == "series_range_invalid"


def test_reconciler_replaces_model_value_with_backend_fact():
    index = _index()
    partition = _partition("partition-source")
    submitted = _bound(
        partition,
        bucket="all_assumption_candidates",
        item=_candidate("Inputs", "B1"),
    )

    outcome = PartitionReconciler().reconcile(index, [submitted])
    accepted = outcome.final_extraction["all_assumption_candidates"][0]

    assert accepted["raw_value"] == 0.25
    assert accepted["formula_status"] == "static_value"
    assert accepted["candidate_id"] == deterministic_candidate_id(
        index.workbook_version,
        "all_assumption_candidates",
        ("Inputs!B1",),
    )


def test_formula_without_cache_remains_null():
    index = _index()
    partition = _partition("partition-null")
    submitted = _bound(
        partition,
        bucket="derived_value_candidates",
        item=_candidate(
            "Inputs",
            "C1",
            raw_value=0,
            role="formula_derived_value",
        ),
    )

    accepted = PartitionReconciler().reconcile(
        index,
        [submitted],
    ).final_extraction["derived_value_candidates"][0]

    assert accepted["raw_value"] is None
    assert accepted["formula_status"] == "formula_no_cache"


def test_same_semantic_source_is_deduplicated_across_partitions():
    index = _index()
    first = _partition("partition-one")
    second = _partition("partition-two")
    duplicate = _candidate("Inputs", "B1", raw_value=0.25)
    partials = [
        _bound(first, bucket="all_assumption_candidates", item=duplicate),
        _bound(second, bucket="all_assumption_candidates", item=duplicate),
    ]

    outcome = PartitionReconciler().reconcile(index, partials)

    assert len(outcome.final_extraction["all_assumption_candidates"]) == 1
    assert outcome.deduplicated_candidates == 1


def test_source_invalid_candidate_is_quarantined_without_losing_valid_candidate():
    index = _index()
    partition = _partition("partition-mixed-source")
    valid = _candidate("Inputs", "B1")
    invalid = _candidate("Inputs", "B1")
    invalid["candidate_id"] = "missing-source"
    invalid["source_references"] = []
    submitted = _bound(partition)
    submitted["result"]["all_assumption_candidates"] = [valid, invalid]

    outcome = PartitionReconciler().reconcile(index, [submitted])

    candidates = outcome.final_extraction["all_assumption_candidates"]
    assert len(candidates) == 1
    assert candidates[0]["reconciliation_status"] == "validated_source"
    review = outcome.final_extraction["review_candidates"]
    assert any(
        candidate.get("candidate_id") == "missing-source"
        and candidate.get("source_references") == []
        and candidate.get("source_contract_bucket")
        == "all_assumption_candidates"
        for candidate in review
    )


def test_string_source_candidate_is_quarantined():
    index = _index()
    partition = _partition("partition-string-source")
    candidate = _candidate("Inputs", "B1")
    candidate["candidate_id"] = "string-source"
    candidate["source_references"] = ["Inputs!B1"]

    outcome = PartitionReconciler().reconcile(
        index,
        [_bound(
            partition,
            bucket="all_assumption_candidates",
            item=candidate,
        )],
    )

    assert outcome.final_extraction["all_assumption_candidates"] == []
    rejected = outcome.final_extraction["review_candidates"][0]
    assert rejected["candidate_id"] == "string-source"
    assert rejected["source_references"] == ["Inputs!B1"]


def test_incompatible_roles_move_to_review_when_resolver_defers():
    index = _index()
    first = _partition("partition-input")
    second = _partition("partition-output")
    output = _candidate(
        "Inputs",
        "B1",
        raw_value=0.25,
        role="hardcoded_display_output",
    )
    output["business_role"] = "unclassified"
    partials = [
        _bound(first, bucket="all_assumption_candidates", item=_candidate("Inputs", "B1")),
        _bound(second, bucket="output_candidates", item=output),
    ]

    outcome = PartitionReconciler(max_reconciliation_calls=1).reconcile(
        index,
        partials,
        conflict_resolver=lambda _conflict: None,
    )

    assert not outcome.final_extraction["all_assumption_candidates"]
    assert not outcome.final_extraction["output_candidates"]
    review = outcome.final_extraction["review_candidates"][0]
    assert review["reconciliation_status"] == "review_required"
    assert review["conflicting_buckets"] == [
        "all_assumption_candidates",
        "output_candidates",
    ]


def test_nonexistent_source_candidate_is_quarantined():
    index = _index()
    partition = _partition("partition-missing")
    submitted = _bound(
        partition,
        bucket="all_assumption_candidates",
        item=_candidate("Missing", "A1"),
    )

    outcome = PartitionReconciler().reconcile(index, [submitted])

    assert outcome.final_extraction["all_assumption_candidates"] == []
    rejected = outcome.final_extraction["review_candidates"][0]
    assert rejected["source_references"] == [
        {"sheet_name": "Missing", "cell": "A1"}
    ]
    assert rejected["source_contract_bucket"] == "all_assumption_candidates"


def test_source_less_structure_moves_to_rejected_review_candidate():
    index = _index()
    partition = _partition("partition-source-less-structure")
    submitted = _bound(
        partition,
        bucket="scenario_structures",
        item={
            "structure_id": "scenario-1",
            "source_references": [],
        },
    )

    outcome = PartitionReconciler().reconcile(index, [submitted])

    assert outcome.final_extraction["scenario_structures"] == []
    review = outcome.final_extraction["review_candidates"]
    assert len(review) == 1
    assert review[0]["source_contract_bucket"] == "scenario_structures"
    assert review[0]["source_references"] == []


@pytest.mark.parametrize(
    ("period_range", "value_range", "sheet_name"),
    [
        ("", "Forecast!C8:J8", "Forecast"),
        ("Forecast!C3:J3", "", "Forecast"),
        ("Forecast!C:C", "Forecast!C8:J8", None),
        ("Forecast!C3:J3", "Forecast!3:3", None),
        ("C3:J3", "C8:J8", None),
        ("!C3:J3", "Forecast!C8:J8", None),
        ("Forecast!not-a-range", "Forecast!C8:J8", "Forecast"),
    ],
)
def test_invalid_series_range_is_quarantined_without_losing_valid_series(
    period_range,
    value_range,
    sheet_name,
):
    index = _index()
    partition = _partition("partition-mixed-series", "Forecast", "A1:J8")
    invalid = _series("Invalid", period_range, value_range)
    invalid["sheet_name"] = sheet_name
    valid = _series(
        "Revenue",
        "Forecast!C3:J3",
        "Forecast!C8:J8",
    )
    submitted = _bound(partition)
    submitted["result"]["financial_series"] = [invalid, valid]

    outcome = PartitionReconciler().reconcile(index, [submitted])

    assert len(outcome.final_extraction["financial_series"]) == 1
    assert outcome.final_extraction["financial_series"][0]["label"] == "Revenue"
    rejected = outcome.final_extraction["review_candidates"]
    assert len(rejected) == 1
    assert rejected[0]["original_label"] == "Invalid"
    assert rejected[0]["submitted_role"] == "financial_series"
    assert rejected[0]["source_contract_bucket"] == "financial_series"
    assert (
        rejected[0]["reconciliation_rejection_reason"]
        == "series_range_invalid"
    )
    assert rejected[0]["source_references"] == []


def test_valid_a1_series_passes_through_without_a_range_resolution():
    index = _index()
    partition = _partition("partition-valid-a1", "Forecast", "A1:J8")

    outcome = PartitionReconciler().reconcile(
        index,
        [_bound(
            partition,
            series=_series("Revenue", "Forecast!C3:J3", "Forecast!C8:J8"),
        )],
    )

    assert (
        outcome.final_extraction["financial_series"][0]["period_range"]
        == "Forecast!C3:J3"
    )
    assert outcome.final_extraction["range_resolutions"] == []


def test_integer_period_span_recovers_unique_horizontal_axis_from_raw_facts():
    index = _index()
    _add_horizontal_axis(
        index,
        sheet="Forecast",
        header_row=10,
        value_row=12,
        first_column=3,
        start=2027,
        count=27,
    )
    partition = _partition("partition-2027-2053", "Forecast", "A1:AC12")

    outcome = PartitionReconciler().reconcile(
        index,
        [_bound(
            partition,
            series=_series("Revenue", "2027-2053", "Forecast!C12:AC12"),
        )],
    )

    descriptor = outcome.final_extraction["financial_series"][0]
    assert descriptor["period_range"] == "Forecast!C10:AC10"
    assert outcome.final_extraction["range_resolutions"] == [{
        "field": "period_range",
        "submitted": "2027-2053",
        "resolved": "Forecast!C10:AC10",
        "strategy": "unique_integer_span_match",
        "partition_id": "partition-2027-2053",
    }]


def test_integer_period_span_recovers_unique_vertical_axis_from_raw_facts():
    index = _index()
    _add_vertical_axis(
        index,
        sheet="Inputs",
        header_column="D",
        value_column="E",
        first_row=2,
        start=0,
        count=27,
    )
    partition = _partition("partition-0-26", "Inputs", "D1:E28")

    outcome = PartitionReconciler().reconcile(
        index,
        [_bound(
            partition,
            series=_series("Volume", "0-26", "Inputs!E2:E28"),
        )],
    )

    descriptor = outcome.final_extraction["financial_series"][0]
    assert descriptor["period_range"] == "Inputs!D2:D28"
    assert (
        outcome.final_extraction["range_resolutions"][0]["resolved"]
        == "Inputs!D2:D28"
    )


def test_integer_period_span_recovers_quoted_sheet_axis():
    index = _index()
    _add_horizontal_axis(
        index,
        sheet="Cash Flow",
        header_row=3,
        value_row=8,
        first_column=2,
        start=2027,
        count=27,
    )
    partition = _partition("partition-quoted", "Cash Flow", "A1:AB8")
    series = _series("Cash flow", "2027-2053", "'Cash Flow'!B8:AB8")
    series["sheet_name"] = "Cash Flow"

    outcome = PartitionReconciler().reconcile(index, [_bound(partition, series=series)])

    assert (
        outcome.final_extraction["financial_series"][0]["period_range"]
        == "'Cash Flow'!B3:AB3"
    )


@pytest.mark.parametrize(
    ("period_range", "value_range", "primary_range", "axis_count", "extra_axis"),
    [
        ("2027-2053", "Forecast!C12:Z12", "A1:AC12", 27, False),
        ("2053-2027", "Forecast!C12:AC12", "A1:AC12", 27, False),
        ("2027-2053", "Forecast!C12:AC12", "A1:AC12", 0, False),
        ("2027-2053", "Forecast!C12:AC12", "A1:AC12", 27, True),
        ("2027-2053", "not-a-range", "A1:AC12", 27, False),
        ("2027-2053", "Inputs!C12:AC12", "A1:AC12", 27, False),
        ("2027-2053", "Forecast!C12:AC12", "A1:AC11", 27, False),
    ],
    ids=(
        "length-mismatch",
        "reversed-span",
        "zero-matches",
        "multiple-matches",
        "invalid-value-range",
        "cross-sheet-value-evidence",
        "evidence-outside-primary-range",
    ),
)
def test_integer_period_span_quarantines_when_recovery_is_not_unique_or_aligned(
    period_range,
    value_range,
    primary_range,
    axis_count,
    extra_axis,
):
    index = _index()
    if axis_count:
        _add_horizontal_axis(
            index,
            sheet="Forecast",
            header_row=10,
            value_row=12,
            first_column=3,
            start=2027,
            count=axis_count,
        )
    if extra_axis:
        _add_horizontal_axis(
            index,
            sheet="Forecast",
            header_row=9,
            value_row=11,
            first_column=3,
            start=2027,
            count=27,
        )
    partition = _partition("partition-boundary", "Forecast", primary_range)

    outcome = PartitionReconciler().reconcile(
        index,
        [_bound(partition, series=_series("Revenue", period_range, value_range))],
    )

    _rejected_series(outcome)


def test_fragment_merging_preserves_and_deduplicates_range_resolutions():
    index = _index()
    _add_horizontal_axis(
        index,
        sheet="Forecast",
        header_row=4,
        value_row=8,
        first_column=3,
        start=2027,
        count=4,
        add_values=False,
    )
    _add_horizontal_axis(
        index,
        sheet="Forecast",
        header_row=4,
        value_row=8,
        first_column=7,
        start=2031,
        count=4,
        add_values=False,
    )
    first = _partition("partition-left", "Forecast", "A1:F8")
    second = _partition("partition-right", "Forecast", "G1:J8")
    left = _series("Revenue", "2027-2030", "Forecast!C8:F8")
    right = _series("Revenue", "2031-2034", "Forecast!G8:J8")

    outcome = PartitionReconciler().reconcile(
        index,
        [
            _bound(first, series=left),
            _bound(second, series=right),
            _bound(second, series=right),
        ],
    )

    assert (
        outcome.final_extraction["financial_series"][0]["period_range"]
        == "Forecast!C4:J4"
    )
    assert outcome.final_extraction["range_resolutions"] == [{
        "field": "period_range",
        "submitted": "2027-2030",
        "resolved": "Forecast!C4:F4",
        "strategy": "unique_integer_span_match",
        "partition_id": "partition-left",
    }, {
        "field": "period_range",
        "submitted": "2031-2034",
        "resolved": "Forecast!G4:J4",
        "strategy": "unique_integer_span_match",
        "partition_id": "partition-right",
    }]
    assert outcome.final_extraction["financial_series"][0][
        "_backend_range_resolutions"
    ] == outcome.final_extraction["range_resolutions"]


def test_series_source_not_found_remains_terminal():
    index = _index()
    partition = _partition("partition-missing-series", "Forecast", "A1:J8")
    missing = _series(
        "Missing",
        "Forecast!K3:N3",
        "Forecast!K8:N8",
    )

    with pytest.raises(ReconciliationError) as exc:
        PartitionReconciler().reconcile(
            index,
            [_bound(partition, series=missing)],
        )

    assert exc.value.code == "series_source_not_found"


def test_horizontal_series_fragments_join_in_source_order():
    index = _index()
    partials = [
        _bound(
            _partition("partition-left", "Forecast", "A1:F8"),
            series=_series("Revenue", "Forecast!C3:F3", "Forecast!C8:F8"),
        ),
        _bound(
            _partition("partition-right", "Forecast", "G1:J8"),
            series=_series("Revenue", "Forecast!G3:J3", "Forecast!G8:J8"),
        ),
    ]

    outcome = PartitionReconciler().reconcile(index, partials)
    descriptor = outcome.final_extraction["financial_series"][0]

    assert descriptor["period_range"] == "Forecast!C3:J3"
    assert descriptor["value_range"] == "Forecast!C8:J8"
    assert descriptor["series_id"] == deterministic_series_id(
        index.workbook_version,
        "Forecast!C3:J3",
        "Forecast!C8:J8",
    )


def test_vertical_series_fragments_join_in_source_order():
    index = _index()
    partials = [
        _bound(
            _partition("partition-top", "Vertical", "A1:B3"),
            series=_series("Volume", "Vertical!A2:A3", "Vertical!B2:B3"),
        ),
        _bound(
            _partition("partition-bottom", "Vertical", "A4:B5"),
            series=_series("Volume", "Vertical!A4:A5", "Vertical!B4:B5"),
        ),
    ]

    descriptor = PartitionReconciler().reconcile(
        index,
        partials,
    ).final_extraction["financial_series"][0]

    assert descriptor["period_range"] == "Vertical!A2:A5"
    assert descriptor["value_range"] == "Vertical!B2:B5"
