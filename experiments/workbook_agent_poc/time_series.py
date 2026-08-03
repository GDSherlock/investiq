"""Backend-owned materialization of canonical financial time series.

The model identifies semantic descriptors and workbook ranges. This module is the sole source of
truth for canonical periods, values, source cells, and calculation telemetry.
"""

from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
from datetime import date, datetime
import math
import re
from typing import Any, Iterable

from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries

from workbook_tools import ToolError, WorkbookToolset


_A1_RANGE = re.compile(
    r"^(?:(?P<sheet>'(?:[^']|'')+'|[^!]+)!)?"
    r"(?P<range>\$?[A-Za-z]{1,3}\$?\d+(?::\$?[A-Za-z]{1,3}\$?\d+)?)$"
)
_ANNUAL = re.compile(r"^(?:FY\s*)?(?P<year>\d{2}|\d{4})(?P<state>[AEF])?$")
_QUARTER_FIRST = re.compile(
    r"^Q(?P<quarter>[1-4])\s*[-/]?\s*(?:FY\s*)?(?P<year>\d{2}|\d{4})(?P<state>[AEF])?$"
)
_YEAR_FIRST_QUARTER = re.compile(
    r"^(?:FY\s*)?(?P<year>\d{2}|\d{4})\s*[-/]?\s*Q(?P<quarter>[1-4])(?P<state>[AEF])?$"
)
_MONTH = re.compile(
    r"^(?P<month>JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)"
    r"[\s\-/']*(?P<year>\d{2}|\d{4})(?P<state>[AEF])?$"
)
_SEQUENCE_PERIOD = re.compile(
    r"^(?P<kind>CONSTRUCTION|OPERATING)(?:\s+(?:YEAR|PERIOD))?\s*(?P<number>\d+)$"
)
_MONTHS = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}
_REQUIRED_DESCRIPTOR_FIELDS = {
    "series_id", "label", "semantic_role", "category", "unit", "frequency",
    "period_range", "value_range",
}


class SeriesMaterializationError(Exception):
    """A structured, series-scoped failure safe for API reporting."""

    def __init__(self, code: str, message: str, **context: Any):
        self.code = code
        self.message = message
        self.context = context
        super().__init__(f"{code}: {message}")


@dataclass(frozen=True)
class RangeSpec:
    sheet_name: str
    cell_range: str
    qualified_range: str
    orientation: str
    coordinates: tuple[str, ...]
    bounds: tuple[int, int, int, int]

    @property
    def length(self) -> int:
        return len(self.coordinates)


def _sheet_for_output(sheet_name: str) -> str:
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_.]*", sheet_name):
        return sheet_name
    return "'" + sheet_name.replace("'", "''") + "'"


def _qualify(sheet_name: str, cell_reference: str) -> str:
    return f"{_sheet_for_output(sheet_name)}!{cell_reference}"


def _display_label(raw_value: Any) -> str | None:
    if raw_value is None:
        return None
    if isinstance(raw_value, (datetime, date)):
        return raw_value.isoformat()
    if isinstance(raw_value, float) and raw_value.is_integer():
        return str(int(raw_value))
    return str(raw_value)


def _four_digit_year(value: str) -> int | None:
    try:
        year = int(value)
    except (TypeError, ValueError):
        return None
    if len(value) == 2:
        return 2000 + year if year < 70 else 1900 + year
    if 1900 <= year <= 2200:
        return year
    return None


def _forecast_state(marker: str | None) -> bool | None:
    if marker == "A":
        return False
    if marker in {"E", "F"}:
        return True
    return None


def normalize_period(raw_value: Any, display_label: str | None = None) -> dict[str, Any]:
    """Normalize only safely recognizable period components while preserving the raw value."""
    normalized = {
        "raw_label": raw_value,
        "display_label": display_label if display_label is not None else _display_label(raw_value),
        "period_type": None,
        "year": None,
        "quarter": None,
        "month": None,
        "is_forecast": None,
    }
    if isinstance(raw_value, (datetime, date)):
        normalized.update(period_type="date", year=raw_value.year, month=raw_value.month)
        return normalized
    if isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool):
        if float(raw_value).is_integer() and 1900 <= int(raw_value) <= 2200:
            normalized.update(period_type="annual", year=int(raw_value))
        return normalized
    if not isinstance(raw_value, str):
        return normalized

    text = " ".join(raw_value.strip().upper().split())
    for pattern, period_type in (
        (_QUARTER_FIRST, "quarterly"),
        (_YEAR_FIRST_QUARTER, "quarterly"),
        (_MONTH, "monthly"),
        (_ANNUAL, "annual"),
    ):
        match = pattern.fullmatch(text)
        if not match:
            continue
        groups = match.groupdict()
        year = _four_digit_year(groups["year"])
        if year is None:
            return normalized
        normalized.update(
            period_type=period_type,
            year=year,
            quarter=int(groups["quarter"]) if groups.get("quarter") else None,
            month=_MONTHS.get(groups.get("month")),
            is_forecast=_forecast_state(groups.get("state")),
        )
        return normalized

    sequence = _SEQUENCE_PERIOD.fullmatch(text)
    if sequence:
        normalized["period_type"] = sequence.group("kind").casefold()
    return normalized


def _values_match(left: Any, right: Any) -> bool:
    if left is None and right is None:
        return True
    if (
        isinstance(left, (int, float))
        and isinstance(right, (int, float))
        and not isinstance(left, bool)
        and not isinstance(right, bool)
    ):
        return math.isclose(float(left), float(right), rel_tol=1e-9, abs_tol=1e-9)
    return left == right or str(left).strip() == str(right).strip()


def _legacy_axis_range(series: dict[str, Any], axis_name: str) -> str | None:
    direct_name = "period_range" if axis_name == "period_axis" else "value_range"
    direct = series.get(direct_name)
    if isinstance(direct, str) and direct.strip():
        return direct
    axis = series.get(axis_name)
    if isinstance(axis, dict) and isinstance(axis.get("source_range"), str):
        return axis["source_range"]
    for reference in series.get("source_references", []) or []:
        if reference.get("reference_type") != axis_name:
            continue
        sheet_name = reference.get("sheet_name")
        cell_range = reference.get("range")
        if sheet_name and cell_range:
            return _qualify(str(sheet_name), str(cell_range))
    return None


def _descriptor_from_series(series: dict[str, Any], *, input_source: str) -> dict[str, Any] | None:
    period_range = _legacy_axis_range(series, "period_axis")
    value_range = _legacy_axis_range(series, "value_axis")
    if not period_range or not value_range:
        return None
    descriptor = {
        key: deepcopy(value)
        for key, value in series.items()
        if key not in {"period_axis", "value_axis", "calculation_type", "formula_pattern", "source_references"}
    }
    descriptor["period_range"] = period_range
    descriptor["value_range"] = value_range
    descriptor["_input_source"] = input_source
    period_axis = series.get("period_axis") or {}
    value_axis = series.get("value_axis") or {}
    if isinstance(period_axis.get("periods"), list):
        descriptor["_legacy_periods"] = deepcopy(period_axis["periods"])
    if isinstance(value_axis.get("values"), list):
        descriptor["_legacy_values"] = deepcopy(value_axis["values"])
    if series.get("calculation_type") is not None:
        descriptor["_legacy_calculation_type"] = series.get("calculation_type")
    return descriptor


def is_compatible_financial_series_object(series: Any) -> bool:
    """Return whether an object supplies complete axis ranges suitable for materialization."""
    return (
        isinstance(series, dict)
        and _legacy_axis_range(series, "period_axis") is not None
        and _legacy_axis_range(series, "value_axis") is not None
    )


def normalize_financial_series_descriptors(
    extraction: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, int], list[dict[str, Any]]]:
    """Collect new descriptors and compatible legacy full-series objects only."""
    submitted = extraction.get("financial_series_descriptors")
    if not isinstance(submitted, list):
        submitted = extraction.get("financial_series", []) or []
    raw_descriptors = [deepcopy(item) for item in submitted if isinstance(item, dict)]
    descriptors: list[dict[str, Any]] = []
    for item in raw_descriptors:
        descriptor = _descriptor_from_series(item, input_source="financial_series")
        if descriptor is None:
            descriptor = deepcopy(item)
            descriptor["_input_source"] = "financial_series"
        descriptors.append(descriptor)

    legacy_detected = 0
    for item in extraction.get("financial_series_candidates", []) or []:
        if not isinstance(item, dict):
            continue
        descriptor = _descriptor_from_series(item, input_source="financial_series_candidates")
        if descriptor is None:
            continue
        legacy_detected += 1
        descriptors.append(descriptor)
    return (
        descriptors,
        {
            "submitted_descriptors": len(raw_descriptors),
            "legacy_series_detected": legacy_detected,
        },
        raw_descriptors,
    )


def _has_backend_period_range_resolution(descriptor: dict[str, Any]) -> bool:
    resolutions = descriptor.get("_backend_range_resolutions")
    if not isinstance(resolutions, list):
        return False
    return any(
        isinstance(resolution, dict)
        and resolution.get("field") == "period_range"
        and resolution.get("strategy") == "unique_integer_span_match"
        for resolution in resolutions
    )


class FinancialSeriesMaterializer:
    """Validate descriptors and build canonical workbook-owned series."""

    def __init__(self, tools: WorkbookToolset):
        self.tools = tools
        self._range_fact_cache: dict[str, list[dict[str, Any]]] = {}
        self.backend_range_reads = 0

    def _parse_range(
        self,
        reference: Any,
        *,
        explicit_sheet: str | None,
        field_name: str,
    ) -> RangeSpec:
        if not isinstance(reference, str) or not reference.strip():
            raise SeriesMaterializationError(
                "INVALID_RANGE", f"{field_name} must be a non-empty contiguous A1 range"
            )
        match = _A1_RANGE.fullmatch(reference.strip())
        if not match:
            raise SeriesMaterializationError(
                "INVALID_RANGE", f"{field_name} is not a valid contiguous A1 range"
            )
        sheet_token = match.group("sheet")
        if sheet_token:
            if sheet_token.startswith("'"):
                sheet_name = sheet_token[1:-1].replace("''", "'")
            else:
                sheet_name = sheet_token
        else:
            sheet_name = explicit_sheet
        if not sheet_name:
            raise SeriesMaterializationError(
                "INVALID_RANGE", f"{field_name} must include a sheet name"
            )

        cell_range = match.group("range").replace("$", "").upper()
        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
            max_sheet_row, max_sheet_col = self.tools.sheet_dimensions(sheet_name)
        except ToolError as exc:
            code = "SHEET_NOT_FOUND" if exc.code == "sheet_not_found" else "INVALID_RANGE"
            raise SeriesMaterializationError(code, exc.message, sheet_name=sheet_name) from exc
        except Exception as exc:
            raise SeriesMaterializationError(
                "INVALID_RANGE", f"{field_name} is not a valid contiguous A1 range"
            ) from exc
        if (
            min_row < 1 or min_col < 1
            or max_row > max_sheet_row or max_col > max_sheet_col
        ):
            raise SeriesMaterializationError(
                "INVALID_RANGE",
                f"{field_name} is outside the worksheet bounds",
                sheet_name=sheet_name,
            )
        rows = max_row - min_row + 1
        columns = max_col - min_col + 1
        if rows > 1 and columns > 1:
            raise SeriesMaterializationError(
                "TWO_DIMENSIONAL_RANGE", f"{field_name} must be one-dimensional"
            )
        orientation = "horizontal" if rows == 1 else "vertical"
        coordinates = tuple(
            f"{get_column_letter(column)}{row}"
            for row in range(min_row, max_row + 1)
            for column in range(min_col, max_col + 1)
        )
        normalized_range = (
            coordinates[0] if len(coordinates) == 1 else f"{coordinates[0]}:{coordinates[-1]}"
        )
        spec = RangeSpec(
            sheet_name=sheet_name,
            cell_range=normalized_range,
            qualified_range=_qualify(sheet_name, normalized_range),
            orientation=orientation,
            coordinates=coordinates,
            bounds=(min_col, min_row, max_col, max_row),
        )
        self._reject_merged_ambiguity(spec, field_name)
        return spec

    def _reject_merged_ambiguity(self, spec: RangeSpec, field_name: str) -> None:
        min_col, min_row, max_col, max_row = spec.bounds
        for merged_range in self.tools.merged_cell_ranges(spec.sheet_name):
            merged_bounds = range_boundaries(merged_range)
            merged_min_col, merged_min_row, merged_max_col, merged_max_row = merged_bounds
            intersects = not (
                max_col < merged_min_col or min_col > merged_max_col
                or max_row < merged_min_row or min_row > merged_max_row
            )
            if intersects and (merged_min_col != merged_max_col or merged_min_row != merged_max_row):
                raise SeriesMaterializationError(
                    "MERGED_CELL_AMBIGUITY",
                    f"{field_name} intersects merged range {merged_range}",
                    sheet_name=spec.sheet_name,
                )

    def _facts(self, spec: RangeSpec) -> list[dict[str, Any]]:
        if spec.qualified_range not in self._range_fact_cache:
            self.backend_range_reads += 1
            self._range_fact_cache[spec.qualified_range] = [
                self.tools.get_cell(spec.sheet_name, coordinate)
                for coordinate in spec.coordinates
            ]
        return deepcopy(self._range_fact_cache[spec.qualified_range])

    @staticmethod
    def _aligned(period: RangeSpec, value: RangeSpec) -> None:
        if period.orientation != value.orientation:
            raise SeriesMaterializationError(
                "ORIENTATION_MISMATCH", "period and value ranges use different orientations"
            )
        if period.length != value.length:
            raise SeriesMaterializationError(
                "PERIOD_VALUE_LENGTH_MISMATCH",
                "period and value ranges contain different numbers of cells",
                period_length=period.length,
                value_length=value.length,
            )
        if period.sheet_name != value.sheet_name:
            return
        p_min_col, p_min_row, p_max_col, p_max_row = period.bounds
        v_min_col, v_min_row, v_max_col, v_max_row = value.bounds
        if period.orientation == "horizontal" and (p_min_col, p_max_col) != (v_min_col, v_max_col):
            raise SeriesMaterializationError(
                "ORIENTATION_MISMATCH", "horizontal period and value ranges are shifted"
            )
        if period.orientation == "vertical" and (p_min_row, p_max_row) != (v_min_row, v_max_row):
            raise SeriesMaterializationError(
                "ORIENTATION_MISMATCH", "vertical period and value ranges are shifted"
            )

    def _representative_cell_only(self, period: RangeSpec, value: RangeSpec) -> bool:
        if period.length != 1 or value.length != 1 or period.sheet_name != value.sheet_name:
            return False
        p_col, p_row, _, _ = period.bounds
        v_col, v_row, _, _ = value.bounds
        max_row, max_col = self.tools.sheet_dimensions(period.sheet_name)
        if p_col == v_col and p_row != v_row:
            present = [
                self._cell_present(period.sheet_name, f"{get_column_letter(column)}{p_row}")
                for column in range(1, max_col + 1)
            ]
            index = p_col - 1
        elif p_row == v_row and p_col != v_col:
            present = [
                self._cell_present(period.sheet_name, f"{get_column_letter(p_col)}{row}")
                for row in range(1, max_row + 1)
            ]
            index = p_row - 1
        else:
            return False
        if not present[index]:
            return False
        start = index
        end = index
        while start > 0 and present[start - 1]:
            start -= 1
        while end + 1 < len(present) and present[end + 1]:
            end += 1
        return end - start + 1 > 1

    def _cell_present(self, sheet_name: str, coordinate: str) -> bool:
        fact = self.tools.get_cell(sheet_name, coordinate)
        return fact.get("raw_value") is not None or fact.get("formula") is not None

    @staticmethod
    def _formula_pattern(facts: list[dict[str, Any]]) -> dict[str, Any]:
        formulas = [fact for fact in facts if fact.get("formula") is not None]
        formula_count = len(formulas)
        blank_count = sum(
            fact.get("formula") is None and fact.get("raw_value") is None for fact in facts
        )
        static_count = len(facts) - formula_count - blank_count
        consistent: bool | None = None
        if formula_count > 1:
            try:
                normalized = {
                    Translator(fact["formula"], origin=fact["cell"]).translate_formula("A1")
                    for fact in formulas
                }
                consistent = len(normalized) == 1
            except Exception:
                consistent = None
        return {
            "is_formula_series": formula_count > 0,
            "formula_cell_count": formula_count,
            "static_cell_count": static_count,
            "blank_cell_count": blank_count,
            "pattern_consistent": consistent,
        }

    @staticmethod
    def _calculation_type(pattern: dict[str, Any]) -> str:
        formulas = pattern["formula_cell_count"]
        static = pattern["static_cell_count"]
        blanks = pattern["blank_cell_count"]
        if formulas and static:
            return "mixed"
        if formulas:
            return "formula"
        if static:
            return "hardcoded"
        if blanks:
            return "blank"
        return "unknown"

    @staticmethod
    def _failure(descriptor: dict[str, Any], error: SeriesMaterializationError) -> dict[str, Any]:
        return {
            "result_type": "financial_series",
            "series_id": descriptor.get("series_id"),
            "label": descriptor.get("label"),
            "semantic_role": descriptor.get("semantic_role", "financial_series"),
            "business_role": descriptor.get("business_role"),
            "materialization_status": "failed",
            "validation_status": "rejected",
            "calculation_type": "unknown",
            "formula_pattern": {
                "is_formula_series": False,
                "formula_cell_count": 0,
                "static_cell_count": 0,
                "blank_cell_count": 0,
                "pattern_consistent": None,
            },
            "warnings": [],
            "validation_warnings": [],
            "error_code": error.code,
            "rejection_reason": error.code.casefold(),
            "rejected_claims": [error.message],
            "error_context": error.context,
            "review_required": True,
            "number_of_periods": 0,
            "orientation": None,
            "validated_periods": [],
            "validated_values": [],
            "value_cell_metadata": [],
            "source_ranges": {
                "period_axis": descriptor.get("period_range"),
                "value_axis": descriptor.get("value_range"),
            },
        }

    def materialize(
        self,
        descriptor: dict[str, Any],
        *,
        period_range_recovered: bool = False,
    ) -> dict[str, Any]:
        missing = sorted(_REQUIRED_DESCRIPTOR_FIELDS - descriptor.keys())
        if missing:
            raise SeriesMaterializationError(
                "INVALID_SERIES_DESCRIPTOR",
                f"missing required descriptor fields: {', '.join(missing)}",
            )
        if descriptor.get("semantic_role") != "financial_series":
            raise SeriesMaterializationError(
                "INVALID_SERIES_DESCRIPTOR", "semantic_role must be financial_series"
            )
        if not str(descriptor.get("series_id") or "").strip() or not str(descriptor.get("label") or "").strip():
            raise SeriesMaterializationError(
                "INVALID_SERIES_DESCRIPTOR", "series_id and label must be non-empty"
            )

        explicit_sheet = descriptor.get("sheet_name")
        period_spec = self._parse_range(
            descriptor.get("period_range"),
            explicit_sheet=explicit_sheet,
            field_name="period_range",
        )
        value_spec = self._parse_range(
            descriptor.get("value_range"),
            explicit_sheet=explicit_sheet,
            field_name="value_range",
        )
        self._aligned(period_spec, value_spec)
        if self._representative_cell_only(period_spec, value_spec):
            raise SeriesMaterializationError(
                "REPRESENTATIVE_CELL_ONLY",
                "one point was submitted from an evidenced multi-period axis",
                period_length=period_spec.length,
                value_length=value_spec.length,
            )

        period_facts = self._facts(period_spec)
        value_facts = self._facts(value_spec)
        warnings: list[str] = []
        if period_range_recovered:
            warnings.append("PERIOD_RANGE_RESOLVED_FROM_WORKBOOK_EVIDENCE")
        period_points: list[dict[str, Any]] = []
        for index, (coordinate, fact) in enumerate(zip(period_spec.coordinates, period_facts)):
            raw_label = fact.get("raw_value")
            point = normalize_period(raw_label, _display_label(raw_label))
            point.update(index=index, source_cell=_qualify(period_spec.sheet_name, coordinate))
            period_points.append(point)
            if raw_label is not None and point["period_type"] is None:
                warnings.append("UNRECOGNIZED_PERIOD_LABEL")
            if raw_label is None:
                warnings.append("BLANK_PERIOD_LABEL")
        displayed = [point["display_label"] for point in period_points if point["display_label"] is not None]
        if len(displayed) != len(set(displayed)):
            warnings.append("DUPLICATE_PERIOD_LABEL")

        value_points = [
            {
                "index": index,
                "value": fact.get("raw_value"),
                "source_cell": _qualify(value_spec.sheet_name, coordinate),
                "formula": fact.get("formula"),
                "cached_value_available": (
                    fact.get("formula") is not None and fact.get("raw_value") is not None
                ),
                "cached_value_freshness": "unknown" if fact.get("formula") is not None else None,
                "formula_status": fact.get("formula_status"),
                "number_format": fact.get("number_format"),
                "data_type": fact.get("data_type"),
            }
            for index, (coordinate, fact) in enumerate(zip(value_spec.coordinates, value_facts))
        ]
        pattern = self._formula_pattern(value_facts)
        calculation_type = self._calculation_type(pattern)
        if pattern["formula_cell_count"]:
            warnings.append("CACHED_VALUE_FRESHNESS_UNKNOWN")
        if 0 < pattern["blank_cell_count"] < len(value_facts):
            warnings.append("PARTIALLY_BLANK_SERIES")
        if calculation_type == "mixed":
            warnings.append("MIXED_FORMULA_STATIC_SERIES")
        if self.tools.recalculation_signal().get("recalculation_warning") and pattern["is_formula_series"]:
            warnings.append("WORKBOOK_RECALCULATION_WARNING")

        legacy_periods = descriptor.get("_legacy_periods")
        actual_periods = [point["raw_label"] for point in period_points]
        if isinstance(legacy_periods, list) and (
            len(legacy_periods) != len(actual_periods)
            or any(not _values_match(left, right) for left, right in zip(legacy_periods, actual_periods))
        ):
            warnings.append("LEGACY_PERIOD_ARRAY_DISAGREEMENT")
        legacy_values = descriptor.get("_legacy_values")
        actual_values = [point["value"] for point in value_points]
        if isinstance(legacy_values, list) and (
            len(legacy_values) != len(actual_values)
            or any(not _values_match(left, right) for left, right in zip(legacy_values, actual_values))
        ):
            warnings.append("LEGACY_VALUE_ARRAY_DISAGREEMENT")
        legacy_calculation_type = descriptor.get("_legacy_calculation_type")
        if legacy_calculation_type not in (None, "unknown", calculation_type):
            warnings.append("LEGACY_CALCULATION_TYPE_DISAGREEMENT")

        label_reference = descriptor.get("label_reference")
        source_references = [
            {
                "sheet_name": period_spec.sheet_name,
                "range": period_spec.cell_range,
                "reference_type": "period_axis",
            },
            {
                "sheet_name": value_spec.sheet_name,
                "range": value_spec.cell_range,
                "reference_type": "value_axis",
            },
        ]
        if label_reference:
            try:
                label_spec = self._parse_range(
                    label_reference,
                    explicit_sheet=explicit_sheet,
                    field_name="label_reference",
                )
                if label_spec.length != 1:
                    raise SeriesMaterializationError(
                        "INVALID_RANGE", "label_reference must identify one cell"
                    )
                label_reference = label_spec.qualified_range
                source_references.append({
                    "sheet_name": label_spec.sheet_name,
                    "range": label_spec.cell_range,
                    "reference_type": "label",
                })
            except SeriesMaterializationError:
                warnings.append("INVALID_LABEL_REFERENCE")
                label_reference = None

        warnings = list(dict.fromkeys(warnings))
        materialization_status = "materialized_with_warning" if warnings else "materialized"
        validation_status = "validated_with_warning" if warnings else "validated"
        canonical = {
            key: deepcopy(descriptor.get(key))
            for key in (
                "series_id", "label", "semantic_role", "category", "unit", "frequency",
                "scenario", "entity", "currency", "reasoning_summary", "llm_confidence",
                "business_role",
            )
            if key in descriptor
        }
        canonical.update({
            "orientation": period_spec.orientation,
            "period_axis": {
                "source_range": period_spec.qualified_range,
                "periods": period_points,
            },
            "value_axis": {
                "source_range": value_spec.qualified_range,
                "values": value_points,
            },
            "label_reference": label_reference,
            "calculation_type": calculation_type,
            "formula_pattern": pattern,
            "source_references": source_references,
            "materialization_status": materialization_status,
            "validation_status": validation_status,
            "warnings": warnings,
            "aliases": [],
        })
        return canonical

    @staticmethod
    def _result(canonical: dict[str, Any]) -> dict[str, Any]:
        result = deepcopy(canonical)
        periods = canonical["period_axis"]["periods"]
        values = canonical["value_axis"]["values"]
        result.update({
            "result_type": "financial_series",
            "submitted_role": "financial_series",
            "validated_role": "financial_series",
            "number_of_periods": len(periods),
            "source_ranges": {
                "period_axis": canonical["period_axis"]["source_range"],
                "value_axis": canonical["value_axis"]["source_range"],
            },
            "validated_periods": [point["display_label"] for point in periods],
            "validated_values": [point["value"] for point in values],
            "value_cell_metadata": deepcopy(values),
            "validation_warnings": deepcopy(canonical["warnings"]),
            "rejected_claims": [],
            "rejection_reason": None,
            "error_code": None,
            "review_required": False,
        })
        return result

    @classmethod
    def _duplicate_result(
        cls,
        canonical: dict[str, Any],
        *,
        duplicate_of: str | None,
    ) -> dict[str, Any]:
        result = cls._result(canonical)
        result.update({
            "validation_status": "rejected",
            "error_code": "DUPLICATE_SERIES",
            "rejection_reason": "duplicate_series",
            "rejected_claims": [f"duplicate of canonical series {duplicate_of}"],
            "review_required": True,
            "duplicate_of": duplicate_of,
        })
        return result

    @staticmethod
    def _preference(canonical: dict[str, Any], index: int) -> tuple[int, int]:
        sheet = canonical["value_axis"]["source_range"].rsplit("!", 1)[0].casefold()
        summary = any(marker in sheet for marker in ("dashboard", "cover", "summary"))
        return (0 if summary else 1, -index)

    @staticmethod
    def _exact_key(canonical: dict[str, Any]) -> tuple[Any, ...]:
        return (
            canonical["period_axis"]["source_range"],
            canonical["value_axis"]["source_range"],
            canonical.get("scenario"),
            canonical.get("entity"),
            canonical.get("unit"),
            canonical.get("currency"),
        )

    @staticmethod
    def _semantic_copy_key(canonical: dict[str, Any]) -> tuple[Any, ...]:
        return (
            " ".join(str(canonical.get("label") or "").casefold().split()),
            canonical.get("category"),
            canonical.get("unit"),
            canonical.get("frequency"),
            canonical.get("scenario"),
            canonical.get("entity"),
            canonical.get("currency"),
            repr([point["raw_label"] for point in canonical["period_axis"]["periods"]]),
            repr([point["value"] for point in canonical["value_axis"]["values"]]),
        )

    @staticmethod
    def _label_context_key(canonical: dict[str, Any]) -> tuple[Any, ...]:
        return (
            " ".join(str(canonical.get("label") or "").casefold().split()),
            canonical.get("scenario"),
            canonical.get("entity"),
            canonical.get("unit"),
            canonical.get("currency"),
        )

    def materialize_collection(
        self,
        descriptors: Iterable[dict[str, Any]],
        *,
        input_counts: dict[str, int],
        trust_backend_range_resolutions: bool = False,
    ) -> dict[str, Any]:
        canonical_by_index: dict[int, dict[str, Any]] = {}
        results: list[dict[str, Any]] = []
        for index, descriptor in enumerate(descriptors):
            try:
                canonical = self.materialize(
                    descriptor,
                    period_range_recovered=(
                        trust_backend_range_resolutions
                        and _has_backend_period_range_resolution(descriptor)
                    ),
                )
            except SeriesMaterializationError as exc:
                results.append(self._failure(descriptor, exc))
                continue
            canonical_by_index[index] = canonical
            results.append(self._result(canonical))

        active = set(canonical_by_index)

        def deduplicate(key_function) -> None:
            groups: dict[tuple[Any, ...], list[int]] = {}
            for index in sorted(active):
                groups.setdefault(key_function(canonical_by_index[index]), []).append(index)
            for indexes in groups.values():
                if len(indexes) < 2:
                    continue
                winner = max(
                    indexes,
                    key=lambda item: self._preference(canonical_by_index[item], item),
                )
                for index in indexes:
                    if index == winner:
                        continue
                    loser = canonical_by_index[index]
                    winner_series = canonical_by_index[winner]
                    if loser.get("label") != winner_series.get("label"):
                        winner_series["aliases"].append(loser.get("label"))
                        winner_series["aliases"] = list(dict.fromkeys(winner_series["aliases"]))
                        results[winner] = self._result(winner_series)
                    results[index] = self._duplicate_result(
                        loser,
                        duplicate_of=winner_series.get("series_id"),
                    )
                    active.discard(index)

        deduplicate(self._exact_key)
        deduplicate(self._semantic_copy_key)

        label_groups: dict[tuple[Any, ...], list[int]] = {}
        for index in sorted(active):
            label_groups.setdefault(
                self._label_context_key(canonical_by_index[index]), []
            ).append(index)
        for indexes in label_groups.values():
            evidence = {
                self._exact_key(canonical_by_index[index])[:2] for index in indexes
            }
            if len(indexes) < 2 or len(evidence) < 2:
                continue
            for index in indexes:
                canonical = canonical_by_index[index]
                canonical["warnings"] = list(dict.fromkeys(
                    [*canonical["warnings"], "DUPLICATE_LABEL_DIFFERENT_RANGE"]
                ))
                canonical["materialization_status"] = "materialized_with_warning"
                canonical["validation_status"] = "validated_with_warning"
                results[index] = self._result(canonical)

        canonical_series = [canonical_by_index[index] for index in sorted(active)]
        rejected = [result for result in results if result.get("validation_status") == "rejected"]
        duplicate_count = sum(result.get("error_code") == "DUPLICATE_SERIES" for result in rejected)
        summary = {
            **input_counts,
            "submitted_series": (
                input_counts.get("submitted_descriptors", 0)
                + input_counts.get("legacy_series_detected", 0)
            ),
            "materialized_series": len(canonical_series),
            "validated_series": len(canonical_series),
            "validated_with_warning": sum(
                series.get("validation_status") == "validated_with_warning"
                for series in canonical_series
            ),
            "rejected_series": len(rejected),
            "representative_cell_only": sum(
                result.get("error_code") == "REPRESENTATIVE_CELL_ONLY" for result in rejected
            ),
            "period_value_mismatches": sum(
                result.get("error_code") == "PERIOD_VALUE_LENGTH_MISMATCH" for result in rejected
            ),
            "duplicate_series": duplicate_count,
            "backend_range_reads": self.backend_range_reads,
            "reclassified_series": 0,
        }
        return {
            "canonical_series": canonical_series,
            "validation_results": results,
            "summary": summary,
        }


def materialize_financial_series(
    tools: WorkbookToolset,
    extraction: dict[str, Any],
    *,
    trust_backend_range_resolutions: bool = False,
) -> dict[str, Any]:
    """Materialize canonical output in-place while preserving raw descriptor/legacy buckets."""
    descriptors, input_counts, raw_descriptors = normalize_financial_series_descriptors(extraction)
    extraction.setdefault("financial_series_descriptors", raw_descriptors)
    outcome = FinancialSeriesMaterializer(tools).materialize_collection(
        descriptors,
        input_counts=input_counts,
        trust_backend_range_resolutions=trust_backend_range_resolutions,
    )
    extraction["financial_series"] = deepcopy(outcome["canonical_series"])
    return outcome


def canonical_series_to_points(series: dict[str, Any]) -> list[dict[str, Any]]:
    """Convert one canonical series to chart-ready aligned records."""
    period_axis = series.get("period_axis")
    value_axis = series.get("value_axis")
    periods = period_axis.get("periods") if isinstance(period_axis, dict) else None
    values = value_axis.get("values") if isinstance(value_axis, dict) else None
    if not isinstance(periods, list) or not isinstance(values, list) or len(periods) != len(values):
        raise SeriesMaterializationError(
            "PERIOD_VALUE_LENGTH_MISMATCH", "canonical period and value axes are not aligned"
        )
    points: list[dict[str, Any]] = []
    for index, (period, value) in enumerate(zip(periods, values)):
        if not isinstance(period, dict) or not isinstance(value, dict):
            raise SeriesMaterializationError(
                "INVALID_SERIES_DESCRIPTOR", "canonical axis points must be objects"
            )
        points.append({
            "period_index": index,
            "period_label": period.get("display_label"),
            "year": period.get("year"),
            "quarter": period.get("quarter"),
            "month": period.get("month"),
            "is_forecast": period.get("is_forecast"),
            "value": value.get("value"),
            "source_cell": value.get("source_cell"),
        })
    return points


# Compatibility wrappers for existing local callers. Arrays are ignored as source of truth.
def validate_financial_series(
    tools: WorkbookToolset,
    series: dict[str, Any],
) -> dict[str, Any]:
    descriptor = _descriptor_from_series(series, input_source="financial_series") or deepcopy(series)
    outcome = FinancialSeriesMaterializer(tools).materialize_collection(
        [descriptor], input_counts={"submitted_descriptors": 1, "legacy_series_detected": 0}
    )
    return outcome["validation_results"][0]


def validate_financial_series_collection(
    tools: WorkbookToolset,
    series_collection: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    descriptors = [
        _descriptor_from_series(series, input_source="financial_series") or deepcopy(series)
        for series in series_collection
    ]
    outcome = FinancialSeriesMaterializer(tools).materialize_collection(
        descriptors,
        input_counts={
            "submitted_descriptors": len(descriptors),
            "legacy_series_detected": 0,
        },
    )
    return outcome["validation_results"]


__all__ = [
    "FinancialSeriesMaterializer",
    "SeriesMaterializationError",
    "canonical_series_to_points",
    "materialize_financial_series",
    "is_compatible_financial_series_object",
    "normalize_financial_series_descriptors",
    "normalize_period",
    "validate_financial_series",
    "validate_financial_series_collection",
]
