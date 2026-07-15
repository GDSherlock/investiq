"""
EXPERIMENTAL — isolated proof of concept. NOT wired into the production upload flow.

Controlled, read-only workbook inspection tools for the LLM function-calling loop.

Design invariants demonstrated here:
  * Every returned fact carries a full evidence envelope (sheet, cell, raw_value,
    displayed_value, formula, formula_status, data_type, number_format,
    source_reference, parse_warnings).
  * Missing formula cache values remain None. They are NEVER coerced to 0.
  * The toolset is bound to a single workbook loaded locally by the backend.
    The model only ever receives tool RESULTS, never a file handle or path.

This mirrors what production tools would do (openpyxl, loaded twice: values + formulas),
but stays in an isolated experiment directory.
"""

from __future__ import annotations

import io
import hashlib
import json
import secrets
from typing import Any

import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

from dependency import is_external_formula


# Bounded result sizes (production would enforce the same).
MAX_CELLS_PER_READ = 500
MAX_SEARCH_HITS = 100
MAX_OBSERVATION_PAYLOAD_BYTES = 12_000
DEFAULT_CHUNK_PAYLOAD_BYTES = 11_000


def _col_letter(idx: int) -> str:
    return get_column_letter(idx)


class WorkbookToolset:
    """Wraps one workbook and exposes controlled inspection tools.

    Loaded twice, exactly like the production ExcelParser:
      _wb_values   -> data_only=True  (cached values; formula cache may be None)
      _wb_formulas -> data_only=False (raw formula strings)
    """

    def __init__(self, *, file_bytes: bytes | None = None, file_path: str | None = None):
        if file_bytes is not None:
            source_bytes = file_bytes
        elif file_path is not None:
            with open(file_path, "rb") as source:
                source_bytes = source.read()
        else:
            raise ValueError("Provide file_bytes or file_path")
        self._wb_values = openpyxl.load_workbook(io.BytesIO(source_bytes), data_only=True)
        self._wb_formulas = openpyxl.load_workbook(io.BytesIO(source_bytes), data_only=False)
        self._workbook_version = hashlib.sha256(source_bytes).hexdigest()
        self._range_requests: dict[str, dict[str, Any]] = {}
        self._continuations: dict[str, tuple[str, int]] = {}
        # The exploration loop and deterministic validator share this toolset. Cache every
        # observed fact so validation reuses the full-context read instead of reopening or
        # reconstructing workbook ranges.
        self._fact_cache: dict[tuple[str, str], dict[str, Any]] = {}

    @property
    def workbook_version(self) -> str:
        return self._workbook_version

    # ---- helpers -------------------------------------------------------

    def _require_sheet(self, sheet_name: str):
        if sheet_name not in self._wb_values.sheetnames:
            raise ToolError(
                "sheet_not_found",
                f"Sheet {sheet_name!r} does not exist. Available: {self._wb_values.sheetnames}",
            )
        return self._wb_values[sheet_name], self._wb_formulas[sheet_name]

    def _cell_fact(self, sheet_name: str, coord: str) -> dict[str, Any]:
        """Build the full evidence envelope for a single cell. Never coerces None to 0."""
        ws_v, ws_f = self._require_sheet(sheet_name)
        try:
            row, col = coordinate_to_tuple(coord)
        except Exception:
            raise ToolError("bad_cell_reference", f"{coord!r} is not a valid A1 reference")
        normalized_coord = f"{_col_letter(col)}{row}"
        cache_key = (sheet_name, normalized_coord)
        if cache_key in self._fact_cache:
            return dict(self._fact_cache[cache_key])

        cell_v = ws_v.cell(row=row, column=col)
        cell_f = ws_f.cell(row=row, column=col)

        raw_formula = cell_f.value if isinstance(cell_f.value, str) and cell_f.value.startswith("=") else None
        warnings: list[str] = []
        is_external_ref = False
        is_error = False

        if raw_formula is not None:
            cached = cell_v.value  # may legitimately be None -> keep it None
            is_external_ref = is_external_formula(raw_formula)
            is_error = isinstance(cached, str) and cached.startswith("#")
            if is_external_ref:
                # An external reference's value is not resolvable locally. Never invent it.
                formula_status = "formula_external"
                raw_value = None
                warnings.append("external_reference_value_unavailable")
            elif is_error:
                # A cached Excel error (#N/A, #REF!, ...) is NOT a value.
                formula_status = "formula_error"
                raw_value = None
                warnings.append(f"formula_error_cache:{cached}")
            elif cached is None:
                formula_status = "formula_no_cache"
                raw_value = None
                warnings.append("formula_cache_missing_value_left_null")
            else:
                formula_status = "formula_with_cached_value"
                raw_value = cached
        else:
            formula_status = "static_value"
            raw_value = cell_v.value

        # openpyxl cannot render the number-format-applied display string reliably.
        displayed_value = None
        if raw_value is not None:
            warnings.append("displayed_value_unavailable_via_openpyxl")

        fact = {
            "sheet_name": sheet_name,
            "cell": normalized_coord,
            "source_reference": f"{sheet_name}!{normalized_coord}",
            "raw_value": raw_value,
            "displayed_value": displayed_value,
            "formula": raw_formula,
            # static_value | formula_with_cached_value | formula_no_cache | formula_error | formula_external
            "formula_status": formula_status,
            "is_external_ref": is_external_ref,
            "is_error": is_error,
            "data_type": cell_v.data_type,             # openpyxl: n/s/f/d/b/e
            "python_type": type(raw_value).__name__ if raw_value is not None else "NoneType",
            "number_format": cell_v.number_format,
            "parse_warnings": warnings,
        }
        self._fact_cache[cache_key] = fact
        return dict(fact)

    # ---- tools (these map 1:1 to LLM function schemas) -----------------

    def list_sheets(self) -> dict[str, Any]:
        out = []
        for ws in self._wb_values.worksheets:
            required_range = f"A1:{_col_letter(ws.max_column)}{ws.max_row}"
            out.append({
                "name": ws.title,
                "state": ws.sheet_state,      # visible | hidden | veryHidden
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "required_range": required_range,
            })
        return {"sheets": out, "count": len(out)}

    def inspect_sheet(self, sheet_name: str, sample_rows: int = 8) -> dict[str, Any]:
        ws_v, ws_f = self._require_sheet(sheet_name)
        n_formulas = sum(
            1 for row in ws_f.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        )
        # small textual preview of the top-left region (bounded)
        preview = []
        for r in range(1, min(ws_v.max_row, sample_rows) + 1):
            cells = []
            for cidx in range(1, min(ws_v.max_column, 8) + 1):
                v = ws_v.cell(row=r, column=cidx).value
                if v is not None:
                    cells.append(f"{_col_letter(cidx)}{r}={v!r}")
            if cells:
                preview.append(" ".join(cells))
        return {
            "sheet_name": sheet_name,
            "state": ws_v.sheet_state,
            "dimensions": ws_v.dimensions,
            "max_row": ws_v.max_row,
            "max_col": ws_v.max_column,
            "formula_cell_count": n_formulas,
            "merged_ranges": [str(m) for m in ws_v.merged_cells.ranges],
            "preview": preview,
        }

    @staticmethod
    def _json_bytes(payload: Any) -> bytes:
        return json.dumps(
            payload,
            default=str,
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8")

    @classmethod
    def _with_serialized_bytes(cls, payload: dict[str, Any]) -> dict[str, Any]:
        payload = dict(payload)
        payload["serialized_bytes"] = 0
        for _ in range(8):
            size = len(cls._json_bytes(payload))
            if size == payload["serialized_bytes"]:
                return payload
            payload["serialized_bytes"] = size
        return payload

    def _facts_in_rect(
        self,
        sheet_name: str,
        min_col: int,
        min_row: int,
        max_col: int,
        max_row: int,
        facts: dict[tuple[int, int], dict[str, Any]],
    ) -> list[dict[str, Any]]:
        return [
            facts[(row, col)]
            for row in range(min_row, max_row + 1)
            for col in range(min_col, max_col + 1)
            if (row, col) in facts
        ]

    def _range_text(self, min_col: int, min_row: int, max_col: int, max_row: int) -> str:
        return f"{_col_letter(min_col)}{min_row}:{_col_letter(max_col)}{max_row}"

    def _chunk_fits(
        self,
        *,
        sheet_name: str,
        requested_range: str,
        returned_range: str,
        cells: list[dict[str, Any]],
        range_cell_count: int,
        max_serialized_bytes: int,
    ) -> bool:
        if range_cell_count > MAX_CELLS_PER_READ:
            return False
        # Conservative final-envelope placeholders ensure sizing includes every field,
        # including the self-reported byte count and a full continuation token.
        candidate = self._with_serialized_bytes({
            "request_id": "r" * 43,
            "chunk_id": f"{sheet_name}:{returned_range}:" + "c" * 43 + ":999999/999999",
            "chunk_index": 999999,
            "chunk_count": 999999,
            "sheet_name": sheet_name,
            "requested_range": requested_range,
            "returned_range": returned_range,
            "cell_count": len(cells),
            "range_cell_count": range_cell_count,
            "workbook_version": self.workbook_version,
            "is_complete": True,
            "has_more": True,
            "next_range": requested_range,
            "continuation_token": "t" * 43,
            "cells": cells,
        })
        return candidate["serialized_bytes"] <= max_serialized_bytes

    def _partition_range(
        self,
        sheet_name: str,
        requested_range: str,
        bounds: tuple[int, int, int, int],
        max_serialized_bytes: int,
    ) -> list[tuple[str, list[dict[str, Any]], int]]:
        min_col, min_row, max_col, max_row = bounds
        facts: dict[tuple[int, int], dict[str, Any]] = {}
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                coord = f"{_col_letter(col)}{row}"
                fact = self._cell_fact(sheet_name, coord)
                if fact["raw_value"] is not None or fact["formula"] is not None:
                    facts[(row, col)] = fact

        partitions: list[tuple[str, list[dict[str, Any]], int]] = []

        def add_rect(c1: int, r1: int, c2: int, r2: int) -> None:
            returned = self._range_text(c1, r1, c2, r2)
            cells = self._facts_in_rect(sheet_name, c1, r1, c2, r2, facts)
            count = (r2 - r1 + 1) * (c2 - c1 + 1)
            partitions.append((returned, cells, count))

        row_start = min_row
        row = min_row
        while row <= max_row:
            returned = self._range_text(min_col, row_start, max_col, row)
            cells = self._facts_in_rect(sheet_name, min_col, row_start, max_col, row, facts)
            count = (row - row_start + 1) * (max_col - min_col + 1)
            if self._chunk_fits(
                sheet_name=sheet_name,
                requested_range=requested_range,
                returned_range=returned,
                cells=cells,
                range_cell_count=count,
                max_serialized_bytes=max_serialized_bytes,
            ):
                row += 1
                continue

            if row > row_start:
                add_rect(min_col, row_start, max_col, row - 1)
                row_start = row
                continue

            # One full row is too large. Preserve the row boundary and split it
            # into adjacent column windows.
            col_start = min_col
            col = min_col
            while col <= max_col:
                returned = self._range_text(col_start, row, col, row)
                cells = self._facts_in_rect(sheet_name, col_start, row, col, row, facts)
                count = col - col_start + 1
                if self._chunk_fits(
                    sheet_name=sheet_name,
                    requested_range=requested_range,
                    returned_range=returned,
                    cells=cells,
                    range_cell_count=count,
                    max_serialized_bytes=max_serialized_bytes,
                ):
                    col += 1
                    continue
                if col == col_start:
                    raise ToolError(
                        "payload_too_large",
                        f"Cell {_col_letter(col)}{row} cannot fit in a complete "
                        f"{max_serialized_bytes}-byte observation payload.",
                    )
                add_rect(col_start, row, col - 1, row)
                col_start = col
            if col_start <= max_col:
                add_rect(col_start, row, max_col, row)
            row += 1
            row_start = row

        if row_start <= max_row:
            add_rect(min_col, row_start, max_col, max_row)
        return partitions

    def read_range(
        self,
        sheet_name: str,
        cell_range: str,
        continuation_token: str | None = None,
        *,
        max_serialized_bytes: int = DEFAULT_CHUNK_PAYLOAD_BYTES,
    ) -> dict[str, Any]:
        self._require_sheet(sheet_name)
        requested_range = cell_range.upper()
        if continuation_token is not None:
            state_pointer = self._continuations.get(continuation_token)
            if state_pointer is None:
                raise ToolError("invalid_continuation_token", "Unknown continuation token.")
            request_id, chunk_index = state_pointer
            state = self._range_requests[request_id]
            if state["workbook_version"] != self.workbook_version:
                raise ToolError(
                    "workbook_version_changed",
                    "The workbook changed after this range request was created.",
                )
            if (
                state["sheet_name"] != sheet_name
                or state["requested_range"] != requested_range
                or state["max_serialized_bytes"] != max_serialized_bytes
            ):
                raise ToolError(
                    "continuation_binding_mismatch",
                    "Continuation token does not belong to this workbook, sheet, range, and budget.",
                )
            return state["chunks"][chunk_index]

        try:
            min_c, min_r, max_c, max_r = range_boundaries(requested_range)
        except Exception:
            raise ToolError("bad_range", f"{cell_range!r} is not a valid A1 range")
        if not 1_000 <= max_serialized_bytes <= MAX_OBSERVATION_PAYLOAD_BYTES:
            raise ToolError(
                "bad_payload_budget",
                f"max_serialized_bytes must be between 1000 and {MAX_OBSERVATION_PAYLOAD_BYTES}.",
            )

        request_id = secrets.token_urlsafe(32)
        partitions = self._partition_range(
            sheet_name,
            requested_range,
            (min_c, min_r, max_c, max_r),
            max_serialized_bytes,
        )
        continuation_tokens = [secrets.token_urlsafe(32) for _ in partitions[:-1]]
        chunks: list[dict[str, Any]] = []
        chunk_count = len(partitions)
        for index, (returned_range, cells, range_cell_count) in enumerate(partitions):
            has_more = index + 1 < chunk_count
            next_range = partitions[index + 1][0] if has_more else None
            token = continuation_tokens[index] if has_more else None
            chunk = self._with_serialized_bytes({
                "request_id": request_id,
                "chunk_id": (
                    f"{sheet_name}:{returned_range}:{request_id[:12]}:"
                    f"{index + 1}/{chunk_count}"
                ),
                "chunk_index": index,
                "chunk_count": chunk_count,
                "sheet_name": sheet_name,
                "requested_range": requested_range,
                "returned_range": returned_range,
                "cell_count": len(cells),
                "range_cell_count": range_cell_count,
                "workbook_version": self.workbook_version,
                "is_complete": True,
                "has_more": has_more,
                "next_range": next_range,
                "continuation_token": token,
                "cells": cells,
            })
            if chunk["serialized_bytes"] > max_serialized_bytes:
                raise ToolError(
                    "payload_too_large",
                    "Final chunk envelope exceeded its safe serialized-byte budget.",
                )
            chunks.append(chunk)

        state = {
            "request_id": request_id,
            "sheet_name": sheet_name,
            "requested_range": requested_range,
            "workbook_version": self.workbook_version,
            "max_serialized_bytes": max_serialized_bytes,
            "chunks": chunks,
        }
        self._range_requests[request_id] = state
        for index, token in enumerate(continuation_tokens, start=1):
            self._continuations[token] = (request_id, index)
        return chunks[0]

    def get_cell(self, sheet_name: str, cell_reference: str) -> dict[str, Any]:
        return self._cell_fact(sheet_name, cell_reference)

    def search_cells(self, query: str, *, in_formulas: bool = False) -> dict[str, Any]:
        q = str(query).lower()
        hits = []
        book = self._wb_formulas if in_formulas else self._wb_values
        for ws in book.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.value is None:
                        continue
                    if q in str(c.value).lower():
                        hits.append({
                            "sheet_name": ws.title,
                            "cell": c.coordinate,
                            "source_reference": f"{ws.title}!{c.coordinate}",
                            "match": str(c.value)[:120],
                        })
                        if len(hits) >= MAX_SEARCH_HITS:
                            return {"query": query, "truncated": True, "hits": hits}
        return {"query": query, "truncated": False, "hit_count": len(hits), "hits": hits}

    def get_workbook_metadata(self) -> dict[str, Any]:
        sheets = [
            {
                "name": ws.title,
                "state": ws.sheet_state,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "required_range": f"A1:{_col_letter(ws.max_column)}{ws.max_row}",
            }
            for ws in self._wb_values.worksheets
        ]
        named = [{"name": n, "target": t} for n, t in self.defined_names().items()]
        external = []
        for sheet, coord, f in self.iter_formulas():
            if is_external_formula(f):
                external.append({"source_reference": f"{sheet}!{coord}", "formula": f})
        return {
            "sheet_count": len(sheets),
            "hidden_sheet_count": sum(1 for s in sheets if s["state"] != "visible"),
            "sheets": sheets,
            "required_sheet_ranges": {
                sheet["name"]: sheet["required_range"] for sheet in sheets
            },
            "named_ranges": named,
            "external_links": external,
        }

    def get_named_ranges(self) -> dict[str, Any]:
        return {"named_ranges": [{"name": n, "target": t} for n, t in self.defined_names().items()]}

    def get_data_validations(self, sheet_name: str) -> list[dict[str, Any]]:
        ws_v, _ = self._require_sheet(sheet_name)
        out = []
        for dv in ws_v.data_validations.dataValidation:
            out.append({
                "range": str(dv.sqref),
                "type": dv.type,
                "formula1": dv.formula1,
                "allow_blank": dv.allow_blank,
            })
        return out

    def get_formulas(self, sheet_name: str, cell_range: str | None = None) -> dict[str, Any]:
        _, ws_f = self._require_sheet(sheet_name)
        formulas: dict[str, str] = {}
        if cell_range:
            try:
                min_c, min_r, max_c, max_r = range_boundaries(cell_range)
            except Exception:
                raise ToolError("bad_range", f"{cell_range!r} is not a valid A1 range")
            for r in range(min_r, max_r + 1):
                for cidx in range(min_c, max_c + 1):
                    cell = ws_f.cell(row=r, column=cidx)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas[cell.coordinate] = cell.value
        else:
            for row in ws_f.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas[cell.coordinate] = cell.value
        return {"sheet_name": sheet_name, "formula_count": len(formulas), "formulas": formulas}

    def submit_extraction_result(self, result: dict[str, Any]) -> dict[str, Any]:
        """Terminal tool. The loop stops here; validation runs against the workbook.
        NOTE: the coverage GATE lives in the agent loop, not here — this only acks receipt."""
        n = sum(len(result.get(k, [])) for k in (
            "all_assumption_candidates", "parameter_candidates", "derived_value_candidates",
            "output_candidates", "financial_series_candidates", "financial_series"))
        return {"received": True, "candidate_count": n}

    # ---- introspection helpers used by the dependency graph + role classifier ----

    def content_sheets(self) -> set[str]:
        """Sheet names that contain at least one non-empty cell (incl. hidden)."""
        out: set[str] = set()
        for ws in self._wb_values.worksheets:
            for row in ws.iter_rows():
                if any(c.value is not None for c in row):
                    out.add(ws.title)
                    break
        return out

    def sheet_dimensions(self, sheet_name: str) -> tuple[int, int]:
        """Return (max_row, max_column) for deterministic range-bound checks."""
        ws_v, _ = self._require_sheet(sheet_name)
        return ws_v.max_row, ws_v.max_column

    def recalculation_signal(self) -> dict[str, Any]:
        """Expose workbook recalc flags without claiming cached-value freshness."""
        calc = getattr(self._wb_formulas, "calculation", None)
        if calc is None:
            return {"recalculation_warning": False, "cached_value_freshness": "unknown"}
        warning = any(
            getattr(calc, attribute, None) is True
            for attribute in ("fullCalcOnLoad", "forceFullCalc", "calcOnSave")
        )
        return {
            "recalculation_warning": warning,
            "cached_value_freshness": "unknown",
            "calc_mode": getattr(calc, "calcMode", None),
            "full_calc_on_load": getattr(calc, "fullCalcOnLoad", None),
            "force_full_calc": getattr(calc, "forceFullCalc", None),
        }

    def merged_cell_ranges(self, sheet_name: str) -> list[str]:
        """Return merged ranges from the already-loaded workbook without a model-visible read."""
        _, ws_f = self._require_sheet(sheet_name)
        return [str(cell_range) for cell_range in ws_f.merged_cells.ranges]

    def non_empty_cell_references(self, sheet_name: str) -> set[str]:
        """Actual non-empty cells, including formulas whose cached value is absent."""
        _, ws_f = self._require_sheet(sheet_name)
        return {
            cell.coordinate
            for row in ws_f.iter_rows()
            for cell in row
            if cell.value is not None
        }

    def defined_names(self) -> dict[str, str]:
        out: dict[str, str] = {}
        collection = self._wb_formulas.defined_names
        if hasattr(collection, "items"):
            entries = collection.items()
        else:  # openpyxl 3.0 exposes DefinedNameList.definedName instead.
            entries = ((dn.name, dn) for dn in collection.definedName)
        for name, dn in entries:
            try:
                out[name] = dn.value
            except Exception:
                out[name] = str(dn)
        return out

    def iter_formulas(self):
        """Yield (sheet, coord, formula) for every formula cell across ALL sheets (incl. hidden)."""
        for ws in self._wb_formulas.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        yield (ws.title, c.coordinate, c.value)

    def neighbor_label(self, sheet_name: str, coord: str) -> dict[str, Any]:
        """Nearest non-empty text cell to the left in the same row (the row label)."""
        ws_v, _ = self._require_sheet(sheet_name)
        row, col = coordinate_to_tuple(coord)
        for c in range(col - 1, 0, -1):
            v = ws_v.cell(row=row, column=c).value
            if isinstance(v, str) and v.strip():
                return {"label": v.strip(), "label_cell": f"{_col_letter(c)}{row}"}
        return {"label": None, "label_cell": None}

    def sheet_formula_ratio(self, sheet_name: str) -> float:
        """Fraction of value cells (numeric or formula) on the sheet that are formulas.
        A results/calc sheet is formula-dominated; an input sheet is not."""
        _, ws_f = self._require_sheet(sheet_name)
        formulas = 0
        numerics = 0
        for row in ws_f.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    formulas += 1
                elif isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                    numerics += 1
        denom = formulas + numerics
        return formulas / denom if denom else 0.0

    def data_validation_cells(self, sheet_name: str) -> set[str]:
        """Set of 'Sheet!Cell' references governed by a list data-validation dropdown."""
        ws_v, _ = self._require_sheet(sheet_name)
        refs: set[str] = set()
        for dv in ws_v.data_validations.dataValidation:
            for token in str(dv.sqref).split():
                try:
                    min_c, min_r, max_c, max_r = range_boundaries(token)
                except Exception:
                    continue
                for r in range(min_r, max_r + 1):
                    for cidx in range(min_c, max_c + 1):
                        refs.add(f"{sheet_name}!{_col_letter(cidx)}{r}")
        return refs


class ToolError(Exception):
    """Structured tool error returned to the model as data (never crashes the loop)."""

    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(f"{code}: {message}")

    def as_result(self) -> dict[str, Any]:
        return {"error": {"code": self.code, "message": self.message}}
