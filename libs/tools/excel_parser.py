"""
Excel Parser Tool — extracts formulas, values, and structured data from financial models.

Uses openpyxl; idempotent, returns structured JSON, 30s timeout.
"""

import io
from typing import Any

import openpyxl


class ExcelParser:
    """Parse uploaded Excel financial model files into structured data."""

    # Map canonical sheet names to common aliases
    _SHEET_ALIASES: dict[str, list[str]] = {
        "Cover": ["Cover", "README", "Summary", "Overview", "Title"],
        "Assumptions": ["Assumptions", "Inputs", "Parameters", "Params"],
        "Revenue": ["Revenue", "Operations", "Revenues", "Income"],
        "Capex": ["Capex", "CapEx", "Capital_Expenditure", "CAPEX"],
        "PnL": ["PnL", "Income_Statement", "P&L", "Profit_Loss", "IS"],
        "Debt_Schedule": ["Debt_Schedule", "Financing", "Debt", "Loan"],
        "CashFlows": ["CashFlows", "Cash_Flow", "Cash_Flows", "CF", "Cashflow"],
        "Returns": ["Returns", "DCF", "Valuation", "IRR"],
        "Sensitivity": ["Sensitivity", "Sensitivities", "Scenario"],
        "Checks": ["Checks", "Audit", "Validation", "Model_Checks"],
        "Dashboard": ["Dashboard", "Summary_Dashboard", "Exec_Summary"],
    }

    def __init__(self, file_bytes: bytes | None = None, file_path: str | None = None):
        if file_bytes:
            self._wb_data = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            self._wb_formulas = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
        elif file_path:
            self._wb_data = openpyxl.load_workbook(file_path, data_only=True)
            self._wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        else:
            raise ValueError("Provide either file_bytes or file_path")

    def _resolve_sheet(self, canonical: str) -> str | None:
        """Resolve a canonical sheet name to the actual sheet name in the workbook."""
        actual_names = self._wb_data.sheetnames
        for alias in self._SHEET_ALIASES.get(canonical, [canonical]):
            if alias in actual_names:
                return alias
        # Case-insensitive fallback
        lower_map = {s.lower(): s for s in actual_names}
        for alias in self._SHEET_ALIASES.get(canonical, [canonical]):
            if alias.lower() in lower_map:
                return lower_map[alias.lower()]
        return None

    @property
    def sheet_names(self) -> list[str]:
        return self._wb_data.sheetnames

    def parse_all(self) -> dict[str, Any]:
        """Parse entire workbook into structured JSON."""
        return {
            "sheets": self.sheet_names,
            "cover": self._parse_cover(),
            "assumptions": self._parse_assumptions(),
            "revenue": self._parse_time_series("Revenue"),
            "capex": self._parse_time_series("Capex"),
            "pnl": self._parse_time_series("PnL"),
            "debt_schedule": self._parse_time_series("Debt_Schedule"),
            "cash_flows": self._parse_time_series("CashFlows"),
            "returns": self._parse_returns(),
            "sensitivity": self._parse_sensitivity(),
            "checks": self._parse_checks(),
            "dashboard": self._parse_dashboard(),
        }

    def _parse_cover(self) -> dict[str, Any]:
        """Extract project metadata from Cover sheet."""
        name = self._resolve_sheet("Cover")
        if not name:
            return {}
        ws = self._wb_data[name]
        cover = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            label_cell = None
            value_cell = None
            for cell in row:
                if cell.column == 2 and cell.value:  # Column B
                    label_cell = str(cell.value).strip()
                if cell.column == 4 and cell.value:  # Column D
                    value_cell = cell.value
            if label_cell and value_cell:
                cover[label_cell] = value_cell
        return cover

    def _parse_assumptions(self) -> list[dict[str, Any]]:
        """Extract all assumptions as structured records."""
        name = self._resolve_sheet("Assumptions")
        if not name:
            return []
        ws = self._wb_data[name]
        assumptions = []
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
            cells = {c.column: c.value for c in row}
            # Column B=row_id, C=name, D=value, E=unit, F=source, G=scenario
            row_id = cells.get(2)
            name = cells.get(3)
            value = cells.get(4)

            if name and not str(name).startswith("──") and value is not None:
                assumptions.append({
                    "row_id": str(row_id) if row_id else None,
                    "name": str(name),
                    "value": value,
                    "unit": str(cells.get(5, "")) if cells.get(5) else None,
                    "source": str(cells.get(6, "")) if cells.get(6) else None,
                    "scenario": str(cells.get(7, "")) if cells.get(7) else "Base",
                })

        return assumptions

    def _parse_time_series(self, sheet_name: str) -> dict[str, Any]:
        """Parse a time-series sheet (Revenue, Capex, PnL, etc.)."""
        resolved = self._resolve_sheet(sheet_name)
        if not resolved:
            return {"error": f"Sheet '{sheet_name}' not found"}

        ws = self._wb_data[resolved]

        # Row 3 has headers (years)
        years = []
        for cell in ws[3]:
            if cell.column >= 3 and cell.value:
                years.append(str(cell.value))

        # Data rows
        line_items = {}
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
            label = None
            values = []
            for cell in row:
                if cell.column == 2 and cell.value:
                    label = str(cell.value).strip()
                if cell.column >= 3:
                    values.append(cell.value if cell.value is not None else 0)

            if label and not label.startswith("──"):
                line_items[label] = values[:len(years)]

        return {
            "title": ws.cell(row=2, column=2).value,
            "years": years,
            "data": line_items,
        }

    def _parse_returns(self) -> dict[str, Any]:
        """Parse Returns summary sheet."""
        name = self._resolve_sheet("Returns")
        if not name:
            return {"metrics": []}
        ws = self._wb_data[name]
        metrics = []
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
            cells = {c.column: c.value for c in row}
            name = cells.get(2)
            if name and not str(name).startswith("──"):
                metrics.append({
                    "metric": str(name),
                    "base_case": cells.get(4),
                    "stress_case": cells.get(5),
                    "upside_case": cells.get(6),
                    "hurdle": cells.get(7),
                    "status": cells.get(8),
                })
        return {"metrics": metrics}

    def _parse_sensitivity(self) -> dict[str, Any]:
        """Parse Sensitivity analysis sheet."""
        name = self._resolve_sheet("Sensitivity")
        if not name:
            return {"one_way": [], "two_way": {}}
        ws = self._wb_data[name]

        # One-way sensitivity (rows 5-13)
        one_way = []
        for row in ws.iter_rows(min_row=6, max_row=13, values_only=False):
            cells = {c.column: c.value for c in row}
            name = cells.get(2)
            if name:
                one_way.append({
                    "assumption": str(name),
                    "stress_minus_20": cells.get(3),
                    "stress_minus_10": cells.get(4),
                    "base_case": cells.get(5),
                    "upside_plus_10": cells.get(6),
                    "upside_plus_20": cells.get(7),
                    "irr_range": cells.get(8),
                    "key_variable": cells.get(9),
                })

        # Two-way sensitivity (rows 16+)
        two_way_columns = []
        for cell in ws[17]:
            if cell.column >= 3 and cell.value is not None:
                two_way_columns.append(cell.value)

        two_way = []
        for row in ws.iter_rows(min_row=18, max_row=25, values_only=False):
            cells = {c.column: c.value for c in row}
            wacc_label = cells.get(2)
            if wacc_label:
                two_way.append({
                    "wacc": str(wacc_label),
                    "values": [cells.get(c) for c in range(3, 10)],
                })

        return {
            "one_way": one_way,
            "two_way": {
                "row_var": "WACC",
                "col_var": "Throughput Fee",
                "columns": two_way_columns,
                "data": two_way,
            },
        }

    def _parse_checks(self) -> list[dict[str, Any]]:
        """Parse model integrity checks."""
        name = self._resolve_sheet("Checks")
        if not name:
            return []
        ws = self._wb_data[name]
        checks = []
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
            cells = {c.column: c.value for c in row}
            desc = cells.get(2)
            if desc and not str(desc).startswith("──"):
                checks.append({
                    "description": str(desc),
                    "value": cells.get(4),
                    "expected": cells.get(5),
                    "status": cells.get(6),
                })
        return checks

    def _parse_dashboard(self) -> dict[str, Any]:
        """Parse executive dashboard."""
        name = self._resolve_sheet("Dashboard")
        if not name:
            return {}
        ws = self._wb_data[name]
        data = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            for cell in row:
                if cell.value is not None:
                    data[cell.coordinate] = cell.value
        return data

    def get_formulas(self, sheet_name: str) -> dict[str, str]:
        """Extract formulas (not computed values) from a sheet."""
        ws = self._wb_formulas[sheet_name]
        formulas = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas[cell.coordinate] = cell.value
        return formulas

    def health_check(self) -> dict[str, Any]:
        """Run basic health checks on the uploaded model."""
        issues = []
        score = 100

        # Check required sheets
        required = ["Assumptions", "Revenue", "Capex", "PnL", "CashFlows", "Returns"]
        for sheet in required:
            if not self._resolve_sheet(sheet):
                issues.append(f"Missing required sheet: {sheet}")
                score -= 15

        # Check assumptions have values
        assumptions = self._parse_assumptions() if "Assumptions" in self.sheet_names else []
        if len(assumptions) < 10:
            issues.append(f"Only {len(assumptions)} assumptions found (expected 30+)")
            score -= 10

        # Check for empty data
        for sheet in ["Revenue", "PnL"]:
            if sheet in self.sheet_names:
                ts = self._parse_time_series(sheet)
                if not ts.get("data"):
                    issues.append(f"Sheet '{sheet}' has no data rows")
                    score -= 10

        return {
            "health_score": max(0, score),
            "sheets_found": self.sheet_names,
            "assumptions_count": len(assumptions),
            "issues": issues,
            "status": "HEALTHY" if score >= 80 else "DEGRADED" if score >= 50 else "UNHEALTHY",
        }
