from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import re
from zipfile import ZipFile

from openpyxl import Workbook

from apps.api.app.model_extraction_models import (
    CanonicalOutput,
    FinancialSeries,
    FinancialSeriesValue,
    ModelParameter,
    ModelVersion,
)
from apps.api.app.model_extraction_repository import WorkbookVersionRepository
from apps.api.app.model_extraction_types import FinancialEntityIdFactory, new_uuid
from apps.api.app.workbook_storage import DatabaseWorkbookStorage


_CALCULATION_PROPERTIES_PATTERN = re.compile(
    rb"<calcPr\b[^>]*/>|<calcPr\b[^>]*>.*?</calcPr>",
    re.DOTALL,
)


def without_calculation_properties(content_bytes: bytes) -> bytes:
    source_buffer = BytesIO(content_bytes)
    output_buffer = BytesIO()
    with ZipFile(source_buffer, "r") as source, ZipFile(output_buffer, "w") as output:
        removed = 0
        for member in source.infolist():
            payload = source.read(member.filename)
            if member.filename == "xl/workbook.xml":
                payload, removed = _CALCULATION_PROPERTIES_PATTERN.subn(b"", payload)
            output.writestr(member, payload)
    if removed != 1:
        raise AssertionError("Expected one workbook calculation-properties element")
    return output_buffer.getvalue()


def calculation_workbook_bytes(
    *,
    include_calculation_properties: bool = True,
    include_kpi_formulas: bool = False,
    include_equity_cash_flow: bool = False,
) -> bytes:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = 2
    inputs["A2"] = 3

    calc = workbook.create_sheet("Calc")
    calc["A1"] = 2026
    calc["A2"] = 2027
    calc["B1"] = "=SUM(Inputs!A1:A2)"
    calc["B2"] = "=B1*2"
    calc["B3"] = '=COUNTIF(Inputs!A1:A2,">0")'
    calc["B4"] = "=B3+1"
    calc["B5"] = "=B6+1"
    calc["B6"] = "=B5+1"
    calc["B7"] = "='[rates.xlsx]Inputs'!A1+1"
    calc["B8"] = "=IF(TRUE,B2,1/0)"
    if include_equity_cash_flow:
        calc["A3"] = 2028
        calc["D1"] = "=0-100"
        calc["D2"] = "=Inputs!A1*20"
        calc["D3"] = "=80"
    if include_kpi_formulas:
        inputs["C1"] = -100
        inputs["C2"] = 110
        inputs["D1"] = 60
        inputs["D2"] = 60
        inputs["E1"] = 1.4
        inputs["E2"] = 1.2
        inputs["E3"] = 0
        inputs["F1"] = 5
        inputs["G1"] = 100
        inputs["G2"] = 250
        calc["C1"] = "=IRR(Inputs!C1:C2)"
        calc["C2"] = "=NPV(10%,Inputs!D1:D2)"
        calc["C3"] = '=MINIFS(Inputs!E1:E3,Inputs!E1:E3,">0")'
        calc["C4"] = "=Inputs!F1"
        calc["C5"] = "=Inputs!G2/Inputs!G1"

    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["C1"] = "=Inputs!A2+1"
    very_hidden = workbook.create_sheet("Very Hidden")
    very_hidden.sheet_state = "veryHidden"
    very_hidden["D1"] = "=Hidden!C1+1"

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    content_bytes = buffer.getvalue()
    if not include_calculation_properties:
        return without_calculation_properties(content_bytes)
    return content_bytes


def create_materialized_rule_model(
    session,
    *,
    include_calculation_properties: bool = True,
    include_kpi_formulas: bool = False,
    include_equity_cash_flow: bool = False,
):
    storage = DatabaseWorkbookStorage(session)
    workbook = WorkbookVersionRepository(session, storage).get_or_create(
        calculation_workbook_bytes(
            include_calculation_properties=include_calculation_properties,
            include_kpi_formulas=include_kpi_formulas,
            include_equity_cash_flow=include_equity_cash_flow,
        ),
        "calculation-rules.xlsx",
    )
    model = ModelVersion(
        id=new_uuid(),
        workbook_version_id=workbook.id,
        upload_filename="calculation-rules.xlsx",
        status="materialized",
        validation_status="validated",
        submitted=True,
        extraction_snapshot_json={
            "final_extraction": {
                "parameter_candidates": [
                    {
                        "label": "Snapshot-only helper",
                        "source_reference": "Inputs!A2",
                    }
                ]
            }
        },
        created_at=datetime.now(timezone.utc),
        completed_at=datetime.now(timezone.utc),
    )
    id_factory = FinancialEntityIdFactory(model.id)
    input_parameter = ModelParameter(
        id=id_factory.parameter_id("Inputs", "A1"),
        model_version_id=model.id,
        entity_kind="parameter",
        source_bucket="parameter_candidates",
        label="Volume",
        submitted_role="hardcoded_input",
        validated_role="hardcoded_input",
        raw_value_json=2,
        validated_value_json=2,
        source_sheet="Inputs",
        source_cell="A1",
        formula_status="static_value",
        source_validation_status="validated",
        role_validation_status="validated",
        validation_status="validated",
        data_type="n",
        number_format="General",
    )
    series_id = id_factory.series_id(
        "Calc!A1",
        "Calc!B2",
        "base",
        "Project",
        "USD",
        "USD",
    )
    series = FinancialSeries(
        id=series_id,
        model_version_id=model.id,
        entity_kind="financial_series",
        label="Calculated output",
        semantic_role="financial_series",
        business_role="revenue",
        unit="USD",
        frequency="annual",
        orientation="horizontal",
        scenario="base",
        entity="Project",
        currency="USD",
        calculation_type="formula",
        period_source_range="Calc!A1",
        value_source_range="Calc!B2",
        materialization_status="materialized",
        validation_status="validated",
    )
    value = FinancialSeriesValue(
        id=id_factory.value_id(series_id, 0),
        financial_series_id=series_id,
        period_index=0,
        raw_period_label_json=2026,
        display_period_label="2026",
        period_type="annual",
        year=2026,
        is_forecast=True,
        value_json=None,
        period_source_sheet="Calc",
        period_source_cell="A1",
        value_source_sheet="Calc",
        value_source_cell="B2",
        exact_formula="=B1*2",
        formula_status="formula_no_cache",
        cached_value_available=False,
        cached_value_freshness="missing",
        number_format="General",
        data_type="f",
    )
    output = CanonicalOutput(
        id=id_factory.output_id("Calc", "B1"),
        model_version_id=model.id,
        entity_kind="canonical_output",
        llm_candidate_alias="total-project-cost",
        label="Total project cost",
        category="summary",
        canonical_name="Total Project Cost",
        business_role="total_project_cost",
        submitted_role="formula_output",
        validated_role="formula_output",
        raw_value_json=None,
        unit="USD",
        scenario="base",
        source_sheet="Calc",
        source_cell="B1",
        exact_formula="=SUM(Inputs!A1:A2)",
        formula_status="formula_no_cache",
        source_validation_status="validated",
        role_validation_status="validated",
        validation_status="validated",
        data_type="f",
        number_format="General",
        validation_warnings_json=[],
    )
    persisted_entities = [workbook, model, input_parameter, output, series, value]
    if include_equity_cash_flow:
        equity_series_id = id_factory.series_id(
            "Calc!A1:A3",
            "Calc!D1:D3",
            "base",
            "Equity",
            "USD",
            "USD",
        )
        equity_series = FinancialSeries(
            id=equity_series_id,
            model_version_id=model.id,
            entity_kind="financial_series",
            label="Equity cash flow",
            semantic_role="financial_series",
            business_role="equity_cash_flow",
            unit="USD",
            frequency="annual",
            orientation="vertical",
            scenario="base",
            entity="Equity",
            currency="USD",
            calculation_type="formula",
            period_source_range="Calc!A1:A3",
            value_source_range="Calc!D1:D3",
            materialization_status="materialized",
            validation_status="validated",
        )
        equity_formulas = ("=0-100", "=Inputs!A1*20", "=80")
        equity_values = [
            FinancialSeriesValue(
                id=id_factory.value_id(equity_series_id, index),
                financial_series_id=equity_series_id,
                period_index=index,
                raw_period_label_json=2026 + index,
                display_period_label=str(2026 + index),
                period_type="annual",
                year=2026 + index,
                is_forecast=True,
                value_json=None,
                period_source_sheet="Calc",
                period_source_cell=f"A{index + 1}",
                value_source_sheet="Calc",
                value_source_cell=f"D{index + 1}",
                exact_formula=equity_formulas[index],
                formula_status="formula_no_cache",
                cached_value_available=False,
                cached_value_freshness="missing",
                number_format="General",
                data_type="f",
            )
            for index in range(3)
        ]
        persisted_entities.extend([equity_series, *equity_values])
    session.add_all(persisted_entities)
    session.commit()
    return storage, workbook, model, input_parameter, series, value
