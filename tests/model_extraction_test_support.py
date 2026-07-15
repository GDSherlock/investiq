from __future__ import annotations

from collections.abc import Iterator
from contextlib import contextmanager
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from sqlalchemy import Engine, create_engine, event
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool


def new_id() -> str:
    return str(uuid4())


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def sample_workbook_bytes(value: float = 1.0) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Assumptions"
    worksheet["A1"] = "Input"
    worksheet["B1"] = value
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def persistence_workbook_bytes() -> bytes:
    workbook = Workbook()
    assumptions = workbook.active
    assumptions.title = "Assumptions"
    assumptions["A1"] = "Project Alpha"
    assumptions["A4"] = "Tax rate"
    assumptions["B4"] = 0.2
    assumptions["B4"].number_format = "0.0%"

    calculations = workbook.create_sheet("Calculations")
    calculations["A4"] = "Discounted value"
    calculations["B4"] = "=Assumptions!B4*100"
    calculations["B4"].number_format = "0.00"

    income_statement = workbook.create_sheet("P&L")
    income_statement["A2"] = "Period"
    income_statement["B2"] = 2026
    income_statement["C2"] = 2027
    income_statement["A3"] = "Revenue"
    income_statement["B3"] = 100.0
    income_statement["C3"] = "=B3*1.1"
    income_statement["B3"].number_format = "#,##0.0"
    income_statement["C3"].number_format = "#,##0.0"

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def deterministic_extraction_result(
    *,
    submitted: bool = True,
    conflicting_parameter: bool = False,
) -> dict[str, object]:
    if not submitted:
        return {
            "endpoint_mode": "experimental_workbook_agent_validation",
            "filename": "model.xlsx",
            "runtime_seconds": 0.01,
            "driver_meta": {"api": "responses", "deployment": "deterministic"},
            "submitted": False,
            "stop_reason": "model_returned_no_tool_call",
            "coverage": {"total_sheets": 3, "inspected_sheets": 0},
            "final_extraction": {
                "all_assumption_candidates": [],
                "output_candidates": [],
                "financial_series": [],
            },
            "validation_summary": {"candidate_count": 0},
            "time_series_summary": {"materialized_series": 0},
            "validation_results": [],
            "warnings": [],
            "errors": [{"code": "AGENT_INCOMPLETE"}],
            "trace": [],
            "trace_truncated": False,
        }

    parameter_candidate = {
        "candidate_id": "tax-rate",
        "original_label": "Tax rate",
        "submitted_role": "hardcoded_input",
        "raw_value": 0.2,
        "unit": "%",
        "period": 2026,
        "scenario": "base",
        "source_references": [{"sheet_name": "Assumptions", "cell": "B4"}],
        "reasoning_summary": "Explicit workbook input",
        "llm_confidence": 0.9,
        "category": "tax",
        "canonical_name": "tax_rate",
    }
    derived_candidate = {
        "candidate_id": "discounted-value",
        "original_label": "Discounted value",
        "submitted_role": "formula_derived_value",
        "raw_value": None,
        "unit": "USDm",
        "period": 2026,
        "scenario": "base",
        "source_references": [{"sheet_name": "Calculations", "cell": "B4"}],
        "reasoning_summary": "Formula-derived intermediate",
        "llm_confidence": 0.85,
        "category": "calculation",
        "canonical_name": "discounted_value",
    }
    output_candidate = {
        "candidate_id": "display-output",
        "original_label": "Displayed revenue",
        "submitted_role": "hardcoded_display_output",
        "raw_value": 100.0,
        "source_references": [{"sheet_name": "P&L", "cell": "B3"}],
        "reasoning_summary": "Output only",
        "llm_confidence": 0.8,
    }
    metadata_candidate = {
        "candidate_id": "project-name",
        "original_label": "Project Alpha",
        "submitted_role": "metadata",
        "raw_value": "Project Alpha",
        "source_references": [{"sheet_name": "Assumptions", "cell": "A1"}],
        "reasoning_summary": "Workbook title",
        "llm_confidence": 0.95,
    }
    review_candidates = []
    validation_results: list[dict[str, object]] = [
        {
            "candidate_id": "tax-rate",
            "_bucket": "parameter_candidates",
            "source_reference": "Assumptions!B4",
            "source_validation_status": "validated",
            "submitted_role": "hardcoded_input",
            "validated_role": "hardcoded_input",
            "role_validation_status": "validated",
            "validation_status": "validated",
            "validated_value": 0.2,
            "formula_status": "static_value",
            "data_type": "n",
            "number_format": "0.0%",
            "validation_confidence": 0.95,
            "validation_warnings": [],
            "dependency_evidence": {"dependents": ["Calculations!B4"]},
        },
        {
            "candidate_id": "discounted-value",
            "_bucket": "derived_value_candidates",
            "source_reference": "Calculations!B4",
            "source_validation_status": "validated_null",
            "submitted_role": "formula_derived_value",
            "validated_role": "formula_derived_value",
            "role_validation_status": "validated",
            "validation_status": "validated_null",
            "validated_value": None,
            "formula_status": "formula_no_cache",
            "data_type": "n",
            "number_format": "0.00",
            "validation_confidence": 0.6,
            "validation_warnings": ["formula_cache_missing"],
            "dependency_evidence": {"precedents": ["Assumptions!B4"]},
        },
        {
            "candidate_id": "display-output",
            "_bucket": "output_candidates",
            "source_reference": "P&L!B3",
            "source_validation_status": "validated",
            "submitted_role": "hardcoded_display_output",
            "validated_role": "hardcoded_display_output",
            "role_validation_status": "validated_deferred",
            "validation_status": "validated",
            "validated_value": 100.0,
            "formula_status": "static_value",
            "validation_confidence": 0.9,
            "validation_warnings": [],
            "dependency_evidence": {},
        },
        {
            "series_id": "llm-revenue",
            "_bucket": "financial_series",
            "validation_status": "validated_with_warning",
            "validation_warnings": ["CACHED_VALUE_FRESHNESS_UNKNOWN"],
            "dependency_evidence": {"series": "not-a-persisted-contract"},
        },
    ]
    if conflicting_parameter:
        review_candidates.append(
            {
                **parameter_candidate,
                "candidate_id": "conflicting-tax-rate",
                "submitted_role": "scenario_selector",
            }
        )
        validation_results.append(
            {
                "candidate_id": "conflicting-tax-rate",
                "_bucket": "review_candidates",
                "source_reference": "Assumptions!B4",
                "source_validation_status": "validated",
                "submitted_role": "scenario_selector",
                "validated_role": "scenario_selector",
                "role_validation_status": "validated",
                "validation_status": "validated",
                "validated_value": 0.2,
                "formula_status": "static_value",
                "validation_confidence": 0.8,
                "validation_warnings": [],
                "dependency_evidence": {},
            }
        )

    canonical_series = {
        "series_id": "llm-revenue",
        "label": "Revenue",
        "semantic_role": "financial_series",
        "category": "income_statement",
        "unit": "USDm",
        "frequency": "annual",
        "scenario": "base",
        "entity": "Project Alpha",
        "currency": "USD",
        "reasoning_summary": "Complete revenue row",
        "llm_confidence": 0.92,
        "orientation": "horizontal",
        "period_axis": {
            "source_range": "'P&L'!B2:C2",
            "periods": [
                {
                    "index": 0,
                    "raw_label": 2026,
                    "display_label": "2026",
                    "period_type": "annual",
                    "year": 2026,
                    "quarter": None,
                    "month": None,
                    "is_forecast": True,
                    "source_cell": "'P&L'!B2",
                },
                {
                    "index": 1,
                    "raw_label": 2027,
                    "display_label": "2027",
                    "period_type": "annual",
                    "year": 2027,
                    "quarter": None,
                    "month": None,
                    "is_forecast": True,
                    "source_cell": "'P&L'!C2",
                },
            ],
        },
        "value_axis": {
            "source_range": "'P&L'!B3:C3",
            "values": [
                {
                    "index": 0,
                    "value": 100.0,
                    "source_cell": "'P&L'!B3",
                    "formula": None,
                    "cached_value_available": False,
                    "cached_value_freshness": None,
                    "formula_status": "static_value",
                    "number_format": "#,##0.0",
                    "data_type": "n",
                },
                {
                    "index": 1,
                    "value": None,
                    "source_cell": "'P&L'!C3",
                    "formula": "=B3*1.1",
                    "cached_value_available": False,
                    "cached_value_freshness": "unknown",
                    "formula_status": "formula_no_cache",
                    "number_format": "#,##0.0",
                    "data_type": "n",
                },
            ],
        },
        "label_reference": "'P&L'!A3",
        "calculation_type": "mixed",
        "formula_pattern": {"formula_cell_count": 1, "static_cell_count": 1},
        "materialization_status": "materialized_with_warning",
        "validation_status": "validated_with_warning",
        "warnings": ["CACHED_VALUE_FRESHNESS_UNKNOWN"],
        "aliases": ["Sales"],
    }
    return {
        "endpoint_mode": "experimental_workbook_agent_validation",
        "filename": "model.xlsx",
        "runtime_seconds": 0.02,
        "driver_meta": {"api": "responses", "deployment": "deterministic"},
        "submitted": True,
        "stop_reason": "submitted",
        "coverage": {"total_sheets": 3, "inspected_sheets": 3},
        "final_extraction": {
            "metadata": [metadata_candidate],
            "all_assumption_candidates": [],
            "parameter_candidates": [parameter_candidate],
            "derived_value_candidates": [derived_candidate],
            "output_candidates": [output_candidate],
            "unclassified_inputs": [],
            "review_candidates": review_candidates,
            "financial_series": [canonical_series],
        },
        "validation_summary": {
            "candidate_count": len(validation_results),
            "validated": 2,
            "validated_null": 1,
            "validated_with_warning": 1,
        },
        "time_series_summary": {
            "materialized_series": 1,
            "validated_with_warning": 1,
        },
        "validation_results": validation_results,
        "warnings": [{"code": "CANDIDATE_VALIDATION_WARNING"}],
        "errors": [],
        "trace": [{"event": "deterministic"}],
        "trace_truncated": False,
    }


def create_sqlite_session_factory(
    database_url: str = "sqlite+pysqlite:///:memory:",
) -> tuple[Engine, sessionmaker[Session]]:
    engine_kwargs: dict[str, object] = {
        "connect_args": {"check_same_thread": False},
    }
    if database_url.endswith(":memory:"):
        engine_kwargs["poolclass"] = StaticPool

    engine = create_engine(database_url, **engine_kwargs)

    @event.listens_for(engine, "connect")
    def _enable_foreign_keys(dbapi_connection, _connection_record) -> None:
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    return engine, sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)


@contextmanager
def sqlite_session() -> Iterator[Session]:
    from apps.api.app.database import Base
    from apps.api.app import model_extraction_models  # noqa: F401

    engine, session_factory = create_sqlite_session_factory()
    Base.metadata.create_all(engine)
    session = session_factory()
    try:
        yield session
    finally:
        session.close()
        engine.dispose()


def sqlite_file_url(path: Path) -> str:
    return f"sqlite+pysqlite:///{path}"
