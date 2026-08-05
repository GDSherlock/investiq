from __future__ import annotations

from io import BytesIO
from pathlib import Path
import uuid

from openpyxl import Workbook
import pytest

from apps.api.app.calculation_rules.compiler import FormulaCompiler
from apps.api.app.calculation_rules.evaluator import SafeCalculationEvaluator, ScalarValue
from apps.api.app.calculation_rules.graph import CalculationGraphBuilder
from apps.api.app.calculation_rules.inventory import WorkbookFormulaInventory
from apps.api.app.calculation_rules.types import CalculationRuleExtractionConfiguration


def _phase2_contracts():
    from apps.api.app.calculation_rules.phase2_registry import (
        PHASE2_FUNCTION_REGISTRY,
    )
    from apps.api.app.calculation_rules.phase2_types import (
        Phase2CalculationConfiguration,
    )

    return Phase2CalculationConfiguration(), PHASE2_FUNCTION_REGISTRY


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = 1
    inputs["A2"] = 2
    inputs["A3"] = 3
    inputs["A4"] = "text"
    inputs["A5"] = None
    calc = workbook.create_sheet("Calc")
    calc["B1"] = '=COUNTIF(Inputs!A1:A5,">=2")'
    calc["B2"] = "=COUNT(Inputs!A1:A5)"
    calc["B3"] = "=COUNTA(Inputs!A1:A5)"
    calc["B4"] = "=SUM(Inputs!A1:A3)"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _compile(configuration, registry):
    catalog = WorkbookFormulaInventory(configuration).scan(
        _workbook_bytes(),
        str(uuid.uuid4()),
    )
    compiler = FormulaCompiler(configuration, function_registry=registry)
    compilations = tuple(compiler.compile(cell, catalog) for cell in catalog.formulas)
    return catalog, compilations


def _compile_formula(formula: str):
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = -100
    inputs["A2"] = 110
    inputs["B1"] = 0
    inputs["B2"] = 0
    calc = workbook.create_sheet("Calc")
    calc["A1"] = formula
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    configuration, registry = _phase2_contracts()
    catalog = WorkbookFormulaInventory(configuration).scan(
        buffer.getvalue(),
        str(uuid.uuid4()),
    )
    return FormulaCompiler(
        configuration,
        function_registry=registry,
    ).compile(catalog.formulas[0], catalog)


def test_v2_compiler_emits_additive_envelope_and_v1_remains_exact() -> None:
    v1_configuration = CalculationRuleExtractionConfiguration()
    v1_catalog = WorkbookFormulaInventory(v1_configuration).scan(
        _workbook_bytes(),
        str(uuid.uuid4()),
    )
    v1 = FormulaCompiler(v1_configuration).compile(
        next(cell for cell in v1_catalog.formulas if cell.ref.cell_address == "B4"),
        v1_catalog,
    )

    configuration, registry = _phase2_contracts()
    _catalog, compilations = _compile(configuration, registry)
    v2 = next(
        item
        for item in compilations
        if item.formula_cell_id
        == next(
            cell.id
            for cell in _catalog.formulas
            if cell.ref.cell_address == "B1"
        )
    )

    assert set(v1.ir_json or {}) == {
        "expression_id",
        "formula_cell_id",
        "ir_version",
        "compiler_version",
        "semantics_profile",
        "formula_sha256",
        "normalized_signature",
        "root",
    }
    assert v2.support_status == "supported"
    assert v2.ir_json is not None
    assert v2.ir_json["ir_version"] == "calc-ir-v2"
    assert v2.ir_json["required_registry_version"] == "calc-functions-v4"
    assert v2.ir_json["capabilities"] == ["conditional-aggregation"]
    assert v2.ir_json["limits"]["node_count"] > 0
    assert v2.ir_json["limits"]["max_depth"] > 0


def test_phase2_registry_is_closed_versioned_and_additive() -> None:
    configuration, registry = _phase2_contracts()

    assert configuration.ir_version == "calc-ir-v2"
    assert configuration.compiler_version == "formula-compiler-v3"
    assert configuration.engine_version == "calc-engine-v4"
    assert configuration.function_registry_version == "calc-functions-v4"
    assert configuration.semantics_profile == "excel-compatible-kpi-v1"
    assert set(registry) == {
        "SUM",
        "AVERAGE",
        "MIN",
        "MAX",
        "ABS",
        "ROUND",
        "IF",
        "COUNT",
        "COUNTA",
        "COUNTIF",
        "IFERROR",
        "AND",
        "OR",
        "MINIFS",
        "IRR",
        "NPV",
        "MOD",
        "YEAR",
        "MATCH",
        "XNPV",
        "XIRR",
    }
    assert registry["COUNTIF"].minimum_arguments == 2
    assert registry["COUNTIF"].maximum_arguments == 2
    assert registry["COUNTIF"].implementation_version == "countif-v2"
    assert registry["IFERROR"].lazy is True
    assert registry["MINIFS"].minimum_arguments == 3
    assert registry["MINIFS"].maximum_arguments == 253
    assert registry["IRR"].maximum_arguments == 2
    assert registry["NPV"].maximum_arguments == 255


@pytest.mark.parametrize(
    ("formula", "capabilities"),
    [
        ("=IFERROR(1/0,0)", ["error-handling"]),
        ("=AND(TRUE,FALSE)", ["logical"]),
        (
            '=MINIFS(Inputs!A1:A2,Inputs!A1:A2,">0")',
            ["conditional-aggregation"],
        ),
        ("=IRR(Inputs!A1:A2)", ["financial"]),
        ("=NPV(10%,Inputs!A1:A2)", ["financial"]),
        (
            "=IRR(Inputs!A1:A2+Inputs!B1:B2)",
            ["financial", "range-arithmetic"],
        ),
    ],
)
def test_kpi_function_ir_records_required_capabilities(
    formula: str,
    capabilities: list[str],
) -> None:
    compilation = _compile_formula(formula)

    assert compilation.support_status == "supported"
    assert compilation.ir_json is not None
    assert compilation.ir_json["capabilities"] == capabilities


def test_v2_count_functions_execute_typed_range_semantics() -> None:
    configuration, registry = _phase2_contracts()
    catalog, compilations = _compile(configuration, registry)
    graph = CalculationGraphBuilder(configuration).build(catalog, compilations)

    results = SafeCalculationEvaluator(function_registry=registry).execute(
        graph,
        catalog,
        compilations,
        configuration,
    )
    by_address = {reference.display: execution for reference, execution in results.items()}

    assert by_address["Calc!B1"].value == ScalarValue.number(2)
    assert by_address["Calc!B2"].value == ScalarValue.number(3)
    assert by_address["Calc!B3"].value == ScalarValue.number(4)
    assert by_address["Calc!B4"].value == ScalarValue.number(6)
    assert all(item.status == "executed" for item in by_address.values())


def test_v2_countif_numeric_comparison_ignores_blank_range_cells() -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = 1
    inputs["A2"] = None
    inputs["A3"] = 2
    calc = workbook.create_sheet("Calc")
    calc["A1"] = '=COUNTIF(Inputs!A1:A3,"<1.5")'
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    configuration, registry = _phase2_contracts()
    catalog = WorkbookFormulaInventory(configuration).scan(
        buffer.getvalue(),
        str(uuid.uuid4()),
    )
    compilations = tuple(
        FormulaCompiler(configuration, function_registry=registry).compile(cell, catalog)
        for cell in catalog.formulas
    )
    graph = CalculationGraphBuilder(configuration).build(catalog, compilations)
    execution = next(
        iter(
            SafeCalculationEvaluator(function_registry=registry).execute(
                graph,
                catalog,
                compilations,
                configuration,
            ).values()
        )
    )

    assert execution.value == ScalarValue.number(1)


def test_v2_countif_rejects_unsupported_wildcard_without_guessing() -> None:
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "alpha"
    calc = workbook.create_sheet("Calc")
    calc["A1"] = '=COUNTIF(Inputs!A1,"a*")'
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    configuration, registry = _phase2_contracts()
    catalog = WorkbookFormulaInventory(configuration).scan(
        buffer.getvalue(),
        str(uuid.uuid4()),
    )
    compilations = tuple(
        FormulaCompiler(configuration, function_registry=registry).compile(cell, catalog)
        for cell in catalog.formulas
    )
    graph = CalculationGraphBuilder(configuration).build(catalog, compilations)
    results = SafeCalculationEvaluator(function_registry=registry).execute(
        graph,
        catalog,
        compilations,
        configuration,
    )

    execution = next(iter(results.values()))
    assert execution.status == "execution_error"
    assert execution.value == ScalarValue.error("#VALUE!")


def test_repository_financial_workbook_is_fully_executable_under_v2() -> None:
    workbook_path = Path(__file__).parents[1] / "Financial_Model_Data.xlsx"
    configuration, registry = _phase2_contracts()
    catalog = WorkbookFormulaInventory(configuration).scan(
        workbook_path.read_bytes(),
        str(uuid.uuid4()),
    )
    compilations = tuple(
        FormulaCompiler(configuration, function_registry=registry).compile(cell, catalog)
        for cell in catalog.formulas
    )
    graph = CalculationGraphBuilder(configuration).build(catalog, compilations)
    results = SafeCalculationEvaluator(function_registry=registry).execute(
        graph,
        catalog,
        compilations,
        configuration,
    )
    by_address = {reference.display: execution for reference, execution in results.items()}

    assert len(catalog.formulas) == 352
    assert sum(item.support_status == "supported" for item in compilations) == 352
    assert all(item.ir_version == "calc-ir-v2" for item in compilations)
    assert all(item.ir_json is not None for item in compilations)
    assert graph.cycles == ()
    assert by_address["Checks!D16"].value == ScalarValue.number(15)
    assert all(item.status == "executed" for item in results.values())
