from __future__ import annotations

from io import BytesIO
import os
from pathlib import Path
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import CALENDAR_MAC_1904
import pytest

from apps.api.app.calculation_rules.compiler import FormulaCompiler
from apps.api.app.calculation_rules.evaluator import SafeCalculationEvaluator, ScalarValue
from apps.api.app.calculation_rules.graph import CalculationGraphBuilder
from apps.api.app.calculation_rules.inventory import WorkbookFormulaInventory
from apps.api.app.calculation_rules.phase2_registry import PHASE2_FUNCTION_REGISTRY
from apps.api.app.calculation_rules.phase2_types import Phase2CalculationConfiguration


def _compile_and_evaluate(
    formula: str,
    values: dict[str, object] | None = None,
    *,
    date_system: str = "1900",
):
    workbook = Workbook()
    if date_system == "1904":
        workbook.epoch = CALENDAR_MAC_1904
    inputs = workbook.active
    inputs.title = "Inputs"
    for address, value in (values or {}).items():
        inputs[address] = value
    calc = workbook.create_sheet("Calc")
    calc["A1"] = formula
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    configuration = Phase2CalculationConfiguration()
    catalog = WorkbookFormulaInventory(configuration).scan(
        buffer.getvalue(),
        str(uuid.uuid4()),
    )
    compilation = FormulaCompiler(
        configuration,
        function_registry=PHASE2_FUNCTION_REGISTRY,
    ).compile(catalog.formulas[0], catalog)
    if compilation.ir_json is None:
        return compilation, None
    graph = CalculationGraphBuilder(configuration).build(catalog, (compilation,))
    execution = next(
        iter(
            SafeCalculationEvaluator(function_registry=PHASE2_FUNCTION_REGISTRY)
            .execute(graph, catalog, (compilation,), configuration)
            .values()
        )
    )
    return compilation, execution


@pytest.mark.parametrize(
    ("formula", "expected"),
    [
        ("=MOD(3,2)", 1),
        ("=MOD(-3,2)", 1),
        ("=MOD(3,-2)", -1),
        ("=MOD(-3,-2)", -1),
    ],
)
def test_mod_uses_the_divisor_sign(formula: str, expected: float) -> None:
    compilation, execution = _compile_and_evaluate(formula)

    assert compilation.support_status == "supported"
    assert execution is not None
    assert execution.value == ScalarValue.number(expected)


def test_mod_returns_division_error_for_zero_divisor() -> None:
    _compilation, execution = _compile_and_evaluate("=MOD(7,0)")

    assert execution is not None
    assert execution.value == ScalarValue.error("#DIV/0!")


def test_or_flattens_ranges_and_preserves_scalar_coercion_rules() -> None:
    _compilation, truthy = _compile_and_evaluate(
        "=OR(FALSE,Inputs!A1:A3)",
        {"A1": 0, "A2": "ignored", "A3": 2},
    )
    _compilation, falsey = _compile_and_evaluate("=OR(FALSE,0)")
    _compilation, scalar_text = _compile_and_evaluate('=OR(FALSE,"text")')
    _compilation, empty_range = _compile_and_evaluate(
        "=OR(Inputs!A1:A2)",
        {"A1": "ignored", "A2": None},
    )
    _compilation, errored = _compile_and_evaluate("=OR(TRUE,#N/A)")

    assert truthy is not None and truthy.value == ScalarValue.boolean(True)
    assert falsey is not None and falsey.value == ScalarValue.boolean(False)
    assert scalar_text is not None
    assert scalar_text.value == ScalarValue.error("#VALUE!")
    assert empty_range is not None
    assert empty_range.value == ScalarValue.error("#VALUE!")
    assert errored is not None and errored.value == ScalarValue.error("#N/A")


@pytest.mark.parametrize(
    ("date_system", "serial", "expected"),
    [
        ("1900", 0, 1900),
        ("1900", 60, 1900),
        ("1900", 45292, 2024),
        ("1904", 0, 1904),
        ("1904", 43830, 2024),
    ],
)
def test_year_respects_the_workbook_date_system(
    date_system: str,
    serial: int,
    expected: int,
) -> None:
    _compilation, execution = _compile_and_evaluate(
        f"=YEAR({serial})",
        date_system=date_system,
    )

    assert execution is not None
    assert execution.value == ScalarValue.number(expected)


@pytest.mark.parametrize(
    ("formula", "expected"),
    [("=YEAR(-1)", "#NUM!"), ('=YEAR("not-a-date")', "#VALUE!")],
)
def test_year_returns_typed_errors_for_invalid_values(
    formula: str,
    expected: str,
) -> None:
    _compilation, execution = _compile_and_evaluate(formula)

    assert execution is not None
    assert execution.status == "execution_error"
    assert execution.value == ScalarValue.error(expected)


@pytest.mark.parametrize(
    ("formula", "values", "expected"),
    [
        ("=MATCH(20,Inputs!A1:A3,0)", {"A1": 10, "A2": 20, "A3": 30}, 2),
        ("=MATCH(25,Inputs!A1:A3,1)", {"A1": 10, "A2": 20, "A3": 30}, 2),
        ("=MATCH(25,Inputs!A1:A3,-1)", {"A1": 30, "A2": 20, "A3": 10}, 1),
        ("=MATCH(2,Inputs!A1:C1,0)", {"A1": 1, "B1": 2, "C1": 3}, 2),
        (
            '=MATCH("proj*",Inputs!A1:A3,0)',
            {"A1": "Base", "A2": "Project", "A3": "Equity"},
            2,
        ),
        ('=MATCH("a~*",Inputs!A1:A2,0)', {"A1": "a?", "A2": "a*"}, 2),
    ],
)
def test_match_supports_excel_one_dimensional_modes(
    formula: str,
    values: dict[str, object],
    expected: int,
) -> None:
    _compilation, execution = _compile_and_evaluate(formula, values)

    assert execution is not None
    assert execution.value == ScalarValue.number(expected)


@pytest.mark.parametrize(
    ("formula", "values", "expected"),
    [
        (
            "=MATCH(99,Inputs!A1:A3,0)",
            {"A1": 1, "A2": 2, "A3": 3},
            "#N/A",
        ),
        (
            "=MATCH(2,Inputs!A1:B2,0)",
            {"A1": 1, "A2": 2, "B1": 3, "B2": 4},
            "#VALUE!",
        ),
        (
            "=MATCH(2,Inputs!A1:A3,2)",
            {"A1": 1, "A2": 2, "A3": 3},
            "#VALUE!",
        ),
    ],
)
def test_match_returns_typed_errors(
    formula: str,
    values: dict[str, object],
    expected: str,
) -> None:
    _compilation, execution = _compile_and_evaluate(formula, values)

    assert execution is not None
    assert execution.value == ScalarValue.error(expected)


def test_xnpv_discounts_irregular_cash_flows_from_the_first_date() -> None:
    _compilation, execution = _compile_and_evaluate(
        "=XNPV(10%,Inputs!A1:A3,Inputs!B1:B3)",
        {
            "A1": -100,
            "A2": 30,
            "A3": 90,
            "B1": 43831,
            "B2": 44013,
            "B3": 44197,
        },
    )

    expected = -100 + 30 / (1.1 ** (182 / 365)) + 90 / (1.1 ** (366 / 365))
    assert execution is not None and execution.value is not None
    assert execution.value.number_value == pytest.approx(expected, abs=1e-9)


@pytest.mark.parametrize(
    ("formula", "values", "expected"),
    [
        (
            "=XNPV(10%,Inputs!A1:A2,Inputs!B1:B3)",
            {
                "A1": -100,
                "A2": 110,
                "B1": 43831,
                "B2": 44197,
                "B3": 44562,
            },
            "#VALUE!",
        ),
        (
            "=XNPV(10%,Inputs!A1:A2,Inputs!B1:B2)",
            {"A1": -100, "A2": "bad", "B1": 43831, "B2": 44197},
            "#VALUE!",
        ),
        (
            "=XNPV(10%,Inputs!A1:A2,Inputs!B1:B2)",
            {"A1": -100, "A2": 110, "B1": 44197, "B2": 43831},
            "#NUM!",
        ),
        (
            "=XNPV(-100%,Inputs!A1:A2,Inputs!B1:B2)",
            {"A1": -100, "A2": 110, "B1": 43831, "B2": 44197},
            "#NUM!",
        ),
    ],
)
def test_xnpv_returns_typed_input_and_domain_errors(
    formula: str,
    values: dict[str, object],
    expected: str,
) -> None:
    _compilation, execution = _compile_and_evaluate(formula, values)

    assert execution is not None
    assert execution.value == ScalarValue.error(expected)


def test_xirr_solves_irregular_dates_deterministically() -> None:
    values = {
        "A1": -10000,
        "A2": 2750,
        "A3": 4250,
        "A4": 3250,
        "A5": 2750,
        "B1": 39448,
        "B2": 39508,
        "B3": 39751,
        "B4": 39859,
        "B5": 39904,
    }
    first = _compile_and_evaluate(
        "=XIRR(Inputs!A1:A5,Inputs!B1:B5)",
        values,
    )[1]
    second = _compile_and_evaluate(
        "=XIRR(Inputs!A1:A5,Inputs!B1:B5,10%)",
        values,
    )[1]

    assert first is not None and first.value is not None
    assert second is not None and second.value is not None
    assert first.value.number_value == pytest.approx(0.373362535, abs=1e-7)
    assert second.value.number_value == pytest.approx(
        first.value.number_value,
        abs=1e-10,
    )


@pytest.mark.parametrize(
    ("values", "expected"),
    [
        ({"A1": 100, "A2": 10, "B1": 43831, "B2": 44197}, "#NUM!"),
        ({"A1": -100, "A2": -10, "B1": 43831, "B2": 44197}, "#NUM!"),
        ({"A1": -100, "A2": 110, "B1": 44197, "B2": 43831}, "#NUM!"),
    ],
)
def test_xirr_rejects_invalid_sign_and_date_inputs(
    values: dict[str, object],
    expected: str,
) -> None:
    _compilation, execution = _compile_and_evaluate(
        "=XIRR(Inputs!A1:A2,Inputs!B1:B2)",
        values,
    )

    assert execution is not None
    assert execution.value == ScalarValue.error(expected)


def test_xirr_returns_num_when_no_root_converges() -> None:
    _compilation, execution = _compile_and_evaluate(
        "=XIRR(Inputs!A1:A3,Inputs!B1:B3,-99%)",
        {
            "A1": -100,
            "A2": 1,
            "A3": -1,
            "B1": 43831,
            "B2": 43832,
            "B3": 43833,
        },
    )

    assert execution is not None
    assert execution.value == ScalarValue.error("#NUM!")
