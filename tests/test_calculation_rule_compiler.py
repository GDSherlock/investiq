from __future__ import annotations

from copy import deepcopy
from dataclasses import replace
from io import BytesIO

from openpyxl import Workbook
import pytest

from apps.api.app.calculation_rules.compiler import (
    CalculationExpressionValidator,
    FormulaCompiler,
)
from apps.api.app.calculation_rules.function_registry import (
    FUNCTION_REGISTRY,
    FUNCTION_REGISTRY_VERSION,
)
from apps.api.app.calculation_rules.inventory import WorkbookFormulaInventory
from apps.api.app.calculation_rules.types import (
    CalculationRuleExtractionConfiguration,
)


WORKBOOK_VERSION_ID = "22345678-1234-5678-9234-567812345678"


def _catalog(formula: str, *, configuration=None):
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = 10
    inputs["A2"] = 20
    inputs["A3"] = 30
    inputs["B2"] = 2
    quoted = workbook.create_sheet("Very Hidden")
    quoted["A1"] = 7
    calc = workbook.create_sheet("Calc")
    calc["C5"] = formula
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return WorkbookFormulaInventory(configuration).scan(
        buffer.getvalue(),
        WORKBOOK_VERSION_ID,
    )


def _compile(formula: str, *, configuration=None):
    configuration = configuration or CalculationRuleExtractionConfiguration()
    catalog = _catalog(formula, configuration=configuration)
    return FormulaCompiler(configuration).compile(catalog.formulas[0], catalog)


@pytest.mark.parametrize(
    ("formula", "root_type", "root_operator"),
    [
        ("=1+2*3", "binary_operation", "add"),
        ("=-2^2%", "binary_operation", "power"),
        ("=+5/2-1", "binary_operation", "subtract"),
        ("=Inputs!A1>=Inputs!$B$2", "comparison", "greater_equal"),
        ('=IF(Inputs!A1>0,SUM(Inputs!A1:A3),"none")', "function_call", None),
        ("=AVERAGE(Inputs!A1:A3)", "function_call", None),
        ("=MIN(1,2)+MAX(3,4)+ABS(-5)+ROUND(2.5,0)", "binary_operation", "add"),
        ("=TRUE<>FALSE", "comparison", "not_equal"),
        ('="exact"', "literal", None),
        ("=#N/A", "error_value", None),
    ],
)
def test_compiler_emits_valid_calc_ir_v1(
    formula: str,
    root_type: str,
    root_operator: str | None,
) -> None:
    compilation = _compile(formula)

    assert compilation.parse_status == "parsed"
    assert compilation.support_status == "supported"
    assert compilation.unsupported_constructs == ()
    assert compilation.ir_json is not None
    assert compilation.ir_json["ir_version"] == "calc-ir-v1"
    assert compilation.ir_json["compiler_version"] == "formula-compiler-v1"
    assert compilation.ir_json["semantics_profile"] == "excel-subset-v1"
    assert compilation.ir_json["root"]["node_type"] == root_type
    if root_operator is not None:
        assert compilation.ir_json["root"]["operator"] == root_operator


def test_reference_evidence_preserves_spans_anchors_ranges_and_exact_sheet() -> None:
    formula = "=SUM(Inputs!$A1:A$3,'Very Hidden'!$A$1)"
    compilation = _compile(formula)

    assert compilation.support_status == "supported"
    assert len(compilation.references) == 2
    range_ref, cell_ref = compilation.references
    assert formula[range_ref.source_span_start : range_ref.source_span_end] == (
        "Inputs!$A1:A$3"
    )
    assert range_ref.reference_kind == "range"
    assert range_ref.target_sheet_name == "Inputs"
    assert range_ref.start_cell_address == "A1"
    assert range_ref.end_cell_address == "A3"
    assert range_ref.start_column_absolute is True
    assert range_ref.start_row_absolute is False
    assert range_ref.end_column_absolute is False
    assert range_ref.end_row_absolute is True
    assert range_ref.range_rows == 3
    assert range_ref.range_columns == 1
    assert cell_ref.source_token == "'Very Hidden'!$A$1"
    assert cell_ref.target_sheet_name == "Very Hidden"
    assert cell_ref.start_column_absolute is True
    assert cell_ref.start_row_absolute is True
    assert cell_ref.resolution_status == "resolved_internal"
    assert range_ref.id != cell_ref.id


def test_normalized_signature_is_coordinate_relative_and_anchor_aware() -> None:
    compilation = _compile("=Inputs!A1+$C$4")

    assert compilation.normalized_signature == (
        "add(ref:Inputs!R[-4]C[-2],ref:Calc!R4C3)"
    )


def test_parentheses_around_literals_and_references_remain_supported() -> None:
    compilation = _compile("=(Inputs!A1)+(2)")

    assert compilation.support_status == "supported"
    assert compilation.ir_json is not None


@pytest.mark.parametrize(
    ("formula", "reason"),
    [
        ('=COUNTIF(Inputs!A1:A3,">0")', "unsupported_function:COUNTIF"),
        ("=VLOOKUP(1,Inputs!A1:A3,1,FALSE)", "unsupported_function:VLOOKUP"),
        ("=TaxRate+1", "named_reference:TaxRate"),
        ("=Inputs!A:A", "whole_column_reference"),
        ("=Inputs!1:1", "whole_row_reference"),
        ("=Inputs!A1&\"x\"", "text_concatenation"),
        ("={1,2}", "array_constant"),
        ("=Inputs:Calc!A1", "three_dimensional_reference"),
    ],
)
def test_explicit_exclusions_are_unsupported_with_exact_evidence(
    formula: str,
    reason: str,
) -> None:
    compilation = _compile(formula)

    assert compilation.ir_json is None
    assert compilation.support_status == "unsupported"
    assert reason in compilation.unsupported_constructs
    assert compilation.parse_status in {"parsed", "not_attempted"}


def test_unknown_function_retains_recognized_reference_evidence() -> None:
    compilation = _compile('=COUNTIF(Inputs!A1:A3,">0")')

    assert len(compilation.references) == 1
    assert compilation.references[0].source_token == "Inputs!A1:A3"
    assert compilation.references[0].resolution_status == "resolved_internal"


def test_external_reference_is_evidence_only_and_never_executable() -> None:
    formula = "='[rates.xlsx]Inputs'!$A$1+1"
    compilation = _compile(formula)

    assert compilation.parse_status == "parsed"
    assert compilation.support_status == "external_reference"
    assert compilation.ir_json is None
    assert len(compilation.references) == 1
    reference = compilation.references[0]
    assert reference.source_token == "'[rates.xlsx]Inputs'!$A$1"
    assert reference.target_classification == "external"
    assert reference.resolution_status == "external"


def test_missing_internal_sheet_is_unresolved_not_guessed() -> None:
    compilation = _compile("=Missing!A1+1")

    assert compilation.support_status == "unsupported"
    assert compilation.ir_json is None
    assert compilation.references[0].resolution_status == "missing_sheet"
    assert "missing_sheet:Missing" in compilation.unsupported_constructs


def test_syntax_error_is_distinct_from_unsupported_formula() -> None:
    compilation = _compile("=1+")

    assert compilation.parse_status == "syntax_error"
    assert compilation.support_status == "unsupported"
    assert compilation.ir_json is None
    assert compilation.unsupported_constructs == ("syntax_error",)


def test_function_arity_is_validated_at_compile_time() -> None:
    compilation = _compile("=ABS(1,2)")

    assert compilation.parse_status == "parsed"
    assert compilation.support_status == "unsupported"
    assert compilation.ir_json is None
    assert compilation.unsupported_constructs == ("invalid_arity:ABS",)


def test_function_registry_is_closed_versioned_and_marks_if_lazy() -> None:
    assert FUNCTION_REGISTRY_VERSION == "function-registry-v1"
    assert set(FUNCTION_REGISTRY) == {
        "SUM",
        "AVERAGE",
        "MIN",
        "MAX",
        "ABS",
        "ROUND",
        "IF",
    }
    assert FUNCTION_REGISTRY["IF"].lazy is True
    assert FUNCTION_REGISTRY["SUM"].accepts_ranges is True
    assert FUNCTION_REGISTRY["ABS"].accepts_ranges is False


@pytest.mark.parametrize(
    ("configuration", "formula", "reason"),
    [
        (CalculationRuleExtractionConfiguration(max_nodes=2), "=1+2", "node_limit"),
        (
            CalculationRuleExtractionConfiguration(max_depth=1),
            "=1+2",
            "nesting limit",
        ),
        (
            CalculationRuleExtractionConfiguration(max_arguments=2),
            "=SUM(1,2,3)",
            "argument_limit",
        ),
        (
            CalculationRuleExtractionConfiguration(max_range_cells=2),
            "=SUM(Inputs!A1:A3)",
            "range_cell_limit",
        ),
    ],
)
def test_compiler_resource_limits_are_evidence_not_partial_ir(
    configuration: CalculationRuleExtractionConfiguration,
    formula: str,
    reason: str,
) -> None:
    compilation = _compile(formula, configuration=configuration)

    assert compilation.support_status == "unsupported"
    assert compilation.ir_json is None
    assert any(reason in item for item in compilation.unsupported_constructs)


def test_deep_but_token_bounded_expression_is_rejected_without_recursion_crash() -> None:
    formula = "=" + "+".join("1" for _ in range(900))

    compilation = _compile(formula)

    assert compilation.support_status == "unsupported"
    assert compilation.ir_json is None
    assert any(
        "nesting" in reason for reason in compilation.unsupported_constructs
    )


def test_out_of_bounds_a1_reference_is_retained_as_invalid_evidence() -> None:
    compilation = _compile("=XFE1+1")

    assert compilation.support_status == "unsupported"
    assert compilation.ir_json is None
    assert compilation.references[0].source_token == "XFE1"
    assert compilation.references[0].resolution_status == "invalid_address"
    assert "invalid_address" in compilation.unsupported_constructs


def test_phase_one_configuration_rejects_unregistered_behavior_versions() -> None:
    with pytest.raises(ValueError, match="Unregistered calculation version"):
        CalculationRuleExtractionConfiguration(ir_version="arbitrary-ir")


def test_formula_and_token_limits_reject_without_partial_ir() -> None:
    length_configuration = CalculationRuleExtractionConfiguration(max_formula_length=4)
    token_configuration = CalculationRuleExtractionConfiguration(max_tokens=3)

    too_long = _compile("=1+23", configuration=length_configuration)
    too_many_tokens = _compile("=1+2+3", configuration=token_configuration)

    assert too_long.unsupported_constructs == ("formula_length_limit",)
    assert too_many_tokens.unsupported_constructs == ("token_limit",)
    assert too_long.ir_json is None
    assert too_many_tokens.ir_json is None


def test_ir_validator_rejects_unknown_nodes_and_tampered_formula_hash() -> None:
    catalog = _catalog("=Inputs!A1+1")
    formula_cell = catalog.formulas[0]
    compilation = FormulaCompiler().compile(formula_cell, catalog)
    validator = CalculationExpressionValidator()
    unknown = deepcopy(compilation.ir_json)
    unknown["root"] = {
        "node_type": "python_expression",
        "source_span": {"start": 1, "end": len(formula_cell.exact_formula)},
        "source": "__import__('os')",
    }
    tampered = deepcopy(compilation.ir_json)
    tampered["formula_sha256"] = "0" * 64

    with pytest.raises(ValueError, match="Unknown calculation node type"):
        validator.validate(
            unknown,
            formula_cell,
            compilation.references,
            CalculationRuleExtractionConfiguration(),
        )
    with pytest.raises(ValueError, match="formula hash"):
        validator.validate(
            tampered,
            formula_cell,
            compilation.references,
            CalculationRuleExtractionConfiguration(),
        )


def test_ir_validator_rejects_reference_target_tampering_and_missing_evidence() -> None:
    catalog = _catalog("=Inputs!A1+1")
    formula_cell = catalog.formulas[0]
    compilation = FormulaCompiler().compile(formula_cell, catalog)
    validator = CalculationExpressionValidator()
    tampered_target = deepcopy(compilation.ir_json)
    tampered_target["root"]["left"]["cell"]["cell_address"] = "A2"
    missing_reference = deepcopy(compilation.ir_json)
    missing_reference["root"] = missing_reference["root"]["right"]
    missing_reference["normalized_signature"] = "literal:number:1"

    with pytest.raises(ValueError, match="reference target"):
        validator.validate(
            tampered_target,
            formula_cell,
            compilation.references,
            CalculationRuleExtractionConfiguration(),
        )
    with pytest.raises(ValueError, match="reference evidence"):
        validator.validate(
            missing_reference,
            formula_cell,
            compilation.references,
            CalculationRuleExtractionConfiguration(),
        )


def test_ir_validator_rejects_forged_reference_evidence_and_literal_types() -> None:
    catalog = _catalog("=Inputs!A1+1")
    formula_cell = catalog.formulas[0]
    compilation = FormulaCompiler().compile(formula_cell, catalog)
    forged_reference = replace(
        compilation.references[0],
        source_token="Inputs!A2",
    )

    with pytest.raises(ValueError, match="reference source token"):
        CalculationExpressionValidator().validate(
            compilation.ir_json,
            formula_cell,
            (forged_reference,),
            CalculationRuleExtractionConfiguration(),
        )

    boolean_catalog = _catalog("=TRUE")
    boolean_cell = boolean_catalog.formulas[0]
    boolean_compilation = FormulaCompiler().compile(boolean_cell, boolean_catalog)
    tampered_boolean = deepcopy(boolean_compilation.ir_json)
    tampered_boolean["root"]["value"] = "TRUE"
    tampered_boolean["normalized_signature"] = "literal:boolean:TRUE"

    with pytest.raises(ValueError, match="Boolean literal"):
        CalculationExpressionValidator().validate(
            tampered_boolean,
            boolean_cell,
            (),
            CalculationRuleExtractionConfiguration(),
        )


def test_ir_validator_rejects_literal_spans_that_do_not_match_the_lexeme() -> None:
    catalog = _catalog("=1+2")
    formula_cell = catalog.formulas[0]
    compilation = FormulaCompiler().compile(formula_cell, catalog)
    tampered = deepcopy(compilation.ir_json)
    tampered["root"]["left"]["source_span"] = {"start": 3, "end": 4}

    with pytest.raises(ValueError, match="source span"):
        CalculationExpressionValidator().validate(
            tampered,
            formula_cell,
            (),
            CalculationRuleExtractionConfiguration(),
        )


def test_ir_validator_rejects_operator_that_disagrees_with_formula_source() -> None:
    catalog = _catalog("=1+2")
    formula_cell = catalog.formulas[0]
    compilation = FormulaCompiler().compile(formula_cell, catalog)
    tampered = deepcopy(compilation.ir_json)
    tampered["root"]["operator"] = "multiply"
    tampered["normalized_signature"] = "multiply(literal:number:1,literal:number:2)"

    with pytest.raises(ValueError, match="operator source"):
        CalculationExpressionValidator().validate(
            tampered,
            formula_cell,
            (),
            CalculationRuleExtractionConfiguration(),
        )


def test_ir_validator_requires_root_and_function_arguments_to_cover_source() -> None:
    catalog = _catalog("=1+2")
    formula_cell = catalog.formulas[0]
    compilation = FormulaCompiler().compile(formula_cell, catalog)
    truncated_root = deepcopy(compilation.ir_json)
    truncated_root["root"] = truncated_root["root"]["left"]
    truncated_root["normalized_signature"] = "literal:number:1"

    with pytest.raises(ValueError, match="root source span"):
        CalculationExpressionValidator().validate(
            truncated_root,
            formula_cell,
            (),
            CalculationRuleExtractionConfiguration(),
        )

    function_catalog = _catalog("=SUM(1,2)")
    function_cell = function_catalog.formulas[0]
    function_compilation = FormulaCompiler().compile(function_cell, function_catalog)
    omitted_argument = deepcopy(function_compilation.ir_json)
    omitted_argument["root"]["arguments"] = omitted_argument["root"]["arguments"][:1]
    omitted_argument["normalized_signature"] = "SUM(literal:number:1)"

    with pytest.raises(ValueError, match="function argument source"):
        CalculationExpressionValidator().validate(
            omitted_argument,
            function_cell,
            (),
            CalculationRuleExtractionConfiguration(),
        )
