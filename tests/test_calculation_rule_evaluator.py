from __future__ import annotations

from copy import deepcopy
from datetime import date
from io import BytesIO

from openpyxl import Workbook
import pytest

from apps.api.app.calculation_rules.comparison import CachedValueComparator
from apps.api.app.calculation_rules.compiler import FormulaCompiler
from apps.api.app.calculation_rules.evaluator import (
    CalculationExecutionContext,
    SafeCalculationEvaluator,
    ScalarValue,
)
from apps.api.app.calculation_rules.graph import CalculationGraphBuilder
from apps.api.app.calculation_rules.inventory import WorkbookFormulaInventory
from apps.api.app.calculation_rules.types import CalculationRuleExtractionConfiguration


WORKBOOK_VERSION_ID = "42345678-1234-5678-9234-567812345678"


def _case(formula: str, values: dict[str, object] | None = None):
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    for address, value in (values or {}).items():
        inputs[address] = value
    calc = workbook.create_sheet("Calc")
    calc["C5"] = formula
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    configuration = CalculationRuleExtractionConfiguration()
    catalog = WorkbookFormulaInventory(configuration).scan(
        buffer.getvalue(), WORKBOOK_VERSION_ID
    )
    formula_cell = catalog.formulas[0]
    compilation = FormulaCompiler(configuration).compile(formula_cell, catalog)
    assert compilation.ir_json is not None
    context = CalculationExecutionContext(
        catalog=catalog,
        formula_cell=formula_cell,
        compilation=compilation,
        calculated_values={},
        configuration=configuration,
    )
    return SafeCalculationEvaluator().evaluate(compilation.ir_json, context)


@pytest.mark.parametrize(
    ("formula", "expected"),
    [
        ("=1+2*3", ScalarValue.number(7)),
        ("=-2^2", ScalarValue.number(4)),
        ("=50%", ScalarValue.number(0.5)),
        ("=ABS(-5)", ScalarValue.number(5)),
        ("=ROUND(-2.5,0)", ScalarValue.number(-3)),
        ("=ROUND(125,-1)", ScalarValue.number(130)),
        ("=IF(TRUE,1,1/0)", ScalarValue.number(1)),
        ("=IF(FALSE,1/0,2)", ScalarValue.number(2)),
        ('="Alpha"="alpha"', ScalarValue.boolean(True)),
        ("=MIN()", ScalarValue.number(0)),
        ("=MAX()", ScalarValue.number(0)),
    ],
)
def test_safe_evaluator_conforms_to_phase_one_scalar_semantics(
    formula: str,
    expected: ScalarValue,
) -> None:
    execution = _case(formula)

    assert execution.status == "executed"
    assert execution.value == expected


def test_range_aggregations_ignore_text_boolean_and_blank_cells() -> None:
    execution = _case(
        "=SUM(Inputs!A1:A5)",
        {"A1": 1, "A2": 2, "A3": "ignored", "A4": True},
    )

    assert execution.value == ScalarValue.number(3)


def test_direct_boolean_is_coerced_but_direct_text_is_value_error() -> None:
    assert _case("=SUM(TRUE,2)").value == ScalarValue.number(3)
    error = _case('=SUM("2",1)')

    assert error.status == "execution_error"
    assert error.value == ScalarValue.error("#VALUE!")


def test_average_without_numeric_range_values_returns_division_error() -> None:
    execution = _case("=AVERAGE(Inputs!A1:A2)", {"A1": "x", "A2": True})

    assert execution.status == "execution_error"
    assert execution.value == ScalarValue.error("#DIV/0!")


@pytest.mark.parametrize(
    ("formula", "error_code"),
    [
        ("=1/0", "#DIV/0!"),
        ('=1+"x"', "#VALUE!"),
        ("=1E308*1E308", "#NUM!"),
        ("=#N/A+1", "#N/A"),
    ],
)
def test_excel_errors_are_values_and_propagate_left_to_right(
    formula: str,
    error_code: str,
) -> None:
    execution = _case(formula)

    assert execution.status == "execution_error"
    assert execution.value == ScalarValue.error(error_code)


def test_aggregate_numeric_overflow_returns_num_error_instead_of_raising() -> None:
    execution = _case("=SUM(1E308,1E308)")

    assert execution.status == "execution_error"
    assert execution.value == ScalarValue.error("#NUM!")


def test_date_static_cell_is_converted_using_workbook_date_system() -> None:
    execution = _case("=Inputs!A1+1", {"A1": date(2026, 1, 1)})

    assert execution.status == "executed"
    assert execution.value.kind == "number"
    assert execution.value.number_value > 46_000


def test_trace_is_bounded_and_contains_only_typed_direct_inputs() -> None:
    execution = _case("=SUM(Inputs!A1:A3)", {"A1": 1, "A2": 2, "A3": 3})

    assert [item["cell_address"] for item in execution.direct_input_trace] == [
        "A1",
        "A2",
        "A3",
    ]
    assert all("formula" not in item for item in execution.direct_input_trace)


def test_evaluator_revalidates_ir_and_rejects_unknown_execution_nodes() -> None:
    workbook = Workbook()
    workbook.active["A1"] = "=1+2"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    configuration = CalculationRuleExtractionConfiguration()
    catalog = WorkbookFormulaInventory(configuration).scan(
        buffer.getvalue(), WORKBOOK_VERSION_ID
    )
    formula_cell = catalog.formulas[0]
    compilation = FormulaCompiler(configuration).compile(formula_cell, catalog)
    tampered = deepcopy(compilation.ir_json)
    tampered["root"]["node_type"] = "python_expression"
    context = CalculationExecutionContext(
        catalog, formula_cell, compilation, {}, configuration
    )

    with pytest.raises(ValueError, match="Unknown calculation node type"):
        SafeCalculationEvaluator().evaluate(tampered, context)


def test_graph_execution_never_uses_cached_value_for_blocked_formula() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calc"
    sheet["A1"] = '=COUNTIF(B1:B2,">0")'
    sheet["A2"] = "=A1+1"
    sheet["B1"] = 1
    sheet["B2"] = 2
    sheet["A3"] = "=B1+B2"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    configuration = CalculationRuleExtractionConfiguration()
    catalog = WorkbookFormulaInventory(configuration).scan(
        buffer.getvalue(), WORKBOOK_VERSION_ID
    )
    compilations = tuple(
        FormulaCompiler(configuration).compile(cell, catalog)
        for cell in catalog.formulas
    )
    plan = CalculationGraphBuilder(configuration).build(catalog, compilations)

    results = SafeCalculationEvaluator().execute(plan, catalog, compilations)
    by_address = {reference.cell_address: result for reference, result in results.items()}

    assert by_address["A1"].status == "not_executable"
    assert by_address["A2"].status == "blocked_by_dependency"
    assert by_address["A3"].status == "executed"
    assert by_address["A3"].value == ScalarValue.number(3)


def test_cached_value_comparator_uses_type_exact_and_numeric_tolerance_rules() -> None:
    comparator = CachedValueComparator(CalculationRuleExtractionConfiguration())

    close = comparator.compare(
        ScalarValue.number(100.0), ScalarValue.number(100.0 + 5e-8), "unknown"
    )
    far = comparator.compare(
        ScalarValue.number(100.0), ScalarValue.number(100.1), "unknown"
    )
    boolean_number = comparator.compare(
        ScalarValue.boolean(True), ScalarValue.number(1), "unknown"
    )
    missing = comparator.compare(ScalarValue.number(1), None, "missing")

    assert close.validation_status == "matched"
    assert close.absolute_error == pytest.approx(5e-8)
    assert far.validation_status == "mismatched"
    assert boolean_number.validation_status == "not_comparable"
    assert missing.validation_status == "no_cached_value"


def test_cached_error_comparison_requires_same_excel_error_code() -> None:
    comparator = CachedValueComparator()

    same = comparator.compare(
        ScalarValue.error("#N/A"), ScalarValue.error("#N/A"), "unknown"
    )
    different = comparator.compare(
        ScalarValue.error("#N/A"), ScalarValue.error("#REF!"), "unknown"
    )

    assert same.validation_status == "matched"
    assert different.validation_status == "mismatched"


def test_cached_date_comparison_uses_date_specific_contract_not_numeric_tolerance() -> None:
    comparator = CachedValueComparator()

    date_exact = comparator.compare(
        ScalarValue.date_serial(46_000, "2025-12-09"),
        ScalarValue.date_serial(46_000, "2025-12-09"),
        "unknown",
    )
    date_fraction = comparator.compare(
        ScalarValue.date_serial(46_000, "2025-12-09"),
        ScalarValue.date_serial(46_000.0000000001, "2025-12-09"),
        "unknown",
    )
    datetime_close = comparator.compare(
        ScalarValue.date_serial(46_000.5, "2025-12-09T12:00:00"),
        ScalarValue.date_serial(46_000.5000000005, "2025-12-09T12:00:00"),
        "unknown",
    )
    datetime_far = comparator.compare(
        ScalarValue.date_serial(46_000.5, "2025-12-09T12:00:00"),
        ScalarValue.date_serial(46_000.500000002, "2025-12-09T12:00:00"),
        "unknown",
    )
    number_date = comparator.compare(
        ScalarValue.number(46_000),
        ScalarValue.date_serial(46_000, "2025-12-09"),
        "unknown",
    )

    assert date_exact.validation_status == "matched"
    assert date_fraction.validation_status == "mismatched"
    assert datetime_close.validation_status == "matched"
    assert datetime_far.validation_status == "mismatched"
    assert number_date.validation_status == "not_comparable"
