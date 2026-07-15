from __future__ import annotations

from io import BytesIO
import uuid

from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from apps.api.app.calculation_rules.inventory import WorkbookFormulaInventory
from apps.api.app.calculation_rules.types import (
    CalculationRuleExtractionConfiguration,
    FormulaIdFactory,
    WorkbookCellRef,
)


WORKBOOK_VERSION_ID = "12345678-1234-5678-9234-567812345678"


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    visible = workbook.active
    visible.title = "Visible"
    visible["A1"] = 2
    visible["B1"] = "=A1*2"
    visible["B1"].number_format = "0.00"

    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = 3
    hidden["B1"] = "=Visible!B1+A1"

    very_hidden = workbook.create_sheet("Very Hidden")
    very_hidden.sheet_state = "veryHidden"
    very_hidden["A1"] = "helper"
    very_hidden["B1"] = "=SUM(Visible!A1,Hidden!A1)"

    workbook.calculation.fullCalcOnLoad = True
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def test_default_configuration_freezes_phase_one_versions_and_budgets() -> None:
    configuration = CalculationRuleExtractionConfiguration()

    assert configuration.ir_version == "calc-ir-v1"
    assert configuration.compiler_version == "formula-compiler-v1"
    assert configuration.engine_version == "calc-engine-v1"
    assert configuration.function_registry_version == "function-registry-v1"
    assert configuration.semantics_profile == "excel-subset-v1"
    assert configuration.max_formula_length == 8_192
    assert configuration.max_tokens == 2_048
    assert configuration.max_nodes == 2_048
    assert configuration.max_depth == 128
    assert configuration.max_arguments == 255
    assert configuration.max_range_cells == 10_000
    assert configuration.max_formula_count == 100_000
    assert configuration.max_total_edges == 1_000_000
    assert configuration.max_trace_inputs == 256
    assert configuration.absolute_tolerance == 1e-9
    assert configuration.relative_tolerance == 1e-9


def test_formula_cell_ids_are_retry_stable_and_use_exact_workbook_identity() -> None:
    reference = WorkbookCellRef(
        workbook_version_id=WORKBOOK_VERSION_ID,
        sheet_name="Visible",
        sheet_position=0,
        cell_address="$b$1",
    )
    factory = FormulaIdFactory(WORKBOOK_VERSION_ID)

    first = factory.formula_cell_id(reference)
    second = factory.formula_cell_id(reference)
    different_sheet = factory.formula_cell_id(
        WorkbookCellRef(WORKBOOK_VERSION_ID, "visible", 0, "B1")
    )

    expected = str(
        uuid.uuid5(
            uuid.UUID(WORKBOOK_VERSION_ID),
            "formula-cell|0|Visible|B1",
        )
    )
    assert first == second == expected
    assert different_sheet != first
    assert reference.cell_address == "B1"


def test_inventory_scans_visible_hidden_and_very_hidden_formula_cells() -> None:
    catalog = WorkbookFormulaInventory().scan(
        _workbook_bytes(),
        WORKBOOK_VERSION_ID,
    )

    assert [
        (cell.ref.sheet_name, cell.ref.cell_address, cell.sheet_state)
        for cell in catalog.formulas
    ] == [
        ("Visible", "B1", "visible"),
        ("Hidden", "B1", "hidden"),
        ("Very Hidden", "B1", "veryHidden"),
    ]
    assert catalog.formulas[0].exact_formula == "=A1*2"
    assert catalog.formulas[0].formula_sha256
    assert catalog.formulas[0].cached_value is None
    assert catalog.formulas[0].cache_status == "missing"
    assert catalog.formulas[0].cache_freshness == "recalculation_required"
    assert catalog.formulas[0].number_format == "0.00"


def test_inventory_catalog_retains_static_and_blank_cell_facts() -> None:
    catalog = WorkbookFormulaInventory().scan(
        _workbook_bytes(),
        WORKBOOK_VERSION_ID,
    )

    static = catalog.cell(WorkbookCellRef(WORKBOOK_VERSION_ID, "Visible", 0, "A1"))
    blank = catalog.cell(WorkbookCellRef(WORKBOOK_VERSION_ID, "Visible", 0, "C1"))

    assert static.value == 2
    assert static.value_type == "number"
    assert static.formula_kind is None
    assert blank.value is None
    assert blank.value_type == "blank"
    assert catalog.sheet_position("Very Hidden") == 2
    assert catalog.sheet_position("very hidden") is None


def test_inventory_rejects_formula_count_over_budget() -> None:
    configuration = CalculationRuleExtractionConfiguration(max_formula_count=2)

    try:
        WorkbookFormulaInventory(configuration).scan(
            _workbook_bytes(),
            WORKBOOK_VERSION_ID,
        )
    except ValueError as exc:
        assert str(exc) == "Workbook formula count exceeds configured limit"
    else:
        raise AssertionError("Expected formula-count limit failure")


def test_inventory_retains_array_and_data_table_ooxml_metadata_without_guessing_text() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = ArrayFormula(ref="A1:A2", text="=B1:B2*2")
    worksheet["C1"] = DataTableFormula(
        ref="C1:D2",
        dt2D=True,
        r1="B1",
        r2="B2",
    )
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    catalog = WorkbookFormulaInventory().scan(
        buffer.getvalue(),
        WORKBOOK_VERSION_ID,
    )
    array_formula, data_table = catalog.formulas

    assert array_formula.formula_kind == "array"
    assert array_formula.exact_formula == "=B1:B2*2"
    assert array_formula.special_metadata == {"t": "array", "ref": "A1:A2"}
    assert data_table.formula_kind == "data_table"
    assert data_table.exact_formula == ""
    assert data_table.special_metadata == {
        "t": "dataTable",
        "ref": "C1:D2",
        "dt2D": "1",
        "r1": "B1",
        "r2": "B2",
    }


def test_inventory_visits_explicit_cells_without_materializing_sparse_rectangles() -> None:
    workbook = Workbook()
    workbook.active["A1"] = "=1+1"
    workbook.active["Z1000"].number_format = "0.00"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    catalog = WorkbookFormulaInventory().scan(
        buffer.getvalue(),
        WORKBOOK_VERSION_ID,
    )

    assert len(catalog.formulas) == 1
    assert len(catalog._cells) == 2
