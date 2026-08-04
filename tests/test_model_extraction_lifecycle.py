from __future__ import annotations

from copy import deepcopy
from io import BytesIO
import os
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.engine import make_url
from sqlalchemy.orm import sessionmaker

from apps.api.app.database import Base
from apps.api.app.model_extraction_models import (
    CanonicalOutput,
    FinancialSeries,
    FinancialSeriesValue,
    ModelParameter,
    ModelSemanticBinding,
    ModelVersion,
    WorkbookVersion,
)
from apps.api.app.model_extraction_read_service import ModelExtractionReadService
from apps.api.app.model_extraction_repository import ModelExtractionRepository
from apps.api.app.model_extraction_service import (
    ModelExtractionPersistenceService,
    _canonicalize_outputs,
)
from apps.api.app.model_extraction_types import (
    CanonicalSourceConflictError,
    FinancialEntityIdFactory,
    ModelExtractionPersistenceError,
    WorkbookTooLargeError,
    new_uuid,
)
from apps.api.app.workbook_storage import DatabaseWorkbookStorage
from apps.api.app.workbook_validation import (
    InvalidWorkbookError,
    WorkbookToolset,
    materialize_financial_series,
)
from tests.model_extraction_test_support import (
    create_sqlite_session_factory,
    deterministic_extraction_result,
    persistence_workbook_bytes,
    sqlite_file_url,
)


class RecordingRunner:
    def __init__(self, result=None, callback=None):
        self.result = result or deterministic_extraction_result()
        self.callback = callback
        self.calls = 0

    def __call__(self, file_bytes: bytes, filename: str):
        self.calls += 1
        if self.callback is not None:
            self.callback(file_bytes, filename)
        return deepcopy(self.result)


class FailOnceAfterCanonicalRepository(ModelExtractionRepository):
    def __init__(self, session):
        super().__init__(session)
        self.failures_remaining = 1

    def persist_canonical_model(self, *args, **kwargs):
        model_version = super().persist_canonical_model(*args, **kwargs)
        if self.failures_remaining:
            self.failures_remaining -= 1
            raise RuntimeError("forced T3 failure")
        return model_version


class FailOnceSnapshotRepository(ModelExtractionRepository):
    def __init__(self, session):
        super().__init__(session)
        self.snapshot_attempts = 0

    def save_extraction_snapshot(self, *args, **kwargs):
        self.snapshot_attempts += 1
        if self.snapshot_attempts == 1:
            raise RuntimeError("transient T2 failure")
        return super().save_extraction_snapshot(*args, **kwargs)


@pytest.fixture
def lifecycle_context(tmp_path: Path):
    engine, session_factory = create_sqlite_session_factory(
        sqlite_file_url(tmp_path / "lifecycle.db")
    )
    Base.metadata.create_all(engine)
    session = session_factory()
    try:
        yield engine, session_factory, session
    finally:
        session.close()
        engine.dispose()


def _count(session, model) -> int:
    return int(session.scalar(select(func.count()).select_from(model)) or 0)


def _contains_key(value, target: str) -> bool:
    if isinstance(value, dict):
        return target in value or any(_contains_key(item, target) for item in value.values())
    if isinstance(value, list):
        return any(_contains_key(item, target) for item in value)
    return False


def test_validated_derived_scalar_is_promoted_to_canonical_output() -> None:
    snapshot = deterministic_extraction_result()
    final_extraction = snapshot["final_extraction"]
    final_extraction["output_candidates"] = []
    final_extraction["derived_value_candidates"] = [
        {
            "candidate_id": "summary-project-irr",
            "original_label": "Project IRR",
            "submitted_role": "formula_derived_value",
            "business_role": "project_irr",
            "raw_value": 0.074,
            "unit": "%",
            "source_references": [{"sheet_name": "P&L", "cell": "B3"}],
            "reasoning_summary": "Displayed project return",
            "llm_confidence": 0.99,
        }
    ]
    snapshot["validation_results"] = [
        {
            "candidate_id": "summary-project-irr",
            "_bucket": "derived_value_candidates",
            "source_reference": "P&L!B3",
            "source_validation_status": "validated",
            "submitted_role": "formula_derived_value",
            "validated_role": "formula_derived_value",
            "role_validation_status": "validated",
            "validation_status": "validated",
            "validated_value": 0.074,
            "validation_confidence": 0.99,
            "validation_warnings": [],
        }
    ]
    tools = WorkbookToolset(file_bytes=persistence_workbook_bytes())

    outputs = _canonicalize_outputs(new_uuid(), snapshot, tools)

    assert len(outputs) == 1
    assert outputs[0]["business_role"] == "project_irr"
    assert outputs[0]["label"] == "Project IRR"
    assert outputs[0]["source_sheet"] == "P&L"
    assert outputs[0]["source_cell"] == "B3"
    assert "BEST_MATCH_SCALAR_PROMOTED" in outputs[0][
        "validation_warnings_json"
    ]


def test_multiple_scalar_candidates_create_best_match_extracted_binding(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context
    workbook = load_workbook(BytesIO(persistence_workbook_bytes()))
    returns = workbook.create_sheet("Returns")
    returns["A7"] = "Project IRR"
    returns["B7"] = "=0.074"
    returns["A8"] = "IRR"
    returns["B8"] = "=0.073"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    result = deterministic_extraction_result()
    final_extraction = result["final_extraction"]
    final_extraction["output_candidates"] = []
    final_extraction["derived_value_candidates"] = [
        {
            "candidate_id": "summary-project-irr",
            "original_label": "Project IRR",
            "submitted_role": "formula_derived_value",
            "business_role": "project_irr",
            "raw_value": None,
            "unit": "%",
            "source_references": [{"sheet_name": "Returns", "cell": "B7"}],
            "reasoning_summary": "Explicit summary return",
            "llm_confidence": 0.99,
        }
    ]
    final_extraction["review_candidates"] = [
        {
            "candidate_id": "calculation-project-irr",
            "original_label": "IRR",
            "submitted_role": "formula_output",
            "business_role": "project_irr",
            "raw_value": None,
            "unit": "%",
            "source_references": [{"sheet_name": "Returns", "cell": "B8"}],
            "reasoning_summary": "Calculation return",
            "llm_confidence": 0.99,
        }
    ]
    result["validation_results"] = [
        {
            "candidate_id": "summary-project-irr",
            "_bucket": "derived_value_candidates",
            "source_reference": "Returns!B7",
            "source_validation_status": "validated_null",
            "submitted_role": "formula_derived_value",
            "validated_role": "formula_derived_value",
            "role_validation_status": "validated",
            "validation_status": "validated_null",
            "validated_value": None,
            "validation_confidence": 0.95,
            "validation_warnings": ["formula_cache_missing"],
        },
        {
            "candidate_id": "calculation-project-irr",
            "_bucket": "review_candidates",
            "source_reference": "Returns!B8",
            "source_validation_status": "validated_null",
            "submitted_role": "formula_output",
            "validated_role": "formula_output",
            "role_validation_status": "validated",
            "validation_status": "validated_null",
            "validated_value": None,
            "validation_confidence": 0.95,
            "validation_warnings": ["formula_cache_missing"],
        },
    ]
    service = ModelExtractionPersistenceService(
        session,
        validation_runner=RecordingRunner(result),
    )

    response = service.process_upload(buffer.getvalue(), "returns.xlsx")

    binding = session.scalar(
        select(ModelSemanticBinding).where(
            ModelSemanticBinding.model_version_id == response["model_version_id"],
            ModelSemanticBinding.semantic_role == "project_irr",
        )
    )
    assert binding is not None
    assert binding.binding_source == "extracted"
    assert binding.canonical_output is not None
    assert binding.canonical_output.source_sheet == "Returns"
    assert binding.canonical_output.source_cell == "B7"
    assert binding.evidence_json["selected_score"] > binding.evidence_json[
        "alternatives"
    ][0]["score"]


def test_t1_commits_workbook_and_extracting_model_before_runner(lifecycle_context) -> None:
    _engine, session_factory, session = lifecycle_context

    def observe_t1(_file_bytes, _filename) -> None:
        observer = session_factory()
        try:
            assert _count(observer, WorkbookVersion) == 1
            model_version = observer.scalar(select(ModelVersion))
            assert model_version.status == "extracting"
            assert model_version.validation_status == "not_run"
        finally:
            observer.close()

    runner = RecordingRunner(callback=observe_t1)
    service = ModelExtractionPersistenceService(session, validation_runner=runner)

    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")

    assert response["model_version_id"]
    assert response["workbook_version_id"]
    assert runner.calls == 1


def test_invalid_workbook_creates_no_rows(lifecycle_context) -> None:
    _engine, _session_factory, session = lifecycle_context
    runner = RecordingRunner()
    service = ModelExtractionPersistenceService(session, validation_runner=runner)

    with pytest.raises(InvalidWorkbookError):
        service.process_upload(b"not-an-xlsx", "broken.xlsx")

    assert _count(session, WorkbookVersion) == 0
    assert _count(session, ModelVersion) == 0
    assert runner.calls == 0


def test_oversized_workbook_is_rejected_before_t1_and_runner(lifecycle_context) -> None:
    _engine, _session_factory, session = lifecycle_context
    runner = RecordingRunner()
    content = persistence_workbook_bytes()
    service = ModelExtractionPersistenceService(
        session,
        validation_runner=runner,
        max_workbook_bytes=len(content) - 1,
    )

    with pytest.raises(WorkbookTooLargeError):
        service.process_upload(content, "model.xlsx")

    assert _count(session, WorkbookVersion) == 0
    assert _count(session, ModelVersion) == 0
    assert runner.calls == 0


def test_runner_exception_marks_extraction_failed_and_preserves_exception_type(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context

    class SentinelRunnerError(RuntimeError):
        pass

    def fail_runner(_file_bytes, _filename):
        raise SentinelRunnerError("runner failed")

    service = ModelExtractionPersistenceService(session, validation_runner=fail_runner)

    with pytest.raises(SentinelRunnerError, match="runner failed"):
        service.process_upload(persistence_workbook_bytes(), "model.xlsx")

    model_version = session.scalar(select(ModelVersion))
    assert model_version.status == "extraction_failed"
    assert model_version.error_code == "EXTRACTION_ERROR"
    assert _count(session, WorkbookVersion) == 1


def test_submitted_false_marks_extraction_failed_and_returns_null_ids(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context
    runner = RecordingRunner(deterministic_extraction_result(submitted=False))
    service = ModelExtractionPersistenceService(session, validation_runner=runner)

    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")

    assert response["submitted"] is False
    assert response["workbook_version_id"] is None
    assert response["model_version_id"] is None
    model_version = session.scalar(select(ModelVersion))
    assert model_version.status == "extraction_failed"
    assert model_version.stop_reason == "model_returned_no_tool_call"


def test_t2_snapshot_commits_before_t3(lifecycle_context) -> None:
    _engine, session_factory, session = lifecycle_context
    runner = RecordingRunner()
    repository = FailOnceAfterCanonicalRepository(session)
    service = ModelExtractionPersistenceService(
        session,
        validation_runner=runner,
        repository=repository,
    )

    with pytest.raises(ModelExtractionPersistenceError):
        service.process_upload(persistence_workbook_bytes(), "model.xlsx")

    observer = session_factory()
    try:
        model_version = observer.scalar(select(ModelVersion))
        assert model_version.status == "persistence_failed"
        assert model_version.extraction_snapshot_json["final_extraction"]
        assert model_version.extracted_at is not None
    finally:
        observer.close()


def test_t2_transient_failure_retries_without_rerunning_extraction(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context
    runner = RecordingRunner()
    repository = FailOnceSnapshotRepository(session)
    service = ModelExtractionPersistenceService(
        session,
        validation_runner=runner,
        repository=repository,
    )

    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")

    assert response["model_version_id"]
    assert repository.snapshot_attempts == 2
    assert runner.calls == 1
    assert session.get(ModelVersion, response["model_version_id"]).status == "materialized"


def test_success_persists_parameter_series_values_and_returns_ids(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context
    service = ModelExtractionPersistenceService(session, validation_runner=RecordingRunner())

    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")

    assert response["workbook_version_id"]
    assert response["model_version_id"]
    assert _count(session, ModelParameter) == 2
    assert _count(session, CanonicalOutput) == 1
    assert _count(session, FinancialSeries) == 1
    assert _count(session, FinancialSeriesValue) == 2
    model_version = session.get(ModelVersion, response["model_version_id"])
    assert model_version.status == "materialized"


def test_snapshot_strips_dependency_evidence(lifecycle_context) -> None:
    _engine, _session_factory, session = lifecycle_context
    service = ModelExtractionPersistenceService(session, validation_runner=RecordingRunner())

    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")
    model_version = session.get(ModelVersion, response["model_version_id"])

    assert not _contains_key(model_version.extraction_snapshot_json, "dependency_evidence")
    assert not _contains_key(model_version.validation_results_json, "dependency_evidence")


def test_metadata_remains_snapshot_only_and_output_is_materialized(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context
    service = ModelExtractionPersistenceService(session, validation_runner=RecordingRunner())
    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")
    read_service = ModelExtractionReadService(session, DatabaseWorkbookStorage(session))

    parameters = read_service.list_parameters(response["model_version_id"])

    assert {parameter.label for parameter in parameters} == {
        "Tax rate",
        "Discounted value",
    }
    output = session.scalar(select(CanonicalOutput))
    assert output.label == "Displayed revenue"
    assert output.business_role == "revenue"
    assert output.source_sheet == "P&L"
    assert output.source_cell == "B3"
    model_version = session.get(ModelVersion, response["model_version_id"])
    final_extraction = model_version.extraction_snapshot_json["final_extraction"]
    assert final_extraction["metadata"][0]["original_label"] == "Project Alpha"
    assert final_extraction["output_candidates"][0]["original_label"] == "Displayed revenue"


@pytest.mark.parametrize("source_references", [None, []])
def test_source_rejected_review_candidate_is_not_canonicalized(
    lifecycle_context,
    source_references,
) -> None:
    _engine, _session_factory, session = lifecycle_context
    result = deterministic_extraction_result()
    result["final_extraction"]["review_candidates"].append({
        "candidate_id": "source-less",
        "original_label": "Unbound candidate",
        "submitted_role": "hardcoded_input",
        "raw_value": 123,
        "source_references": source_references,
        "source_contract_bucket": "all_assumption_candidates",
    })
    result["validation_results"].append({
        "candidate_id": "source-less",
        "original_label": "Unbound candidate",
        "submitted_value": 123,
        "source_reference": None,
        "source_validation_status": "rejected",
        "submitted_role": "hardcoded_input",
        "validated_role": None,
        "role_validation_status": "not_evaluated",
        "validation_status": "rejected",
        "validated_value": None,
        "invalid_source": True,
        "rejection_reason": "no_source",
        "review_required": True,
        "_bucket": "review_candidates",
    })
    service = ModelExtractionPersistenceService(
        session,
        validation_runner=RecordingRunner(result),
    )

    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")

    assert _count(session, ModelParameter) == 2
    assert _count(session, CanonicalOutput) == 1
    model_version = session.get(ModelVersion, response["model_version_id"])
    review = model_version.extraction_snapshot_json["final_extraction"][
        "review_candidates"
    ]
    source_less = next(
        candidate for candidate in review
        if candidate["candidate_id"] == "source-less"
    )
    assert source_less["source_references"] == source_references
    assert any(
        validation.get("candidate_id") == "source-less"
        and validation.get("validation_status") == "rejected"
        for validation in model_version.validation_results_json
    )


def test_range_rejected_series_is_audited_but_not_canonicalized(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context
    result = deterministic_extraction_result()
    result["final_extraction"]["review_candidates"].append({
        "candidate_id": "range-rejected-series",
        "original_label": "Invalid series",
        "submitted_role": "financial_series",
        "raw_value": None,
        "source_references": [],
        "source_contract_bucket": "financial_series",
        "reconciliation_rejection_reason": "series_range_invalid",
        "period_range": "",
        "value_range": "P&L!B2:C2",
    })
    result["validation_results"].append({
        "candidate_id": "range-rejected-series",
        "original_label": "Invalid series",
        "source_validation_status": "rejected",
        "submitted_role": "financial_series",
        "validated_role": None,
        "validation_status": "rejected",
        "invalid_source": False,
        "rejection_reason": "series_range_invalid",
        "review_required": True,
        "_bucket": "review_candidates",
    })
    service = ModelExtractionPersistenceService(
        session,
        validation_runner=RecordingRunner(result),
    )

    response = service.process_upload(
        persistence_workbook_bytes(),
        "model.xlsx",
    )

    assert _count(session, FinancialSeries) == 1
    assert _count(session, FinancialSeriesValue) == 2
    assert _count(session, ModelParameter) == 2
    assert _count(session, CanonicalOutput) == 1
    model_version = session.get(ModelVersion, response["model_version_id"])
    assert any(
        item.get("candidate_id") == "range-rejected-series"
        and item.get("reconciliation_rejection_reason")
        == "series_range_invalid"
        for item in model_version.extraction_snapshot_json[
            "final_extraction"
        ]["review_candidates"]
    )
    assert any(
        item.get("candidate_id") == "range-rejected-series"
        and item.get("rejection_reason") == "series_range_invalid"
        and item.get("validation_status") == "rejected"
        for item in model_version.validation_results_json
    )


def test_formula_derived_parameter_reloads_exact_formula_and_null_cache(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context
    service = ModelExtractionPersistenceService(session, validation_runner=RecordingRunner())
    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")
    read_service = ModelExtractionReadService(session, DatabaseWorkbookStorage(session))

    derived = next(
        parameter
        for parameter in read_service.list_parameters(response["model_version_id"])
        if parameter.label == "Discounted value"
    )

    assert derived.validated_role == "formula_derived_value"
    assert derived.validated_value_json is None
    assert derived.exact_formula == "=Assumptions!B4*100"
    assert derived.formula_status == "formula_no_cache"


def test_financial_series_reloads_explicit_business_role(lifecycle_context) -> None:
    _engine, _session_factory, session = lifecycle_context
    service = ModelExtractionPersistenceService(session, validation_runner=RecordingRunner())
    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")
    read_service = ModelExtractionReadService(session, DatabaseWorkbookStorage(session))

    series = read_service.list_financial_series(response["model_version_id"])

    assert len(series) == 1
    assert series[0].business_role == "revenue"


def test_descriptor_materialization_persists_recovery_audit_and_business_role(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context

    def descriptor_runner(file_bytes: bytes, _filename: str) -> dict:
        recovery_resolution = {
            "field": "period_range",
            "submitted": "2026-2027",
            "resolved": "P&L!B2:C2",
            "strategy": "unique_integer_span_match",
            "partition_id": "partition-pnl",
        }
        extraction = {
            "financial_series": [
                {
                    "series_id": "recovered-cfads",
                    "label": "CFADS",
                    "semantic_role": "financial_series",
                    "business_role": "cfads",
                    "category": "cash_flow",
                    "unit": "USDm",
                    "frequency": "annual",
                    "period_range": "P&L!B2:C2",
                    "value_range": "P&L!B3:C3",
                    "label_reference": "P&L!A3",
                    "reasoning_summary": "Recovered annual CFADS row",
                    "llm_confidence": 0.93,
                    "period_axis": {"periods": ["wrong", "wrong"]},
                    "value_axis": {"values": [999, 999]},
                    "_backend_range_resolutions": [recovery_resolution],
                }
            ],
            "range_resolutions": [recovery_resolution],
        }
        materialized = materialize_financial_series(
            WorkbookToolset(file_bytes=file_bytes),
            extraction,
            trust_backend_range_resolutions=True,
        )
        result = deterministic_extraction_result()
        result["final_extraction"]["financial_series"] = extraction["financial_series"]
        result["final_extraction"]["financial_series_descriptors"] = extraction[
            "financial_series_descriptors"
        ]
        result["final_extraction"]["range_resolutions"] = extraction["range_resolutions"]
        result["validation_results"] = [
            item
            for item in result["validation_results"]
            if item.get("_bucket") != "financial_series"
        ] + materialized["validation_results"]
        result["time_series_summary"] = materialized["summary"]
        return result

    service = ModelExtractionPersistenceService(session, validation_runner=descriptor_runner)
    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")
    read_service = ModelExtractionReadService(session, DatabaseWorkbookStorage(session))

    series = read_service.list_financial_series(response["model_version_id"])
    values = session.scalars(
        select(FinancialSeriesValue)
        .where(FinancialSeriesValue.financial_series_id == series[0].id)
        .order_by(FinancialSeriesValue.period_index)
    ).all()

    assert len(series) == 1
    assert series[0].business_role == "cfads"
    assert series[0].period_source_range == "'P&L'!B2:C2"
    assert series[0].value_source_range == "'P&L'!B3:C3"
    assert "PERIOD_RANGE_RESOLVED_FROM_WORKBOOK_EVIDENCE" in series[0].warnings_json
    assert [value.raw_period_label_json for value in values] == [2026, 2027]
    assert [value.value_json for value in values] == [100.0, None]


def test_same_source_assumption_family_candidates_use_bucket_priority(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context
    result = deterministic_extraction_result()
    parameter = result["final_extraction"]["parameter_candidates"][0]
    parameter["submitted_role"] = "parameter"
    validation = result["validation_results"][0]
    validation["submitted_role"] = "parameter"
    validation["validated_role"] = "parameter"
    validation["role_validation_status"] = "validated_deferred"

    assumption = deepcopy(parameter)
    assumption["candidate_id"] = "tax-rate-assumption"
    assumption["original_label"] = "Tax rate assumption"
    assumption["submitted_role"] = "hardcoded_input"
    result["final_extraction"]["all_assumption_candidates"] = [assumption]
    result["validation_results"].append(
        {
            **deepcopy(validation),
            "candidate_id": assumption["candidate_id"],
            "_bucket": "all_assumption_candidates",
            "submitted_role": "hardcoded_input",
            "validated_role": "hardcoded_input",
        }
    )

    service = ModelExtractionPersistenceService(
        session,
        validation_runner=RecordingRunner(result),
    )

    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")

    parameters = session.scalars(
        select(ModelParameter).where(
            ModelParameter.model_version_id == response["model_version_id"]
        )
    ).all()
    assert len(parameters) == 2
    tax_rate = next(
        item
        for item in parameters
        if item.source_sheet == "Assumptions" and item.source_cell == "B4"
    )
    assert tax_rate.source_bucket == "parameter_candidates"
    assert tax_rate.llm_candidate_alias == "tax-rate"
    assert tax_rate.validated_role == "parameter"


def test_t3_failure_rolls_back_children_and_marks_persistence_failed(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context
    runner = RecordingRunner(
        deterministic_extraction_result(conflicting_parameter=True)
    )
    service = ModelExtractionPersistenceService(session, validation_runner=runner)

    with pytest.raises(CanonicalSourceConflictError):
        service.process_upload(persistence_workbook_bytes(), "model.xlsx")

    assert _count(session, ModelParameter) == 0
    assert _count(session, FinancialSeries) == 0
    assert _count(session, FinancialSeriesValue) == 0
    assert _count(session, ModelSemanticBinding) == 0
    assert session.scalar(select(ModelVersion.status)) == "persistence_failed"


def test_persistence_retry_reuses_snapshot_ids_without_runner_call(
    lifecycle_context,
) -> None:
    _engine, _session_factory, session = lifecycle_context
    runner = RecordingRunner()
    repository = FailOnceAfterCanonicalRepository(session)
    service = ModelExtractionPersistenceService(
        session,
        validation_runner=runner,
        repository=repository,
    )
    with pytest.raises(ModelExtractionPersistenceError):
        service.process_upload(persistence_workbook_bytes(), "model.xlsx")
    model_version = session.scalar(select(ModelVersion))
    expected_parameter_id = FinancialEntityIdFactory(model_version.id).parameter_id(
        "Assumptions",
        "B4",
    )

    model_id, workbook_id = service.retry_canonical_persistence(model_version.id)

    assert model_id == model_version.id
    assert workbook_id == model_version.workbook_version_id
    assert runner.calls == 1
    assert session.get(ModelParameter, expected_parameter_id) is not None
    assert session.get(ModelVersion, model_id).status == "materialized"


def test_same_bytes_new_upload_reuses_workbook_and_creates_model(lifecycle_context) -> None:
    _engine, _session_factory, session = lifecycle_context
    runner = RecordingRunner()
    service = ModelExtractionPersistenceService(session, validation_runner=runner)
    content = persistence_workbook_bytes()

    first = service.process_upload(content, "first.xlsx")
    second = service.process_upload(content, "renamed.xlsx")

    assert first["workbook_version_id"] == second["workbook_version_id"]
    assert first["model_version_id"] != second["model_version_id"]
    assert _count(session, WorkbookVersion) == 1
    assert _count(session, ModelVersion) == 2


def test_already_materialized_retry_is_idempotent(lifecycle_context) -> None:
    _engine, _session_factory, session = lifecycle_context
    runner = RecordingRunner()
    service = ModelExtractionPersistenceService(session, validation_runner=runner)
    response = service.process_upload(persistence_workbook_bytes(), "model.xlsx")

    model_id, workbook_id = service.retry_canonical_persistence(
        response["model_version_id"]
    )

    assert (model_id, workbook_id) == (
        response["model_version_id"],
        response["workbook_version_id"],
    )
    assert runner.calls == 1
    assert _count(session, ModelParameter) == 2


@pytest.mark.postgres
def test_postgres_t3_failure_rolls_back_every_canonical_child() -> None:
    from alembic import command
    from alembic.config import Config

    database_url = os.getenv("TEST_POSTGRES_URL")
    if not database_url:
        pytest.skip("TEST_POSTGRES_URL is required for PostgreSQL acceptance tests")
    database_name = make_url(database_url).database or ""
    if "test" not in database_name.lower():
        pytest.fail("TEST_POSTGRES_URL must identify an isolated test database")

    engine = create_engine(database_url)
    try:
        with engine.begin() as connection:
            for table_name in (
                "calculation_run_values",
                "calculation_runs",
                "calculation_rule_dependencies",
                "calculation_rule_members",
                "grouped_calculation_rules",
                "calculation_graph_components",
                "calculation_graph_versions",
                "workbook_named_expressions",
                "formula_execution_results",
                "formula_canonical_mappings",
                "formula_references",
                "executable_formula_rules",
                "workbook_formula_cells",
                "calculation_rule_extractions",
                "canonical_outputs",
                "financial_series_values",
                "financial_series",
                "model_parameters",
                "model_versions",
                "workbook_versions",
                "alembic_version",
            ):
                connection.exec_driver_sql(
                    f'DROP TABLE IF EXISTS "{table_name}" CASCADE'
                )
        config_path = Path(__file__).parents[1] / "apps" / "api" / "alembic.ini"
        config = Config(str(config_path))
        config.set_main_option("sqlalchemy.url", database_url)
        command.upgrade(config, "head")

        session_factory = sessionmaker(
            bind=engine,
            autoflush=False,
            expire_on_commit=False,
        )
        session = session_factory()
        try:
            runner = RecordingRunner()
            repository = FailOnceAfterCanonicalRepository(session)
            service = ModelExtractionPersistenceService(
                session,
                validation_runner=runner,
                repository=repository,
            )

            with pytest.raises(ModelExtractionPersistenceError):
                service.process_upload(persistence_workbook_bytes(), "model.xlsx")

            assert _count(session, ModelParameter) == 0
            assert _count(session, FinancialSeries) == 0
            assert _count(session, FinancialSeriesValue) == 0
            assert session.scalar(select(ModelVersion.status)) == "persistence_failed"
        finally:
            session.close()
    finally:
        engine.dispose()
